"""
gen_html.py  —  CUL Daily Movement HTML Generator
每次运行会从 Excel 提取最新数据，将 TODAY_DATA 和 FULL_SCHEDULE_DATA 注入到 HTML 中，
保留历史快照机制，输出 cul_daily_movement.html（含 Summary + 完整船期 两个视图）。

新增功能 (2026-07-01):
  - 解析全部列（PORT, man in, wait, Proforma, ltm eta/etd, Voy, date, ETA, ETB, ETD,
    run, Port Stay, fsp distance, speed, ETA Delay, ETD Delay）
  - 两个视图均支持"显示列"下拉选择器（可勾选显示/隐藏列）
  - Full Schedule 默认展示 Proforma 列

用法:
    python gen_html.py
    python gen_html.py --excel "P:/path/to/CUL DAILY MOVEMENT.xlsx"
    python gen_html.py --out "P:/path/to/output/cul_daily_movement.html"
"""

import openpyxl, json, re, sys, os, argparse, glob
from datetime import datetime, date, timedelta

# ── Defaults ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# 数据更新目录: 本机(leahliu)=P:, 另一台(culadmin)=Z:。自动探测, 两机通用, 无需传参。
_BASES = [
    r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新",
    r"P:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新",
]
UPD_DIR = next((b for b in _BASES if os.path.isdir(b)), _BASES[0])
from build_fleet_movement import GEN_DIR
DEFAULT_EXCEL = os.path.join(GEN_DIR, "CUL DAILY MOVEMENT.rebuilt.xlsx")
DEFAULT_HTML  = os.path.join(SCRIPT_DIR, "cul_daily_movement.html")

# ── Column definitions ─────────────────────────────────────────────────────
# Full column list (all columns in the Excel schedule rows)
# Used for Full Schedule view; Summary view uses a subset.
FULL_COLUMNS = [
    ('route',       'Route',            True,   True),
    ('vessel',      'Vessel',           True,   True),
    ('code',        'Code',             True,   True),
    ('port',        'Port',             True,   True),
    ('manIn',       'Man In',           True,   False),  # C2 - default hidden
    ('wait',        'Wait',             True,   True),   # C3
    ('proforma',    'Proforma',         False,  True),   # C4
    ('ltmEta',      'LTM ETA / LTS ETB',False,  False),
    ('ltmEtd',      'LTM ETD / LTS ETD',False,  False),
    ('voy',         'Voy. No',          True,   True),   # C7
    ('date',        'Date',             True,   True),   # C8
    ('eta',         'ETA',              True,   True),   # C9
    ('etb',         'ETB',              True,   True),   # C10
    ('etd',         'ETD',              True,   True),   # C11
    ('run',         'Run',              True,   True),   # C12
    ('portStay',    'Port Stay(hr)',    True,   True),   # C13
    ('fspDistance', 'FSP Distance',     True,   False),  # C14 - default hidden
    ('speed',       'Speed',            True,   True),   # C15
    ('etaDelay',    'ETA Delay',        True,   True),   # C16
    ('etdDelay',    'ETD Delay',        True,   False),  # C17
    ('remark',      'Remark',           True,   False),
    ('pic',         'PIC',              True,   True),
]
# Summary view columns (subset)
SUMMARY_COLUMNS = [
    ('_idx',       '#',                True),   # Row number, not sortable
    ('route',      'Route',            True),
    ('vessel',     'Vessel',           True),
    ('code',       'Code',             True),
    ('port',       'Port',             True),
    ('wait',       'Wait',             True),
    ('voy',        'Voy. No',          True),
    ('eta',        'ETA',              True),
    ('etb',        'ETB',              True),
    ('etd',        'ETD',              True),
    ('portStay',   'Port Stay(hr)',    True),
    ('etaDelay',   'ETA Delay',        True),
    ('etdDelay',   'ETD Delay',        True),
    ('remark',     'Remark',           False),
    ('pic',        'PIC',              True),
]

# ── Helpers ────────────────────────────────────────────────────────────────
def fmt_dt(v):
    if v is None: return ''
    if isinstance(v, datetime):
        return v.strftime('%m/%d %H:%M')   # ETA/ETB/ETD 始终显示时间(含00:00)
    return str(v).strip()

def get_str(v):
    if v is None: return ''
    if isinstance(v, datetime): return v.strftime('%m/%d %H:%M')
    return str(v).strip()

# ── Extract data ──────────────────────────────────────────────────────────
def extract(excel_path):
    today = date.today()
    data_eta_min = None   # capture ALL rows from Excel (no date window)
    data_eta_max = None
    wb_src = openpyxl.load_workbook(excel_path)
    ws_src = wb_src.active

    vessel_blocks = []
    rows_total = ws_src.max_row
    i = 1
    while i <= rows_total:
        c16 = ws_src.cell(i, 16).value
        if c16 and isinstance(c16, str) and 'PIC' in c16:
            route       = VESSEL_ROUTE_OVERRIDE.get(get_str(ws_src.cell(i, 1).value), get_str(ws_src.cell(i, 1).value))
            vessel_full = get_str(ws_src.cell(i, 4).value)
            vessel_code = get_str(ws_src.cell(i, 9).value)
            pic = c16.replace('PIC:', '').replace('PIC :', '').strip()

            schedule_rows = []
            remark = ''
            remarks_by_row = {}  # row_number -> remark for per-row association
            j = i + 2
            while j <= rows_total:
                c1_j = ws_src.cell(j, 1).value
                if c1_j and isinstance(c1_j, str) and c1_j.strip().startswith('Remark'):
                    remark_text = c1_j.strip().replace('Remark:', '').replace('Remark :', '').strip()
                    if remark_text:
                        # Parse remark to find target row: first two tokens = voyage + terminal/port
                        parts = remark_text.split()
                        if len(parts) >= 2 and schedule_rows:
                            target_voy, target_port = parts[0], parts[1]
                            # Strip the leading voyage + terminal/port tokens; keep only the remark content
                            remark_body = ' '.join(parts[2:])
                            matched = False
                            for sr in schedule_rows:
                                sr_voy = get_str(ws_src.cell(sr, 7).value)
                                sr_port = get_str(ws_src.cell(sr, 1).value)
                                if sr_voy == target_voy and sr_port == target_port:
                                    if sr in remarks_by_row:
                                        remarks_by_row[sr] = remarks_by_row[sr] + '; ' + remark_body
                                    else:
                                        remarks_by_row[sr] = remark_body
                                    matched = True
                                    break
                            if not matched:
                                # Fallback: assign to last schedule row
                                remarks_by_row[schedule_rows[-1]] = remark_body
                        elif schedule_rows:
                            remarks_by_row[schedule_rows[-1]] = remark_text
                        else:
                            remark = remark_text  # fallback: no schedule rows yet
                    j += 1
                    continue
                c16_j = ws_src.cell(j, 16).value
                if c16_j and isinstance(c16_j, str) and 'PIC' in c16_j:
                    i = j
                    break
                eta_val = ws_src.cell(j, 9).value
                if isinstance(eta_val, datetime):
                    eta_d = eta_val.date()
                    if data_eta_min is None or eta_d < data_eta_min:
                        data_eta_min = eta_d
                    if data_eta_max is None or eta_d > data_eta_max:
                        data_eta_max = eta_d
                    schedule_rows.append(j)
                j += 1
            else:
                i = rows_total + 1

            vessel_blocks.append({'route': route, 'vessel_full': vessel_full,
                                   'vessel_code': vessel_code, 'pic': pic,
                                   'schedule_rows': schedule_rows,
                                   'remarks_by_row': remarks_by_row})
        else:
            i += 1

    # ── Summary: nearest ETB per vessel (dedup by vessel across routes) ──
    # Rule: prefer ETB >= today (next upcoming call); among futures pick earliest;
    #       if no future ETB, pick the most recent past ETB (closest to today).
    vessel_rows = {}  # vessel_full -> list of (etb_d, row, vb)
    for vb in vessel_blocks:
        for r in vb['schedule_rows']:
            etb_v = ws_src.cell(r, 10).value
            if isinstance(etb_v, datetime):
                vessel_rows.setdefault(vb['vessel_full'], []).append((etb_v.date(), r, vb))

    vessel_best = {}
    summary_row_set = set()
    for vname, entries in vessel_rows.items():
        def _key(e):
            etb_d = e[0]
            if etb_d >= today:
                return (0, etb_d)          # future: always preferred, earlier = better
            return (1, today - etb_d)      # past: (1, gap); smaller gap (more recent) = better
        best_etb, best_row, vb = sorted(entries, key=_key)[0]
        vessel_best[vname] = (best_etb, best_row, vb)

    results = []
    for vname, (best_etb, best_row, vb) in vessel_best.items():
        summary_row_set.add(best_row)
        r = best_row
        rec = {
            'route':       vb['route'],
            'vessel':      vb['vessel_full'],
            'code':        vb['vessel_code'],
            'pic':         vb['pic'],
            'port':        get_str(ws_src.cell(r, 1).value),
            'manIn':       get_str(ws_src.cell(r, 2).value),
            'wait':        get_str(ws_src.cell(r, 3).value),
            'proforma':    get_str(ws_src.cell(r, 4).value),
            'voy':         get_str(ws_src.cell(r, 7).value),
            'ltmEta':      fmt_dt(ws_src.cell(r, 5).value),
            'ltmEtd':      fmt_dt(ws_src.cell(r, 6).value),
            'date':        fmt_dt(ws_src.cell(r, 8).value),
            'eta':         fmt_dt(ws_src.cell(r, 9).value),
            'etaRaw':      ws_src.cell(r, 9).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 9).value, datetime) else '',
            'etbRaw':      ws_src.cell(r, 10).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 10).value, datetime) else '',
            'etdRaw':      ws_src.cell(r, 11).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 11).value, datetime) else '',
            'etb':         fmt_dt(ws_src.cell(r, 10).value),
            'etd':         fmt_dt(ws_src.cell(r, 11).value),
            'run':         get_str(ws_src.cell(r, 12).value),
            'portStay':    get_str(ws_src.cell(r, 13).value),
            'fspDistance': get_str(ws_src.cell(r, 14).value),
            'speed':       get_str(ws_src.cell(r, 15).value),
            'etaDelay':    get_str(ws_src.cell(r, 16).value),
            'etdDelay':    get_str(ws_src.cell(r, 17).value),
            'remark':      vb['remarks_by_row'].get(r, ''),
        }
        results.append(rec)

    # ── Full Schedule: ALL port rows for ALL vessels ──
    full_schedule = []
    for vb in vessel_blocks:
        for r in vb['schedule_rows']:
            full_schedule.append({
                'route':       vb['route'],
                'vessel':      vb['vessel_full'],
                'code':        vb['vessel_code'],
                'pic':         vb['pic'],
                'port':        get_str(ws_src.cell(r, 1).value),
                'manIn':       get_str(ws_src.cell(r, 2).value),
                'wait':        get_str(ws_src.cell(r, 3).value),
                'proforma':    get_str(ws_src.cell(r, 4).value),
                'voy':         get_str(ws_src.cell(r, 7).value),
                'ltmEta':      fmt_dt(ws_src.cell(r, 5).value),
                'ltmEtd':      fmt_dt(ws_src.cell(r, 6).value),
                'date':        fmt_dt(ws_src.cell(r, 8).value),
                'eta':         fmt_dt(ws_src.cell(r, 9).value),
                'etaRaw':      ws_src.cell(r, 9).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 9).value, datetime) else '',
                'etbRaw':      ws_src.cell(r, 10).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 10).value, datetime) else '',
                'etdRaw':      ws_src.cell(r, 11).value.strftime('%Y-%m-%d') if isinstance(ws_src.cell(r, 11).value, datetime) else '',
                'etb':         fmt_dt(ws_src.cell(r, 10).value),
                'etd':         fmt_dt(ws_src.cell(r, 11).value),
                'run':         get_str(ws_src.cell(r, 12).value),
                'portStay':    get_str(ws_src.cell(r, 13).value),
                'fspDistance': get_str(ws_src.cell(r, 14).value),
                'speed':       get_str(ws_src.cell(r, 15).value),
                'etaDelay':    get_str(ws_src.cell(r, 16).value),
                'etdDelay':    get_str(ws_src.cell(r, 17).value),
                'remark':      vb['remarks_by_row'].get(r, ''),
                'isSummary':   r in summary_row_set,
            })

    return {
        'date': today.strftime('%Y-%m-%d'),
        'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'vessels': results,
        'fullSchedule': full_schedule,
        'defaultEtaFrom': (today - timedelta(days=14)).strftime('%Y-%m-%d'),
        'defaultEtaTo':   (today + timedelta(days=30)).strftime('%Y-%m-%d'),
        'defaultEtbFrom': (today - timedelta(days=14)).strftime('%Y-%m-%d'),
        'defaultEtbTo':   (today + timedelta(days=30)).strftime('%Y-%m-%d'),
        'dataEtaMin':     data_eta_min.strftime('%Y-%m-%d') if data_eta_min else '',
        'dataEtaMax':     data_eta_max.strftime('%Y-%m-%d') if data_eta_max else '',
    }

# ── BOA 映射 (Port & Lane Mapping + 补充映射) ─────────────────────────────
# BOA 数据行从 Daily Movement 船期实时聚合 (前端跟随 Port Wait 日期筛选)，
# Lane→Trade / Port→Region 归类从映射表找；映射表未覆盖的在此补充。
# 映射表源: P:\04 上海操作中心\01 船期管理科\船期管理\准班率BOA\2026\船期统计 2026 week 31.xlsx
BOA_SRC = r"P:\04 上海操作中心\01 船期管理科\船期管理\准班率BOA\2026\船期统计 2026 week 31.xlsx"

# 代理联系表（各港口代理/本地操作岗联系人），用于 Maintenance 未维护明细的 Agent/OP 列
# 仅抽取岗位含 operation / ops / vsl operation 者（船东代表处 Operations Representative、
# 本地代理 Operation Manager / Vsl Operation 等），排除 Equipment Control 等非操作岗。
AGENT_SRC = r"P:\04 上海操作中心\01 船期管理科\船期管理\CUL Agent Contact List 20260825 Thailand.xlsx"

# 维护率数据源：Vessel Schedule / Port Log 维护情况台账
# 列: B=Service C=Vessel D=Voyage E=Direction F=Operator G=Port H=ETD
#      I=Port Log Y/N (Y/N)  J=Vessel Schedule Maintain Status (Maintain timely / Not maintained)
MAINT_SRC = r"C:\CULINES\Claw Report\Vessel Schedule Maintain Over Time.xlsx"
# 其他机器(如 culadmin, 挂 Z 盘)无 C:\CULINES 本地路径，改为从 SFTP(10.5.4.2:6622) 拉取，
# 保证任意机器生成的看板都含 Maintenance 数据。路径可用环境变量覆盖。
MAINT_SFTP_HOST = os.environ.get("MAINT_SFTP_HOST", "10.5.4.2")
MAINT_SFTP_PORT = int(os.environ.get("MAINT_SFTP_PORT", "6622"))
MAINT_SFTP_USER = os.environ.get("MAINT_SFTP_USER", "leah")
MAINT_SFTP_PASS = os.environ.get("MAINT_SFTP_PASS", "Fine@B!")
# SFTP 上 MAINT xlsx 的路径（默认与 Bapfile 同目录 /finebi/Master Data - Leah/）
MAINT_SFTP_REMOTE = os.environ.get("MAINT_SFTP_REMOTE", "/finebi/Master Data - Leah/Vessel Schedule Maintain Over Time.xlsx")
# 下载到 .cache/ (已被 .gitignore 忽略)，避免污染仓库
MAINT_LOCAL_CACHE = os.path.join(SCRIPT_DIR, ".cache", "Vessel Schedule Maintain Over Time.xlsx")

# ── BOA 映射：完整兜底映射表 ─────────────────────────────────────────────
# 说明：以下 lane→trade 与 port→region 为「Port & Lane Mapping」表全量快照
#       (船期统计 2026 week 31, 读取时间 2026-08-24) + 代码补充映射。
#       运行时优先读 P 盘映射表覆盖；P 盘不可达时用本兜底，保证 0 Unknown。
# 更新方式：映射表有增删时，手动同步本字典（或运行 gen_html.py 打印 diff）。

# Lane → Trade（42 条来自映射表 + 8 条补充 = 50 条）
BOA_LANE_TRADE_FALLBACK = {
    # ── 来自映射表 Port & Lane Mapping (42) ──
    'AEM': 'MD', 'AEX': 'EU', 'AG2': 'ME', 'AGX': 'ME', 'CCT': 'TH',
    'CES': 'EU', 'CGX': 'ME', 'CHT': 'TH', 'CIS': 'IN', 'CP1': 'PH',
    'CPX': 'MN', 'CST': 'TH', 'CV3': 'VN', 'CVT': 'TH', 'CVX': 'VN',
    'CVX2': 'VN', 'CVX3': 'VN', 'HDT': 'TW', 'IMR': 'ME', 'ISS': 'IN',
    'JPS': 'ME', 'NP2': 'PH', 'NSCT1': 'TW', 'NSX': 'TH', 'RBC1': 'TH',
    'REX': 'ME', 'SCT': 'TH', 'SCT2': 'TH', 'SGX': 'ME', 'SHX': 'VN',
    'SJA': 'ME', 'SL1': 'TH', 'ST3': 'TW', 'STD': 'TW', 'STX': 'TH',
    'SV2': 'VN', 'SVG': 'VN', 'TP1': 'US', 'TPC': 'TP', 'TPN': 'TP',
    'TPX': 'TP', 'VGX': 'ME',
    # ── 补充映射（映射表未覆盖，2026-08-13 用户确认/数据推断）──
    'RTS':  'ME',  # 红海-中东 (SAJED/EGSOK/OMSOH)，船期统计中 RTS→ME
    'NAX':  'MD',  # 用户确认 NAX=NAF→MD；地中海-北非线 (TRALI/TRIST)，与 AEM(MD) 港口重叠
    'RES':  'ME',  # 船期统计中 RES→ME (JOAQJ/EGSOK/SAJED)
    'HLX':  'ME',  # 船期统计中 HLX→ME (THLCH/SGSIN/INNSA/PKKHI)
    'GTS':  'ME',  # KR TASMAN: SAJED/YEADE/EGSOK/OMSOH 中东红海；同区域映射航线 SGX/CGX→ME
    'CGS':  'ME',  # 靠 AEKLF 为主；映射表中靠 AEKLF 的 SGX/CGX 都归 ME
    'CST/SL1': 'TH',  # 组合航线 CST/SL1 → TH (映射表中 CST→TH、SL1→TH)
}

# 某些船期的 route 列会错误地放 vessel code（如 ZGCD），需纠正为实际航线
VESSEL_ROUTE_OVERRIDE = {
    'ZGCD': 'AEM',  # ZHONG GU CHENG DU 实际属 AEM 航线
}

# Port → Region（75 条来自映射表 + 19 条补充 = 94 条）
BOA_PORT_REGION_FALLBACK = {
    # ── 来自映射表 Port & Lane Mapping (75) ──
    'AEJEA': 'Intra Asia', 'BEANR': 'Europe', 'CNGCT': 'China Mainland',
    'CNHMN': 'China Mainland', 'CNHUA': 'China Mainland', 'CNHUI': 'China Mainland',
    'CNNAS': 'China Mainland', 'CNNGB': 'China Mainland', 'CNSHA': 'China Mainland',
    'CNSHH': 'China Mainland', 'CNSHK': 'China Mainland', 'CNSWA': 'China Mainland',
    'CNTAO': 'China Mainland', 'CNTNJ': 'China Mainland', 'CNWIT': 'China Mainland',
    'CNXGG': 'China Mainland', 'CNXMN': 'China Mainland', 'CNXNA': 'China Mainland',
    'CNXNG': 'China Mainland', 'CNYPN': 'China Mainland', 'CNYTN': 'China Mainland',
    'DEHAM': 'Europe', 'DJJIB': 'AF', 'EGALY': 'Intra Asia', 'EGSOK': 'Intra Asia',
    'EGSUE': 'Intra Asia', 'GBSOU': 'Europe', 'GBTIL': 'Europe', 'GRPIR': 'Europe',
    'HKHKG': 'China HK & TW', 'IDJKT': 'Intra Asia', 'ILASD': 'Europe',
    'ILHFA': 'Europe', 'INMUN': 'Intra Asia', 'INNSA': 'Intra Asia',
    'KHKOS': 'Intra Asia', 'KRPUS': 'Intra Asia', 'MYPKG': 'Intra Asia',
    'MYPKN': 'Intra Asia', 'MYPKW': 'Intra Asia', 'NLAMS': 'Europe',
    'NLRTM': 'Europe', 'OMSOH': 'Intra Asia', 'PHMNL': 'Intra Asia',
    'PHMNN': 'Intra Asia', 'PHSPS': 'Intra Asia', 'PKKHI': 'Intra Asia',
    'QAHMD': 'Intra Asia', 'SADMM': 'Intra Asia', 'SAJED': 'Intra Asia',
    'SDPZU': 'AF', 'SGSIN': 'Intra Asia', 'THBKK': 'Intra Asia',
    'THBKS': 'Intra Asia', 'THLCH': 'Intra Asia', 'THSCS': 'Intra Asia',
    'TRALI': 'Europe', 'TRGEB': 'Intra Asia', 'TRIST': 'Europe',
    'TRIZT': 'Europe', 'TRMER': 'Europe', 'TWKEL': 'China HK & TW',
    'TWKHH': 'China HK & TW', 'TWTPE': 'China HK & TW', 'TWTXG': 'China HK & TW',
    'USLAX': 'US', 'USLGB': 'US', 'USOAK': 'US', 'VNDAD': 'Intra Asia',
    'VNDAN': 'Intra Asia', 'VNHCM': 'Intra Asia', 'VNHPH': 'Intra Asia',
    'VNSGN': 'Intra Asia', 'VNVUT': 'Intra Asia', 'YEADE': 'Intra Asia',
    # ── 补充映射（映射表未覆盖，2026-08-13）──
    'DZALG': 'AF', 'AEKLF': 'Intra Asia', 'LYMRA': 'AF', 'LYBEN': 'AF',
    'EGSUZ': 'Intra Asia', 'SAGIZ': 'Intra Asia', 'MALTA': 'Europe', 'JOAQJ': 'Intra Asia',
    'EGSGA': 'Intra Asia', 'INKDL': 'Intra Asia', 'CNDCB': 'China Mainland', 'EGSAF': 'Intra Asia',
    'AOAQJ': 'AF', 'EGDAM': 'Intra Asia', 'TNRDS': 'AF', 'GRSKG': 'Europe',
    'THPAT': 'Intra Asia', 'THSSW': 'Intra Asia',  # 泰国港口，与映射表 THBKK/THSCS/THLCH/THBKS 一致
    'TUZLA': 'Europe',  # 土耳其伊斯坦布尔附近，与映射表 TRIST/TRIZT 一致
}

def load_boa_mappings():
    """读取映射表 Port & Lane Mapping 覆盖兜底映射；映射表缺失(P 盘不可达)时用兜底，不报错。
    返回 (lane_trade, port_region) dict。任何环境下都保证完整映射 → BOA 0 Unknown。
    """
    lane_trade  = dict(BOA_LANE_TRADE_FALLBACK)
    port_region = dict(BOA_PORT_REGION_FALLBACK)
    n_map_lane = n_map_port = 0
    try:
        wb = openpyxl.load_workbook(BOA_SRC, data_only=True, read_only=True)
        if 'Port & Lane Mapping' in wb.sheetnames:
            ws_map = wb['Port & Lane Mapping']
            for row in ws_map.iter_rows(min_row=2, max_col=6, values_only=True):
                port, region, _, lane, trade = row[0], row[1], row[3], row[4], row[5]
                if port and region:
                    p_norm = re.sub(r'\s*\(.*?\)', '', str(port).strip()).strip()
                    port_region[p_norm] = str(region).strip()
                    n_map_port += 1
                if lane and trade:
                    lane_trade[str(lane).strip()] = str(trade).strip()
                    n_map_lane += 1
            print(f'  [BOA] mapping source OK: {n_map_lane} lanes, {n_map_port} ports from sheet')
        else:
            print('  [BOA] sheet "Port & Lane Mapping" not found, using fallback')
    except Exception as e:
        print('  [BOA] mapping source unavailable, using FALLBACK mapping:', e)
    print(f'  [BOA] final mappings: {len(lane_trade)} lanes, {len(port_region)} ports')
    return lane_trade, port_region

# ── Maintenance (Vessel Schedule / Port Log 维护率) 数据源 ────────────────
def ensure_maint_source():
    """返回可用的 MAINT xlsx 本地路径。读取优先级：
    1) 本机 C:\\CULINES 副本（leahliu 机器，无需 VPN）；
    2) 从 SFTP(10.5.4.2:6622) 下载到 .cache/（culadmin 等无本地副本的机器，需 VPN）。
    两者皆不可达返回 None，load_maint_data 将输出空 records 而不崩溃。"""
    if os.path.exists(MAINT_SRC):
        mt = datetime.fromtimestamp(os.path.getmtime(MAINT_SRC)).strftime('%Y-%m-%d %H:%M')
        return MAINT_SRC, mt + ' (local copy)'
    try:
        import paramiko
        os.makedirs(os.path.dirname(MAINT_LOCAL_CACHE), exist_ok=True)
        t = paramiko.Transport((MAINT_SFTP_HOST, MAINT_SFTP_PORT))
        try:
            t.connect(username=MAINT_SFTP_USER, password=MAINT_SFTP_PASS)
            sftp = paramiko.SFTPClient.from_transport(t)
            st = sftp.stat(MAINT_SFTP_REMOTE)
            rmt_time = datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M')
            print(f"  [MAINT] SFTP found {MAINT_SFTP_REMOTE}  mtime={rmt_time}", flush=True)
            tmp = MAINT_LOCAL_CACHE + ".part"
            sftp.get(MAINT_SFTP_REMOTE, tmp)
            os.replace(tmp, MAINT_LOCAL_CACHE)
            print(f"  [MAINT] downloaded from SFTP -> {MAINT_LOCAL_CACHE} ({os.path.getsize(MAINT_LOCAL_CACHE):,} bytes)", flush=True)
        finally:
            try: sftp.close()
            except Exception: pass
            try: t.close()
            except Exception: pass
        return MAINT_LOCAL_CACHE, rmt_time + ' (SFTP)'
    except Exception as e:
        msg = str(e)
        print(f"  [MAINT] SFTP fetch failed: {msg}", file=sys.stderr, flush=True)
        if "10060" in msg or "timed out" in msg.lower() or "Unable to connect" in msg:
            print("  [MAINT] Hint: 10.5.4.2 is internal — connect VPN first.", file=sys.stderr, flush=True)
        return None, ''

# ── 维护率台账列头识别（新旧文件兼容）───────────────────────────────────
# 旧文件列: B=Service C=Vessel D=Voyage E=Direction F=Operator G=Port H=ETD
#           I=Port Log Y/N J=Vessel Schedule Maintain Status
# 新文件(SFTP 改名后)删除了 Vessel 列、整体左移，ETD 时间列落在原 Port 位置，
# 必须按表头名映射，不能依赖固定列位置。
MAINT_HEADER_TOKENS = [
    ('plog',     ('port log', 'portlog', 'plog'),  None),
    ('vsched',   ('vessel schedule', 'vsched'),    None),
    ('service',  ('service',),    None),
    ('vessel',   ('vessel',),     ('vessel schedule',)),
    ('voyage',   ('voyage',),     None),
    ('dir',      ('direction', 'dir'),  None),
    ('operator', ('operator', 'op'),   None),
    ('port',     ('port',),       ('port log',)),
    ('etd',      ('etd',),        None),
]

def _norm_hdr(h):
    """表头归一化：转小写、去非字母数字。"""
    return re.sub(r'[^a-z0-9]', '', str(h or '').strip().lower())

def _hdr_has(nh, *tokens):
    for t in tokens:
        nt = _norm_hdr(t)
        if nh == nt or nh.startswith(nt) or nt in nh:
            return True
    return False

def detect_maint_columns(header_row):
    """按表头名解析列索引(0-based)；表头不可识别返回 None。
    复合名('port log'/'vessel schedule')优先匹配，防止 'port'/'vessel' 误吞。"""
    if not header_row:
        return None
    cols = {}
    for key, tokens, guard in MAINT_HEADER_TOKENS:
        for ci, h in enumerate(header_row):
            nh = _norm_hdr(h)
            if not nh or not _hdr_has(nh, *tokens):
                continue
            if guard and _hdr_has(nh, *guard):
                continue
            cols[key] = ci
            break
    if 'service' not in cols or 'port' not in cols:
        return None
    return cols

# 旧固定布局（0-based）: service=1 vessel=2 voyage=3 dir=4 operator=5
#                        port=6 etd=7 plog=8 vsched=9
LEGACY_MAINT_COLS = {'service': 1, 'vessel': 2, 'voyage': 3, 'dir': 4,
                     'operator': 5, 'port': 6, 'etd': 7, 'plog': 8, 'vsched': 9}
PORT_CODE_RE = re.compile(r'^[A-Z]{4,5}[A-Z0-9]{0,2}$')

def load_maint_data():
    """读取维护率台账，返回 {date, generatedAt, source, records:[...], sourceMtime}。
    列映射优先级：1) 表头名识别(兼容列顺序变化/缺列)；2) port 列内容校验(港口码 vs 时间)；
    3) 旧固定布局回退。任何环境（含源文件不可达）都返回合法结构，生成脚本不崩溃。"""
    src_ret = ensure_maint_source()
    if isinstance(src_ret, tuple) and len(src_ret) >= 2:
        src, src_mtime = src_ret[0], src_ret[1]
    else:
        src, src_mtime = src_ret, ''
    out = {'date': '', 'generatedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
           'source': src or '(no source)', 'records': [], 'sourceMtime': src_mtime,
           'vschedAvailable': False}
    if not src or not os.path.exists(src):
        print('  [MAINT] no source available -> empty records')
        return out
    try:
        wb = openpyxl.load_workbook(src, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out['source'] = (src or '') + '  ·  UNAVAILABLE: empty workbook'
            print('  [MAINT] source unavailable: empty workbook')
            return out

        def _get(row, key):
            ci = cols.get(key)
            return row[ci] if ci is not None and ci < len(row) else None

        def _looks_like_port_col(r0):
            sample = [_get(r, 'port') for r in rows[r0:r0 + 300]]
            vals = [str(v).strip() for v in sample if v is not None and str(v).strip()]
            if not vals:
                return False
            hits = sum(1 for v in vals if PORT_CODE_RE.match(v))
            return hits / len(vals) >= 0.5

        # 1) 表头名识别（第 1 行可能是标题行，再试行 2）
        cols, data_start = None, 1
        for ri in (0, 1):
            if ri >= len(rows):
                break
            c = detect_maint_columns(rows[ri])
            if c:
                cols, data_start = c, ri + 1
                print('  [MAINT] columns mapped by header (row %d): %s' %
                      (ri + 1, ', '.join('%s=%s' % (k, v) for k, v in sorted(c.items()))))
                break
        # 2) 回退旧固定布局
        if cols is None:
            cols = dict(LEGACY_MAINT_COLS)
            data_start = 1
            print('  [MAINT] header not recognized, fallback to legacy fixed layout')
        # 3) port 列内容校验：超过一半非空值是港口码才认为正确，否则试旧布局
        if not _looks_like_port_col(data_start) and cols != LEGACY_MAINT_COLS:
            print('  [MAINT] port column validation failed (looks like non-port values), trying legacy layout')
            cols = dict(LEGACY_MAINT_COLS)
            data_start = 1

        recs = []
        etds = []
        vsched_seen = False
        for row in rows[data_start:]:
            svc = _get(row, 'service')
            port = _get(row, 'port')
            if (svc is None or str(svc).strip() == '') and (port is None or str(port).strip() == ''):
                continue  # 跳过空行
            etd_raw = _get(row, 'etd')
            if etd_raw is not None:
                etd = etd_raw.strftime('%Y-%m-%d') if hasattr(etd_raw, 'strftime') else str(etd_raw)[:10]
            else:
                etd = ''
            if etd:
                etds.append(etd)
            vsched_raw = _get(row, 'vsched')
            plog_raw = _get(row, 'plog')
            vsched_s = '' if vsched_raw is None else str(vsched_raw).strip().lower()
            if vsched_s:
                vsched_seen = True   # 上游仍导出 Vessel Schedule 列(存在非空值)
            recs.append({
                'service':  '' if svc is None else str(svc).strip(),
                'vessel':   '' if _get(row, 'vessel') is None else str(_get(row, 'vessel')).strip(),
                'voyage':   '' if _get(row, 'voyage') is None else str(_get(row, 'voyage')).strip(),
                'dir':      '' if _get(row, 'dir') is None else str(_get(row, 'dir')).strip(),
                'operator': '' if _get(row, 'operator') is None else str(_get(row, 'operator')).strip(),
                'port':     '' if port is None else str(port).strip(),
                'etd':      etd,
                'plog':     '' if plog_raw is None else str(plog_raw).strip().upper(),
                # Vessel Schedule Maintain Status: 仅 "Maintain timely" 视为已维护，其余(Not maintained/空)视为未维护
                'vsched':   1 if vsched_s == 'maintain timely' else 0,
            })
        out['records'] = recs
        out['vschedAvailable'] = vsched_seen
        out['date'] = (min(etds) + ' ~ ' + max(etds)) if etds else ''
        out['source'] = src + f"  ·  {len(recs)} rows  ·  ETD {out['date']}"
        print(f'  [MAINT] source OK: {len(recs)} rows from {os.path.basename(src)}')
    except Exception as e:
        out['source'] = (src or '') + '  ·  UNAVAILABLE: ' + str(e)
        print('  [MAINT] source unavailable:', e)
    return out

# ── 防回退保护：坏源(有记录却 vsched 全 0) 时恢复 last-good 数据 ──────────
def _maint_is_broken(maint):
    """坏源特征：记录数充足、且 Vessel Schedule 列【存在】却全 0。
    2026-08 起上游 SFTP 导出删除了 'Vessel Schedule' 列(新格式)——列缺失
    (vschedAvailable=False)属正常格式：照常采用最新数据，前端 vsched 部分显示 '—'，
    不算坏源。仅当列有值却没有任何一条 'Maintain timely'(疑似坏下载)才回退 last-good。"""
    recs = maint.get('records') or []
    if len(recs) < 50:
        return False
    if not maint.get('vschedAvailable', True):
        return False   # 新格式：Vessel Schedule 列已停供，数据本身有效
    vsched_ok = sum(1 for r in recs if r.get('vsched') == 1)
    return vsched_ok == 0

def _extract_maint_from_html(html_path):
    """从已生成的 HTML 抽取 const MAINT_DATA = {...}; 返回 dict 或 None。"""
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            txt = f.read()
    except Exception:
        return None
    m = re.search(r'const MAINT_DATA\s*=\s*(\{.*\})\s*;$', txt, re.M)
    if not m:
        return None
    try:
        return json.loads(m.group(1))
    except Exception:
        return None

MAINT_SNAPSHOT = os.path.join(SCRIPT_DIR, 'maint_snapshot.json')

def _save_maint_snapshot(maint):
    try:
        with open(MAINT_SNAPSHOT, 'w', encoding='utf-8') as f:
            json.dump(maint, f, ensure_ascii=False)
    except Exception as e:
        print('  [MAINT] snapshot save failed:', e, file=sys.stderr)

def load_agent_contacts(src=None):
    """从《CUL Agent Contact List》xlsx 抽取各港口「操作岗(operation/ops)」代理联系人，
    返回 {PORT_CODE: [{'name','email','tel','mobile'}, ...]}。
    港口码 = 分表名（与 MAINT_DATA.port 一致：EGSOK / SAJED / THLCH …）。
    无操作岗联系人的港口不进入结果（前端显示 —）。"""
    import re as _re, glob as _glob
    if src is None:
        src = AGENT_SRC
    # 文件不存在时回退到同目录最新版本
    if not os.path.exists(src):
        cand = sorted(_glob.glob(os.path.join(os.path.dirname(src) or '.',
                        'CUL Agent Contact List*.xlsx')))
        if cand:
            src = cand[-1]
    if not os.path.exists(src):
        print('  [AGENT] WARN: contact list not found (%s) — Agent/OP column will be empty.' % src,
              file=sys.stderr, flush=True)
        return {}
    try:
        wb = openpyxl.load_workbook(src, data_only=True)
    except Exception as e:
        print('  [AGENT] WARN: failed to open %s: %s' % (src, e), file=sys.stderr, flush=True)
        return {}
    OP_RE = _re.compile(r'(operation|ops|vsl operation)', _re.I)
    EMAIL_RE = _re.compile(r'@')
    def norm(s): return _re.sub(r'[^a-z0-9]', '', str(s or '').strip().lower())
    def find_header(row):
        cells = [str(v or '').strip() for v in row]
        nh = [norm(c) for c in cells]
        h = {'name':None,'email':None,'tel':None,'mobile':None,'pos':None}
        for i,c in enumerate(nh):
            if 'name' in c and h['name'] is None: h['name']=i
            if ('email' in c or 'e-mail' in c) and h['email'] is None: h['email']=i
            if ('telephone' in c or c=='tel' or '电话' in c) and h['tel'] is None: h['tel']=i
            if ('mobile' in c or '手机' in c) and h['mobile'] is None: h['mobile']=i
            if ('position' in c or 'designation' in c or '岗位' in c or '部门' in c or 'department' in c) and h['pos'] is None: h['pos']=i
        if h['name'] is not None and h['email'] is not None:
            return h
        return None
    result = {}
    for sheet in wb.sheetnames:
        if norm(sheet) == 'masterlist':
            continue
        ws = wb[sheet]
        port = sheet.strip().upper()
        contacts = []
        header = None
        for row in ws.iter_rows(values_only=True):
            h = find_header(row)
            if h:
                header = h
                continue
            if header is None:
                continue
            cells = [("" if v is None else (v.strftime('%Y-%m-%d') if hasattr(v,'strftime') else str(v)).strip()) for v in row]
            def get(k):
                i = header[k]
                return cells[i] if (i is not None and i < len(cells)) else ''
            pos_val = get('pos'); name = get('name'); email = get('email')
            tel = get('tel'); mobile = get('mobile')
            if not OP_RE.search(pos_val or ''):
                continue
            if not name or not _re.search(r'[A-Za-z\u4e00-\u9fff]', name):
                continue
            if name.lower().startswith('department') or norm(name) in ('ops','operation','operations'):
                continue
            if not (EMAIL_RE.search(email or '') or tel or mobile):
                continue
            em  = (email.split(';')[0].split(',')[0].strip()  if email  else '')
            em  = em if EMAIL_RE.search(em) else ''
            t1  = (tel.split(';')[0].split(',')[0].strip()     if tel    else '')
            m1  = (mobile.split(';')[0].split(',')[0].strip()  if mobile else '')
            contacts.append({'name':name.strip(), 'email':em, 'tel':t1, 'mobile':m1})
        if contacts:
            result[port] = contacts
    print('  [AGENT] loaded OP contacts for %d ports from %s' % (len(result), os.path.basename(src)))
    return result

def resolve_maint(out_path):
    """加载 MAINT；若检测到坏源(有记录却 vsched 全 0)，从「现有 HTML 已嵌入数据」或
    「仓库内置快照」恢复，避免发布错误的 0% 维护率。源正常时刷新内置快照供其他机器回退。

    注意：本保护对「运行本脚本」的机器生效。culadmin 机器若长期不 git pull 旧代码，
    仍需先修复 SFTP 源文件(见技能文档)才能根除——因为 culadmin 的坏数据来自 SFTP 下载，
    而非本脚本逻辑。"""
    maint = load_maint_data()
    if not _maint_is_broken(maint):
        _save_maint_snapshot(maint)   # 源正常 → 刷新快照
        return maint
    print('  [MAINT] WARNING: source looks broken (records present but vsched all 0). '
          'Recovering last-good MAINT_DATA to avoid shipping 0% rate...',
          file=sys.stderr, flush=True)
    # 1) 复用现有 HTML 内已嵌入的正确 MAINT_DATA（不依赖 git pull / VPN）
    if os.path.exists(out_path):
        rec = _extract_maint_from_html(out_path)
        if rec and not _maint_is_broken(rec):
            print('  [MAINT] recovered from existing HTML (kept last-good data).', flush=True)
            return rec
    # 2) 回退到内置快照
    if os.path.exists(MAINT_SNAPSHOT):
        try:
            with open(MAINT_SNAPSHOT, 'r', encoding='utf-8') as f:
                rec = json.load(f)
            if not _maint_is_broken(rec):
                print('  [MAINT] recovered from bundled snapshot.', flush=True)
                return rec
        except Exception:
            pass
    print('  [MAINT] no recoverable MAINT_DATA; shipping source data as-is (rate may be 0).',
          file=sys.stderr, flush=True)
    return maint

# ── HTML Template ────────────────────────────────────────────────────────
# JS: COLUMN_DEFS_SUMMARY and COLUMN_DEFS_FULL are injected from Python.
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CUL Daily Movement</title>
<!-- Lazy-load xlsx-js-style -->
<script>
var _xlsxReady=false,_xlsxStyled=true;
function _loadXlsx(cb){
  if(window.XLSX){cb();return;}
  var s=document.createElement('script');
  s.src='https://unpkg.com/xlsx-js-style@1.2.0/dist/xlsx.bundle.js';
  s.onerror=function(){
    _xlsxStyled=false;
    s.src='https://cdn.bootcdn.net/ajax/libs/SheetJS/xlsx.full.min.js';
    s.onerror=function(){
      s.src='https://unpkg.com/xlsx@0.18.5/dist/xlsx.full.min.js';
      s.onerror=function(){alert('Failed to load Excel library. Please check your network.');};
      document.head.appendChild(s);
    };
    document.head.appendChild(s);
  };
  s.onload=cb;
  document.head.appendChild(s);
}
</script>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', Arial, sans-serif; background: #f0f4f8; color: #1a2332; }

  /* ── Header ─────────────────────────────────────────────────────────── */
  .header {
    background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%);
    color: #fff; padding: 14px 28px 10px;
    box-shadow: 0 3px 12px rgba(31,78,121,.35);
  }
  .header-top { display: flex; align-items: center; justify-content: space-between; }
  .header-left h1 { font-size: 18px; font-weight: 700; letter-spacing: .5px; }
  .header-left .sub { font-size: 11px; opacity: .75; margin-top: 2px; }
  .header-right { display: flex; gap: 10px; align-items: center; }
  .btn { padding: 7px 16px; border: none; border-radius: 5px; font-size: 13px; font-weight: 600; cursor: pointer; transition: .15s; }
  .btn-export { background: #F6A623; color: #fff; }
  .btn-export:hover { background: #d4891a; }
  .btn-history { background: rgba(255,255,255,.18); color: #fff; border: 1px solid rgba(255,255,255,.4); }
  .btn-history:hover { background: rgba(255,255,255,.30); }

  /* ── Tabs ────────────────────────────────────────────────────────────── */
  .tabs { display: flex; gap: 0; padding: 0 28px; background: #fff; border-bottom: 2px solid #dde4ed; }
  .tab-btn {
    padding: 10px 24px; font-size: 13px; font-weight: 600; cursor: pointer;
    border: none; background: none; color: #5a6e82; border-bottom: 3px solid transparent;
    transition: .15s; margin-bottom: -2px;
  }
  .tab-btn:hover { color: #1F4E79; }
  .tab-btn.active { color: #1F4E79; border-bottom-color: #1F4E79; }
  .tab-content { display: none; }
  .tab-content.active { display: block; }

  /* ── Sticky top bar: header + tabs ─────────────────────────────────────── */
  .top-pinned {
    position: sticky; top: 0; z-index: 100;
  }

  /* ── Controls (sticky: stays pinned below the header+tabs bar) ───────── */
  .controls {
    padding: 14px 28px; background: #fff; border-bottom: 1px solid #dde4ed;
    display: flex; gap: 14px; align-items: center; flex-wrap: wrap;
    position: sticky; top: var(--top-h, 105px); z-index: 60;
    box-shadow: 0 3px 10px rgba(31,78,121,.12);
  }
  /* Section-level filter bars (non-sticky) so their dropdowns never overlay top controls */
  .sub-controls {
    padding: 14px 28px; background: #fff; border-bottom: 1px solid #dde4ed;
    display: flex; gap: 14px; align-items: center; flex-wrap: wrap;
  }
  .controls input, .controls select {
    padding: 6px 12px; border: 1px solid #c9d5e2; border-radius: 5px;
    font-size: 13px; outline: none; height: 34px;
  }
  .controls input:focus, .controls select:focus { border-color: #2E75B6; box-shadow: 0 0 0 2px rgba(46,117,182,.15); }
  .controls input { width: 220px; }
  .controls input[type="date"] {
    width: 138px; padding: 6px 10px; font-size: 12px;
    color: #1F4E79; font-weight: 500; font-family: inherit;
  }
  .eta-label { font-size: 12px; color: #5a6e82; font-weight: 600; margin-right: -6px; }
  .date-type-select { padding: 6px 8px; border: 1px solid #c8d6e5; border-radius: 6px; background: #fff; color: #333; font-size: 12px; font-weight: 600; cursor: pointer; }
  .eta-sep { color: #a8b8c8; font-weight: 400; margin: 0 -2px; }
  .controls select { min-width: 130px; }
  .col-toggle-btn {
    padding: 6px 14px; border: 1px solid #c9d5e2; border-radius: 5px;
    font-size: 12px; cursor: pointer; background: #fff; color: #1F4E79;
    font-weight: 600; transition: .15s; height: 34px;
  }
  .col-toggle-btn:hover { background: #EBF3FB; border-color: #2E75B6; }
  .filter-btn {
    padding: 6px 14px; border: 1px solid #c9d5e2; border-radius: 5px;
    font-size: 12px; cursor: pointer; background: #fff; color: #1F4E79;
    font-weight: 600; transition: .15s; min-width: 130px; text-align: left;
    height: 34px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 180px;
  }
  .filter-btn:hover { background: #EBF3FB; border-color: #2E75B6; }
  .filter-btn.has-selection { background: #EBF3FB; border-color: #2E75B6; color: #0d3b5e; }
  .decom-toggle {
    display: inline-flex; align-items: center; gap: 4px;
    height: 34px; padding: 0 10px;
    border: 1px solid #c9d5e2; border-radius: 5px;
    background: #fff; color: #8a9bb0; font-size: 12px; font-weight: 600;
    cursor: pointer; white-space: nowrap; user-select: none;
    transition: .15s;
  }
  .decom-toggle:hover { background: #F5F8FB; }
  .decom-toggle .decom-dot {
    width: 14px; height: 14px; border-radius: 50%;
    background: #e4ecf5; color: #fff; font-size: 9px; font-weight: 700;
    display: inline-flex; align-items: center; justify-content: center;
    transition: .15s;
  }
  .decom-toggle.on { color: #1F4E79; border-color: #2E75B6; background: #EBF3FB; }
  .decom-toggle.on .decom-dot { background: #2E75B6; }
  .decom-toggle input { display: none; }
  .filter-dropdown { min-width: 200px; }
  .filter-dropdown .filter-actions {
    display: flex; gap: 0; border-bottom: 1px solid #e4ecf5; padding: 4px 8px;
    margin-bottom: 4px;
  }
  .filter-dropdown .filter-actions button {
    flex: 1; padding: 4px 8px; border: none; background: none; font-size: 11px;
    color: #2E75B6; cursor: pointer; font-weight: 600; border-radius: 3px; transition:.1s;
  }
  .filter-dropdown .filter-actions button:hover { background: #EBF3FB; }
  .filter-dropdown label { font-size: 12px; }
  .col-dropdown {
    display: none; position: absolute; background: #fff; border: 1px solid #c9d5e2;
    border-radius: 8px; padding: 10px 0; min-width: 200px;
    box-shadow: 0 8px 30px rgba(0,0,0,.18); z-index: 50; max-height: 400px; overflow-y: auto;
  }
  .col-dropdown.open { display: block; }
  .col-dropdown label {
    display: flex; align-items: center; gap: 8px; padding: 6px 16px;
    font-size: 12.5px; cursor: pointer; transition: .1s;
  }
  .col-dropdown label:hover { background: #f0f7ff; }
  .col-dropdown input[type="checkbox"] { accent-color: #1F4E79; width: 15px; height: 15px; }
  .stat-chip { margin-left: auto; background: #EBF3FB; border: 1px solid #c3d9f0; border-radius: 20px; padding: 4px 14px; font-size: 12px; color: #1F4E79; font-weight: 600; }
  .delay-chip { background: #fff0f0; border: 1px solid #f5c6c6; border-radius: 20px; padding: 4px 14px; font-size: 12px; color: #c00000; font-weight: 600; }
  td.delay { background: #fff0f0 !important; color: #c00000; font-weight: 700; }
  td.ontime { color: #27ae60; font-weight: 600; }
  tr.port-row:hover { filter: brightness(0.96); }
  tr.detail-wrap td { border-bottom: 2px solid #d4e0eb; }
  tr.detail-row:hover { background: #f5f8fb !important; }

  /* ── Tables ──────────────────────────────────────────────────────────── */
  .table-wrap {
    overflow-x: auto; padding: 0 28px 28px; position: relative;
    /* inner vertical scroll so the sticky header stays visible:
       viewport minus pinned top bar height minus controls height minus a small margin */
    max-height: calc(100vh - var(--top-h, 105px) - var(--ctrl-h, 62px) - 12px);
    overflow-y: auto;
  }
  table { width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 12.5px; min-width: 1200px; }
  th {
    background: #1F4E79; color: #fff; font-weight: 600; padding: 9px 8px;
    text-align: center; white-space: nowrap; position: sticky; top: 0; z-index: 5;
    cursor: pointer; user-select: none;
  }
  th:hover { background: #163b5f; }
  th .sort-arrow { display: inline-block; margin-left: 4px; opacity: .5; font-size: 10px; }
  th.sort-asc .sort-arrow::after { content: '\25b2'; opacity: 1; }
  th.sort-desc .sort-arrow::after { content: '\25bc'; opacity: 1; }
  th:not(.sort-asc):not(.sort-desc) .sort-arrow::after { content: '\21c5'; }
  td { padding: 7px 9px; border-bottom: 1px solid #e4ecf5; vertical-align: middle; text-align: center; }

  /* Summary table row striping */
  #summaryView tr:nth-child(even) td { background: #f5f9fe; }
  #summaryView tr:nth-child(odd)  td { background: #fff; }
  #summaryView tr:hover td { background: #dcedf9 !important; }

  /* Full schedule: color-band by vessel group */
  .vessel-group-even td { background: #f5f9fe !important; }
  .vessel-group-odd  td { background: #fff !important; }
  .vessel-group-even:hover td, .vessel-group-odd:hover td { background: #dcedf9 !important; }
  .summary-highlight td { background: #fff3cd !important; }
  .summary-highlight:hover td { background: #ffe69c !important; }
  .vessel-group-first td { border-top: 3px solid #2E75B6; }
  .vessel-group-last  td { border-bottom: 2px solid #c3d9f0; }

  .td-center { text-align: center; }
  .td-idx { text-align: center; color: #8fa3b8; font-size: 11px; font-weight: 500; width: 40px; }
  .td-mono { font-family: 'Consolas', 'Courier New', monospace; font-size: 12px; }
  .badge-route { display: inline-block; padding: 2px 8px; border-radius: 4px; font-weight: 700; font-size: 11px; background: #1F4E79; color: #fff; letter-spacing: .5px; }
  .badge-code { display: inline-block; padding: 2px 7px; border-radius: 4px; font-size: 11px; font-weight: 600; background: #E8F4FD; color: #1F4E79; border: 1px solid #a8cfe8; }
  .weekday-tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 11px; font-weight: 600; background: #1F4E79; color: #fff; }
  .delay-tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 11px; font-weight: 700; background: #fff0f0; color: #c00000; border: 1px solid #f5c6c6; }
  .ahead-tag { display: inline-block; padding: 2px 7px; border-radius: 3px; font-size: 11px; font-weight: 700; background: #f0fff4; color: #1a7340; border: 1px solid #b7dfca; }
  .no-data { text-align: center; padding: 40px; color: #8a9bb0; font-size: 14px; }
  .remark-cell { max-width: 260px; white-space: normal; line-height: 1.4; color: #5a6e82; font-size: 11.5px; }
  .remark-note { display: block; max-width: 200px; white-space: normal; line-height: 1.3; color: #e67e22; font-size: 10.5px; font-weight: normal; margin-top: 2px; }
  .vessel-label {
    font-weight: 700; color: #1F4E79; font-size: 12px;
    display: inline-block; padding: 1px 6px; border-radius: 3px;
    background: #E8F4FD; border: 1px solid #a8cfe8;
  }
  .proforma-cell { font-family: 'Consolas', monospace; font-size: 11px; color: #5a6e82; background: #f8fafc; border-radius: 3px; padding: 1px 5px; }

  /* ── Modal ───────────────────────────────────────────────────────────── */
  .modal-overlay { display: none; position: fixed; inset: 0; background: rgba(0,0,0,.5); z-index: 100; align-items: center; justify-content: center; }
  .modal-overlay.open { display: flex; }
  .modal { background: #fff; border-radius: 10px; width: 92%; max-width: 980px; max-height: 85vh; display: flex; flex-direction: column; box-shadow: 0 20px 60px rgba(0,0,0,.3); }
  .modal-header { padding: 16px 22px; background: #1F4E79; color: #fff; border-radius: 10px 10px 0 0; display: flex; justify-content: space-between; align-items: center; }
  .modal-header h3 { font-size: 15px; }
  .modal-close { cursor: pointer; font-size: 20px; opacity: .7; background: none; border: none; color: #fff; }
  .modal-close:hover { opacity: 1; }
  .modal-body { overflow-y: auto; padding: 16px 22px; }
  .history-date-list { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 16px; }
  .history-date-btn { padding: 5px 14px; border: 1px solid #c9d5e2; border-radius: 20px; font-size: 12px; cursor: pointer; background: #fff; transition: .15s; }
  .history-date-btn:hover, .history-date-btn.active { background: #1F4E79; color: #fff; border-color: #1F4E79; }
  .history-table-wrap { overflow-x: auto; }
  .history-table-wrap table { font-size: 11.5px; min-width: 900px; }
  .footer { text-align: center; padding: 12px; color: #8a9bb0; font-size: 11px; }

  /* ── Lane Order modal ─────────────────────────────────────────────── */
  .lane-row { display: flex; align-items: center; gap: 10px; padding: 5px 8px; border-radius: 5px; }
  .lane-row:nth-child(odd) { background: #f6f9fc; }
  .lane-row .lane-idx { width: 24px; text-align: center; color: #8a9bb0; font-size: 12px; font-variant-numeric: tabular-nums; }
  .lane-row .lane-code { flex: 1; font-weight: 600; color: #1F4E79; font-size: 13px; font-family: 'Consolas', monospace; }
  .lane-move { width: 30px; height: 26px; border: 1px solid #c9d5e2; background: #fff; border-radius: 5px; cursor: pointer; font-size: 13px; color: #3b4a5a; line-height: 1; }
  .lane-move:hover:not(:disabled) { background: #1F4E79; color: #fff; border-color: #1F4E79; }
  .lane-move:disabled { opacity: .3; cursor: default; }

  /* ── Login Overlay ──────────────────────────────────────────────── */
  .login-overlay { display: flex; position: fixed; inset: 0; background: #0a1628; z-index: 9999; align-items: center; justify-content: center; }
  .login-overlay.hidden { display: none; }
  .login-box { background: #fff; border-radius: 10px; padding: 32px 40px; width: 360px; box-shadow: 0 20px 60px rgba(0,0,0,.5); text-align: center; }
  .login-box h2 { margin: 0 0 6px; color: #1F4E79; font-size: 20px; }
  .login-box .sub { color: #8a9bb0; font-size: 12px; margin-bottom: 24px; }
  .login-box input { width: 100%; padding: 10px 14px; border: 1px solid #c9d5e2; border-radius: 6px; font-size: 14px; box-sizing: border-box; margin-bottom: 12px; outline: none; }
  .login-box input:focus { border-color: #2E75B6; box-shadow: 0 0 0 2px rgba(46,117,182,.15); }
  .login-box .login-err { color: #e74c3c; font-size: 12px; margin-bottom: 8px; min-height: 18px; }
  .login-box button { width: 100%; padding: 10px; background: #1F4E79; color: #fff; border: none; border-radius: 6px; font-size: 14px; cursor: pointer; font-weight: 600; }
  .login-box button:hover { background: #2E75B6; }

  /* ── BOA berth-on-arrival stats section ─────────────────────────────── */
  .range-opt { display: inline-flex; align-items: center; gap: 6px; background: #eef2f7; border: 1px solid #d5dee8; border-radius: 16px; padding: 4px 11px; cursor: pointer; font-size: 12px; color: #3b4a5a; user-select: none; }
  .range-opt input { margin: 0; accent-color: #1F4E79; }
  .range-opt.sel { background: #1F4E79; border-color: #1F4E79; color: #fff; }
  .range-opt.sel input { accent-color: #fff; }
  .boa-chips { display: flex; gap: 10px; flex-wrap: wrap; margin-top: 10px; }
  .boa-chip { background: #fff; border: 1px solid #e2e8f0; border-radius: 8px; padding: 8px 16px; min-width: 120px; }
  .boa-chip .num { font-size: 20px; font-weight: 700; color: #1F4E79; }
  .boa-chip .lbl2 { font-size: 11px; color: #8a9bb0; }
  .boa-chip.green .num { color: #2e8b57; }
  .boa-chip.orange .num { color: #e67e22; }
  #boaTblTrade th, #boaTblLane th, #boaTblRegion th, #boaTblPort th { cursor: pointer; white-space: nowrap; font-size: 12px; }
  #boaTblTrade th:hover, #boaTblLane th:hover, #boaTblRegion th:hover, #boaTblPort th:hover { background: #e2ebf5; }
  #boaTblTrade td, #boaTblLane td, #boaTblRegion td, #boaTblPort td { font-size: 12px; }
  .rate-bad { color: #c0392b; font-weight: 600; }
  .rate-mid { color: #e67e22; font-weight: 600; }
  .rate-good { color: #2e8b57; font-weight: 600; }
  .boa-grp { font-weight: 600; color: #33475b; }
  .boa-subrow td { background: #fbfcfe; }
  .boa-sumrow td { background: #eef4fa; font-weight: 600; color: #1F4E79; }
  .boa-grand td { background: #1F4E79; color: #fff; font-weight: 700; }
  .boa-grand td.rate-bad, .boa-grand td.rate-mid, .boa-grand td.rate-good { color: #fff; }
</style>
</head>
<body>

<!-- ── Login Overlay ──────────────────────────────────────────────── -->
<div class="login-overlay" id="loginOverlay">
  <div class="login-box">
    <h2>&#128274; Access Restricted</h2>
    <div class="sub">CUL Daily Movement Dashboard</div>
    <input type="password" id="loginPwd" placeholder="Enter password" onkeydown="if(event.key==='Enter')doLogin()" autofocus>
    <div class="login-err" id="loginErr"></div>
    <button onclick="doLogin()">Sign In</button>
  </div>
</div>

<!-- ── Sticky top bar: header + tabs stay pinned on scroll ───────────── -->
<div class="top-pinned">
  <!-- ── Header ──────────────────────────────────────────────────────────── -->
  <div class="header">
    <div class="header-top">
      <div class="header-left">
        <h1>&#9875; CUL VESSEL DAILY MOVEMENT</h1>
        <div class="sub" id="headerDate">Loading&hellip;</div>
      </div>
      <div class="header-right">
        <button class="btn btn-history" onclick="openHistory()">&#128203; History</button>
        <button class="btn btn-export" id="btnExport" onclick="exportExcel()">&#8595; Export Excel</button>
        <button class="btn btn-export" onclick="openExportTablesModal()">&#128196; Export Tables</button>
      </div>
    </div>
  </div>

  <!-- ── Tabs ────────────────────────────────────────────────────────────── -->
  <div class="tabs">
    <button class="tab-btn active" data-tab="summaryView" onclick="switchTab('summaryView',this)">&#128202; Summary</button>
    <button class="tab-btn" data-tab="fullScheduleView" onclick="switchTab('fullScheduleView',this)">&#128203; Full Schedule</button>
    <button class="tab-btn" data-tab="portView" onclick="switchTab('portView',this)">&#9889; Port Wait</button>
    <button class="tab-btn" data-tab="maintView" onclick="switchTab('maintView',this)">&#128203; Maintenance</button>
    <button class="tab-btn" data-tab="speedView" onclick="switchTab('speedView',this)">&#128168; Speed</button>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════════
     VIEW 1: Summary (one row per vessel, nearest ETA)
     ═════════════════════════════════════════════════════════════════════ -->
<div id="summaryView" class="tab-content active">
  <div class="controls">
    <input type="text" id="searchBox" placeholder="&#128269; Search vessel / port / PIC&hellip;" oninput="renderSummary()">
    <div style="position:relative;">
      <button class="filter-btn" id="filterRouteBtn1" onclick="toggleFilterDropdown('route','1')">All Routes</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownRoute1"></div>
    </div>
    <div style="position:relative;">
      <button class="filter-btn" id="filterVesselBtn1" onclick="toggleFilterDropdown('vessel','1')">All Vessels</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownVessel1"></div>
    </div>
    <div style="position:relative;">
      <button class="filter-btn" id="filterPicBtn1" onclick="toggleFilterDropdown('pic','1')">All PIC</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownPic1"></div>
    </div>
    <select id="filterDelay" onchange="renderSummary()">
      <option value="">All Status</option>
      <option value="delay">Delay Only</option>
      <option value="ahead">Ahead Only</option>
      <option value="normal">No Delay</option>
    </select>
    <div style="position:relative;">
      <button class="col-toggle-btn" id="colToggleBtn1" onclick="toggleColDropdown('1')">&#9881; Columns</button>
      <div class="col-dropdown" id="colDropdown1"></div>
    </div>
    <label class="decom-toggle" title="显示/隐藏已下线船舶 (deprecated vessels)">
      <input type="checkbox" class="decom-cb" onchange="onDecomToggle(this)">
      <span class="decom-dot">+</span>
      <span>Show 已下线</span>
    </label>
    <span class="stat-chip" id="statTotal">&#8212; vessels</span>
    <span class="delay-chip" id="statDelay">&#8212; delayed</span>
  </div>
  <div class="table-wrap">
    <table id="summaryTable">
      <thead><tr id="summaryThead"></tr></thead>
      <tbody id="summaryTbody"></tbody>
    </table>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════════
     VIEW 2: Full Schedule (all ports for all vessels)
     ═════════════════════════════════════════════════════════════════════ -->
<div id="fullScheduleView" class="tab-content">
  <div class="controls">
    <input type="text" id="searchBox2" placeholder="&#128269; Search vessel / port / PIC / voyage&hellip;" oninput="renderFullSchedule()">
    <select id="dateFilterType2" class="date-type-select" onchange="onDateFilterTypeChange()">
      <option value="eta" selected>ETA</option>
      <option value="etb">ETB</option>
    </select>
    <input type="date" id="etaFrom2" title="ETA from" onchange="renderFullSchedule()">
    <span class="eta-sep">–</span>
    <input type="date" id="etaTo2" title="ETA to" onchange="renderFullSchedule()">
    <div style="position:relative;">
      <button class="filter-btn" id="filterRouteBtn2" onclick="toggleFilterDropdown('route','2')">All Routes</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownRoute2"></div>
    </div>
    <div style="position:relative;">
      <button class="filter-btn" id="filterVesselBtn2" onclick="toggleFilterDropdown('vessel','2')">All Vessels</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownVessel2"></div>
    </div>
    <div style="position:relative;">
      <button class="filter-btn" id="filterPicBtn2" onclick="toggleFilterDropdown('pic','2')">All PIC</button>
      <div class="filter-dropdown col-dropdown" id="filterDropdownPic2"></div>
    </div>
    <div style="position:relative;">
      <button class="col-toggle-btn" id="colToggleBtn2" onclick="toggleColDropdown('2')">&#9881; Columns</button>
      <div class="col-dropdown" id="colDropdown2"></div>
    </div>
    <button class="filter-btn" id="laneOrderBtn" onclick="openLaneOrderModal()">&#8644; Lane Order</button>
    <label class="decom-toggle" title="显示/隐藏已下线船舶 (deprecated vessels)">
      <input type="checkbox" class="decom-cb" onchange="onDecomToggle(this)">
      <span class="decom-dot">+</span>
      <span>Show 已下线</span>
    </label>
    <span class="stat-chip" id="statTotal2">&#8212; rows</span>
  </div>
  <div class="table-wrap">
    <table id="fullTable">
      <thead><tr id="fullThead"></tr></thead>
      <tbody id="fullTbody"></tbody>
    </table>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════════
     VIEW 3: Port Wait Analytics
     ═══════════════════════════════════════════════════════════════════ -->
<div id="portView" class="tab-content">
  <div class="controls" style="flex-wrap:wrap;">
    <span style="font-weight:600;font-size:14px;margin-right:4px;">&#128197; Date:</span>
    <input type="date" id="portDateFrom" title="From" onchange="onPortDateChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;width:130px;">
    <span style="margin:0 4px;font-size:13px;">to</span>
    <input type="date" id="portDateTo" title="To" onchange="onPortDateChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;width:130px;">
    <span style="font-weight:600;font-size:14px;margin:0 6px;">&#9201; By:</span>
    <select id="portDateBasis" onchange="onPortDateChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;">
      <option value="eta" selected>ETA</option>
      <option value="etb">ETB</option>
      <option value="etd">ETD</option>
    </select>
    <span style="font-weight:600;font-size:14px;margin:0 8px;">&#128205; Port:</span>
    <div style="position:relative;">
      <button class="filter-btn" id="portFilterBtn" onclick="togglePortFilter()">All Ports</button>
      <div class="filter-dropdown col-dropdown" id="portFilterDropdown"></div>
    </div>
    <input type="text" id="portSearch" placeholder="&#128269; Search ports..." oninput="onPortSearch()" style="margin-left:8px;padding:5px 8px;border:1px solid #b8c6d6;border-radius:4px;font-size:12px;width:160px;">
    <span style="font-weight:600;font-size:14px;margin:0 8px;">&#128205; Remark:</span>
    <div style="position:relative;">
      <button class="filter-btn" id="remarkFilterBtn" onclick="toggleRemarkFilter()">All Remarks</button>
      <div class="filter-dropdown col-dropdown" id="remarkFilterDropdown"></div>
    </div>
    <span class="stat-chip" id="statPortWait" style="margin-left:12px;">&#8212; port entries</span>
  </div>

  <!-- Port Wait Analysis -->
  <h3 style="margin:16px 0 8px;color:#1F4E79;">&#9889; Port Wait Time Analysis</h3>
  <p style="font-size:11px;color:#8a9bb0;margin:0 0 8px;">Ports normalized (terminal suffixes merged). Bunkering-only calls excluded. Berth Rate = calls berthed within threshold (wait &#8804; 12h for CNSHA/CNNGB, &#8804; 6h for other ports) / total calls in selected range. Remark filters apply to both numerator and denominator. Ranked best&#8594;worst by default.</p>
  <div class="sub-controls" style="padding:8px 12px;margin-bottom:8px;">
    <span style="font-weight:600;font-size:13px;margin-right:6px;">&#9201; Over-range:</span>
    <label class="range-opt sel" id="port-opt-all"><input type="radio" name="portrange" value="all" checked> All over (&gt; standard)</label>
    <label class="range-opt" id="port-opt-24"><input type="radio" name="portrange" value="24+"> 24h+</label>
    <label class="range-opt" id="port-opt-48"><input type="radio" name="portrange" value="48+"> 48h+</label>
  </div>
  <div class="table-wrap">
    <table id="portWaitTable">
      <thead><tr id="portWaitThead"></tr></thead>
      <tbody id="portWaitTbody"></tbody>
    </table>
  </div>

  <!-- Wait by Port Region -->
  <h3 style="margin:20px 0 8px;color:#1F4E79;">&#9889; Wait by Port Region</h3>
  <p style="font-size:11px;color:#8a9bb0;margin:0 0 8px;">Port wait aggregated by region (Port&#8594;Region per mapping). Over (range) follows the over-range selector above. Shows Calls / Total Wait / Avg Wait / Over / Over rate.</p>
  <div class="table-wrap">
    <table id="portWaitRegionTable"><thead><tr id="portWaitRegionThead"></tr></thead><tbody id="portWaitRegionTbody"></tbody></table>
  </div>

  <!-- Port Call Count by Lane / Region (side by side) -->
  <div style="display:flex;flex-wrap:wrap;gap:16px;margin-top:8px;">
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h3 style="margin:20px 0 8px;color:#1F4E79;">&#128202; Port Call Count by Lane</h3>
      <p style="font-size:11px;color:#8a9bb0;margin:0 0 8px;">Count of port calls per lane, grouped by trade (Lane&#8594;Trade mapping). Follows Port Wait date/port/remark filters.</p>
      <table id="callCountLaneTable" style="min-width:auto;"><thead><tr id="callCountLaneThead"></tr></thead><tbody id="callCountLaneTbody"></tbody></table>
    </div>
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h3 style="margin:20px 0 8px;color:#1F4E79;">&#128202; Port Call Count by Region</h3>
      <p style="font-size:11px;color:#8a9bb0;margin:0 0 8px;">Count of port calls per port, grouped by region (Port&#8594;Region mapping). Follows Port Wait date/port/remark filters.</p>
      <table id="callCountRegionTable" style="min-width:auto;"><thead><tr id="callCountRegionThead"></tr></thead><tbody id="callCountRegionTbody"></tbody></table>
    </div>
  </div>

  <!-- Remark Category Wait Breakdown & Monthly Trend (side by side) -->
  <div style="display:flex;flex-wrap:wrap;gap:16px;margin-top:8px;">
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h4 style="margin:6px 0 10px;color:#1F4E79;font-size:13px;text-align:center;">&#128202; Wait Time by Remark Category</h4>
      <p style="font-size:10px;color:#8a9bb0;margin:0 0 8px;">Category breakdown of port wait within selected filters and time range.</p>
      <div id="remarkSummary" style="display:flex;flex-direction:column;gap:6px;"></div>
    </div>
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h4 style="margin:6px 0 10px;color:#1F4E79;font-size:13px;text-align:center;">&#128200; Monthly Port Wait Trend</h4>
      <p style="font-size:10px;color:#8a9bb0;margin:0 0 8px;">Monthly aggregation of port wait data within the selected time range and filters.</p>
      <div id="monthlyTrend" style="display:flex;flex-direction:column;gap:6px;"></div>
    </div>
  </div>

  <!-- BOA berth-on-arrival stats (computed from Daily Movement, follows Port Wait date filter) -->
  <div style="margin-top:24px;border-top:2px solid #1F4E79;padding-top:14px;">
    <h3 style="margin:0 0 4px;color:#1F4E79;">&#128202; BOA Berth-On-Arrival Stats <span id="boaLabelSpan" style="font-weight:400;font-size:12px;color:#8a9bb0;"></span></h3>
    <p style="font-size:11px;color:#8a9bb0;margin:0 0 10px;">Source: CUL Daily Movement (follows Port Wait date range above) &middot; Berth = WAIT &le; threshold (CNSHA &amp; CNNGB 12h, others 6h) &middot; Lane&#8594;Trade / Port&#8594;Region per &#8220;Port &amp; Lane Mapping&#8221;. Over-range options: All over = WAIT &gt; threshold; 24h+ = WAIT &ge;24h; 48h+ = WAIT &ge;48h. Only the over count changes with range; berth count is unchanged.</p>

    <div class="sub-controls" style="padding:8px 12px;">
      <span style="font-weight:600;font-size:13px;margin-right:6px;">&#9201; Over-range:</span>
      <label class="range-opt sel" id="boa-opt-all"><input type="radio" name="boarange" value="all" checked> All over (&gt; standard)</label>
      <label class="range-opt" id="boa-opt-24"><input type="radio" name="boarange" value="24+"> 24h+</label>
      <label class="range-opt" id="boa-opt-48"><input type="radio" name="boarange" value="48+"> 48h+</label>
    </div>

    <div class="boa-chips">
      <div class="boa-chip"><div class="num" id="boaChipTotal">&#8212;</div><div class="lbl2">Total calls</div></div>
      <div class="boa-chip green"><div class="num" id="boaChipBerth">&#8212;</div><div class="lbl2">Berth</div></div>
      <div class="boa-chip orange"><div class="num" id="boaChipOver">&#8212;</div><div class="lbl2">Over (current range)</div></div>
      <div class="boa-chip"><div class="num" id="boaChipRate">&#8212;</div><div class="lbl2">Berth rate (current range)</div></div>
    </div>

    <div style="display:flex;flex-wrap:wrap;gap:16px;margin-top:10px;">
      <div class="table-wrap" style="flex:1;min-width:380px;">
        <h4 style="margin:6px 0 8px;color:#1F4E79;font-size:13px;">&#9312; BOA by Trade</h4>
        <table id="boaTblTrade" style="min-width:auto"><thead><tr></tr></thead><tbody></tbody></table>
      </div>
      <div class="table-wrap" style="flex:1;min-width:380px;">
        <h4 style="margin:6px 0 8px;color:#1F4E79;font-size:13px;">&#9313; BOA by Trade &rarr; Lane (Lane mapped to its Trade)</h4>
        <table id="boaTblLane" style="min-width:auto"><thead><tr></tr></thead><tbody></tbody></table>
      </div>
    </div>
    <div style="display:flex;flex-wrap:wrap;gap:16px;margin-top:10px;">
      <div class="table-wrap" style="flex:1;min-width:380px;">
        <h4 style="margin:6px 0 8px;color:#1F4E79;font-size:13px;">&#9314; BOA by Port Region</h4>
        <table id="boaTblRegion" style="min-width:auto"><thead><tr></tr></thead><tbody></tbody></table>
      </div>
      <div class="table-wrap" style="flex:1;min-width:380px;">
        <h4 style="margin:6px 0 8px;color:#1F4E79;font-size:13px;">&#9315; BOA by Port (with Region subtotals)</h4>
        <table id="boaTblPort" style="min-width:auto"><thead><tr></tr></thead><tbody></tbody></table>
      </div>
    </div>
    <p style="font-size:10px;color:#8a9bb0;margin:4px 0 0;">Click headers to sort &middot; Green &#8805;80% &middot; Orange 50&#8211;80% &middot; Red &lt;50%</p>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════════
     VIEW 4: Speed Analytics
     ═══════════════════════════════════════════════════════════════════ -->
<div id="speedView" class="tab-content">
  <div class="controls">
    <span style="font-weight:600;font-size:14px;margin-right:8px;">&#128674; Vessel Filter:</span>
    <div style="position:relative;">
      <button class="filter-btn" id="speedVesselFilterBtn" onclick="toggleSpeedVesselFilter()">All Vessels</button>
      <div class="filter-dropdown col-dropdown" id="speedVesselFilterDropdown"></div>
    </div>
    <input type="text" id="speedVesselSearch" placeholder="&#128269; Search vessel..." oninput="onSpeedVesselSearch()" style="margin-left:10px;padding:5px 8px;border:1px solid #b8c6d6;border-radius:4px;font-size:12px;width:170px;">
    <span class="stat-chip" id="statSpeed" style="margin-left:12px;">&#8212; vessels</span>
    <button class="filter-btn" style="margin-left:12px;" onclick="exportSpeedExcel()">&#8595; Export Speed</button>
  </div>

  <!-- Vessel Speed Analysis -->
  <h3 style="margin:16px 0 8px;color:#1F4E79;">&#128168; Vessel Speed Analysis</h3>
  <p style="font-size:11px;color:#8a9bb0;margin:0 0 8px;">Speed values from source Excel column "SPEED" (calculated as leg distance &#247; sailing time). Min/Max = extreme values per vessel; values &#8804;0 or >20kn excluded as dirty data. Hover values to see raw source.</p>
  <div class="table-wrap">
    <table id="speedTable">
      <thead><tr id="speedThead"></tr></thead>
      <tbody id="speedTbody"></tbody>
    </table>
  </div>
</div>

<!-- ── History Modal ──────────────────────────────────────────────────── -->
<div class="modal-overlay" id="historyModal">
  <div class="modal">
    <div class="modal-header">
      <h3>&#128203; Historical Records</h3>
      <button class="modal-close" onclick="closeHistory()">&#10005;</button>
    </div>
    <div class="modal-body">
      <div class="history-date-list" id="historyDateList"></div>
      <div class="history-table-wrap" id="historyTableWrap"></div>
    </div>
  </div>
</div>

<!-- ═══════════════════════════════════════════════════════════════════
     VIEW 3b: Maintenance Rate (Vessel Schedule / Port Log 维护率)
     ═══════════════════════════════════════════════════════════════════ -->
<div id="maintView" class="tab-content">
  <div class="controls" style="flex-wrap:wrap;">
    <span style="font-weight:600;font-size:14px;margin-right:4px;">&#128100; Operator:</span>
    <select id="maintOpFilter" onchange="onMaintChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;">
      <option value="ALL">All Operators</option>
    </select>
    <span style="font-weight:600;font-size:14px;margin:0 8px;">&#128197; ETD:</span>
    <input type="date" id="maintFrom" title="ETD from" onchange="onMaintChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;width:130px;">
    <span style="margin:0 4px;font-size:13px;">to</span>
    <input type="date" id="maintTo" title="ETD to" onchange="onMaintChange()" style="font-size:12px;padding:3px 6px;border:1px solid #ccc;border-radius:4px;width:130px;">
    <span class="stat-chip" id="statMaint" style="margin-left:12px;">&#8212; calls</span>
    <span style="font-weight:600;font-size:13px;margin:0 6px 0 14px;">&#128205; Service:</span>
    <div style="position:relative;display:inline-block;">
      <button class="filter-btn" id="maintServiceFilterBtn" onclick="toggleMaintServiceFilter()">All Services</button>
      <div class="filter-dropdown col-dropdown" id="maintServiceFilterDropdown"></div>
    </div>
    <span style="font-weight:600;font-size:13px;margin:0 6px 0 14px;">&#128674; Vessel:</span>
    <div style="position:relative;display:inline-block;">
      <button class="filter-btn" id="maintVesselFilterBtn" onclick="toggleMaintVesselFilter()">All Vessels</button>
      <div class="filter-dropdown col-dropdown" id="maintVesselFilterDropdown"></div>
    </div>
    <span style="font-weight:600;font-size:13px;margin:0 6px 0 14px;">&#128205; Port:</span>
    <div style="position:relative;display:inline-block;">
      <button class="filter-btn" id="maintPortFilterBtn" onclick="toggleMaintPortFilter()">All Ports</button>
      <div class="filter-dropdown col-dropdown" id="maintPortFilterDropdown"></div>
    </div>
    <input type="text" id="maintPortSearch" placeholder="&#128269; Search ports..." oninput="onMaintPortSearch()" style="margin-left:8px;padding:5px 8px;border:1px solid #b8c6d6;border-radius:4px;font-size:12px;width:160px;">
  </div>

  <p style="font-size:11px;color:#8a9bb0;margin:8px 0 6px;">Source: Vessel Schedule / Port Log 维护台账 (follows Operator &amp; ETD filters above). 判定规则: Port Log 维护 = I 列 Y；Vessel Schedule 维护 = J 列 &quot;Maintain timely&quot;（Not maintained 即未维护）。两类未维护明细已分开列出。</p>
  <p style="font-size:12px;color:#1F4E79;margin:0 0 8px;font-weight:600;">&#128336; Maintenance 数据更新时间：<span id="maintSourceTs">—</span></p>

  <!-- Overall chips -->
  <div class="boa-chips" id="maintChips">
    <div class="boa-chip"><div class="num" id="maintChipTotal">&#8212;</div><div class="lbl2">Total calls</div></div>
    <div class="boa-chip"><div class="num" id="maintChipPortLog">&#8212;</div><div class="lbl2">Port Log maintained</div></div>
    <div class="boa-chip" style="background:#fdecea;border-color:#e6b8af;"><div class="num" id="maintChipPortLogNo" style="color:#C0392B;">&#8212;</div><div class="lbl2" style="color:#C0392B;">Port Log NOT maintained</div></div>
    <div class="boa-chip"><div class="num" id="maintChipVSched">&#8212;</div><div class="lbl2">Vessel Sched maintained</div></div>
    <div class="boa-chip" style="background:#fdecea;border-color:#e6b8af;"><div class="num" id="maintChipVSchedNo" style="color:#C0392B;">&#8212;</div><div class="lbl2" style="color:#C0392B;">Vessel Sched NOT maintained</div></div>
  </div>

  <!-- Per-Port Maintenance Rate (dedicated rate summary) -->
  <h3 style="margin:22px 0 8px;color:#1F4E79;">&#128202; Per-Port Maintenance Rate
    <span style="font-size:11px;font-weight:400;margin-left:8px;color:#8a9bb0;">每个港口 Port Log / Vessel Schedule 维护率（当前筛选下，点表头可排序）</span>
  </h3>
  <p style="font-size:11px;color:#8a9bb0;margin:0 0 6px;">Port Log Rate = 已维护 Port Log 的 calls ÷ 该港总 calls；Vessel Sched Rate = Maintain timely 的 calls ÷ 该港总 calls。绿 ≥80%，橙 ≥50%，红 &lt;50%。</p>
  <div class="table-wrap" style="max-height:520px;overflow-y:auto;">
    <table id="maintRateTable" style="min-width:auto;"><thead><tr id="maintRateThead"></tr></thead><tbody id="maintRateTbody"></tbody></table>
  </div>

  <!-- By Port -->
  <h3 style="margin:20px 0 8px;color:#1F4E79;">&#128205; Maintenance by Port
    <span style="font-size:11px;font-weight:400;margin-left:10px;white-space:nowrap;">
      <button onclick="expandAllMaintPort()" style="font-size:11px;padding:2px 8px;cursor:pointer;border:1px solid #ccc;border-radius:4px;background:#fff;">Expand all</button>
      <button onclick="collapseAllMaintPort()" style="font-size:11px;padding:2px 8px;cursor:pointer;border:1px solid #ccc;border-radius:4px;background:#fff;margin-left:4px;">Collapse all</button>
    </span>
  </h3>
  <p style="font-size:11px;color:#8a9bb0;margin:0 0 6px;">Click a port row to expand <b>all calls</b> for that port, each with its Port Log (Y/N) and Vessel Schedule (Maintain timely / Not maintained) status. Red = not maintained.</p>
  <div class="table-wrap">
    <table id="maintPortTable" style="min-width:auto;"><thead><tr id="maintPortThead"></tr></thead><tbody id="maintPortTbody"></tbody></table>
  </div>

  <!-- Unmaintained Details: Port Log and Vessel Schedule separated -->
  <h3 style="margin:24px 0 4px;color:#C0392B;">&#128308; Port Log 未维护明细 <span id="maintPlogNoCount" style="font-size:11px;font-weight:400;color:#C0392B;">&#8212;</span>
    <span style="font-size:11px;font-weight:400;margin-left:8px;color:#8a9bb0;">Port Log Y/N = N 的船期（未维护 Port Log）</span>
    <span style="font-size:11px;font-weight:400;margin-left:8px;white-space:nowrap;">
      <input id="maintPlogNoSearch" type="search" placeholder="Filter: port / vessel / voyage / operator…" oninput="renderMaintUnmaintained()" style="font-size:11px;padding:3px 8px;border:1px solid #ccc;border-radius:4px;width:280px;">
    </span>
  </h3>
  <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
    <table id="maintPlogNoTable" style="min-width:auto;"><thead><tr id="maintPlogNoThead"></tr></thead><tbody id="maintPlogNoTbody"></tbody></table>
  </div>

  <h3 style="margin:24px 0 4px;color:#C0392B;">&#128308; Vessel Schedule 未维护明细 <span id="maintVsNoCount" style="font-size:11px;font-weight:400;color:#C0392B;">&#8212;</span>
    <span style="font-size:11px;font-weight:400;margin-left:8px;color:#8a9bb0;">Vessel Schedule Maintain Status = Not maintained 的船期（未维护船期）</span>
    <span style="font-size:11px;font-weight:400;margin-left:8px;white-space:nowrap;">
      <input id="maintVsNoSearch" type="search" placeholder="Filter: port / vessel / voyage / operator…" oninput="renderMaintUnmaintained()" style="font-size:11px;padding:3px 8px;border:1px solid #ccc;border-radius:4px;width:280px;">
    </span>
  </h3>
  <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
    <table id="maintVsNoTable" style="min-width:auto;"><thead><tr id="maintVsNoThead"></tr></thead><tbody id="maintVsNoTbody"></tbody></table>
  </div>

  <!-- Monthly Trend -->
  <div style="display:flex;flex-wrap:wrap;gap:16px;margin-top:8px;">
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h4 style="margin:6px 0 10px;color:#1F4E79;font-size:13px;text-align:center;">&#128200; Monthly Maintenance Trend</h4>
      <p style="font-size:10px;color:#8a9bb0;margin:0 0 8px;">Port Log rate (blue) &amp; Vessel Schedule rate (orange) by ETD month, within current Operator/ETD filters.</p>
      <div id="maintMonthBars" style="display:flex;flex-direction:column;gap:10px;"></div>
    </div>
    <div class="table-wrap" style="flex:1;min-width:380px;">
      <h4 style="margin:6px 0 10px;color:#1F4E79;font-size:13px;text-align:center;">&#128202; Monthly Rates</h4>
      <table id="maintMonthTable" style="min-width:auto;"><thead><tr id="maintMonthThead"></tr></thead><tbody id="maintMonthTbody"></tbody></table>
    </div>
  </div>
</div>

<!-- ── Export Tables Modal (multi-select → one workbook) ──────────────── -->
<div class="modal-overlay" id="exportTablesModal" onclick="if(event.target===this)closeExportTablesModal()">
  <div class="modal" style="max-width:480px;">
    <div class="modal-header">
      <h3>&#128196; Export Tables to Excel</h3>
      <button class="modal-close" onclick="closeExportTablesModal()">&#10005;</button>
    </div>
    <div class="modal-body">
      <p style="font-size:11px;color:#8a9bb0;margin:0 0 10px;">勾选需要导出的表；未勾选的不导出。所有勾选的表将写入同一个 Excel 文件（每张表一个工作表）。导出内容反映当前筛选 / 列可见性。</p>
      <div id="exportTableChecks" style="max-height:52vh;overflow:auto;border:1px solid #e3e9f0;border-radius:6px;padding:6px 10px;"></div>
      <div style="display:flex;gap:8px;margin-top:14px;align-items:center;">
        <button class="btn" onclick="exportSelToggleAll(true)" style="font-size:12px;padding:4px 10px;">&#10003; All</button>
        <button class="btn" onclick="exportSelToggleAll(false)" style="font-size:12px;padding:4px 10px;">&#10007; Clear</button>
        <span style="flex:1;"></span>
        <button class="btn" onclick="closeExportTablesModal()">Cancel</button>
        <button class="btn btn-export" onclick="exportSelectedTables()">&#8595; Export Selected</button>
      </div>
    </div>
  </div>
</div>

<!-- ── Lane Order Modal (reorder Full Schedule lanes) ─────────────────── -->
<div class="modal-overlay" id="laneOrderModal" onclick="if(event.target===this)closeLaneOrderModal()">
  <div class="modal" style="max-width:420px;">
    <div class="modal-header">
      <h3>&#8644; Lane Display Order</h3>
      <button class="modal-close" onclick="closeLaneOrderModal()">&#10005;</button>
    </div>
    <div class="modal-body">
      <p style="font-size:11px;color:#8a9bb0;margin:0 0 10px;">调整列表中 lane 的上下顺序，Full Schedule 即按此顺序展示。点击 ▲/▼ 移动，顺序自动保存（刷新后保留）。</p>
      <div id="laneOrderList" style="max-height:54vh;overflow:auto;border:1px solid #e3e9f0;border-radius:6px;padding:4px 6px;"></div>
      <div style="display:flex;gap:8px;margin-top:14px;align-items:center;">
        <button class="btn" onclick="resetLaneOrder()" style="font-size:12px;padding:4px 10px;">&#8634; Reset to default</button>
        <span style="flex:1;"></span>
        <button class="btn btn-export" onclick="closeLaneOrderModal()">Done</button>
      </div>
    </div>
  </div>
</div>

<div class="footer">CUL Daily Movement Dashboard &nbsp;|&nbsp; Data from CUL DAILY MOVEMENT.xlsx &nbsp;|&nbsp; <span id="footerTs"></span></div>

<!-- ── JavaScript ───────────────────────────────────────────────────────── -->
<script>
const SNAPSHOTS = {};
const TODAY_DATA = __TODAY_DATA__;
const COLUMN_DEFS_SUMMARY = __COLUMN_DEFS_SUMMARY__;
const COLUMN_DEFS_FULL    = __COLUMN_DEFS_FULL__;
const MAINT_DATA          = __MAINT_DATA__;
const AGENT_BY_PORT       = __AGENT_BY_PORT__;

// Default route display order (user-specified 2026-08-05). Unknown routes sort to the end.
// Expanded from combined tokens: NP2-REX -> NP2,REX | RES-CGX -> RES,CGX | CGS-AEM-IMR -> CGS,AEM,IMR
var ROUTE_ORDER = ['ST3','NSCT1','HDT','NSX','CST','CCT','NP2','REX','RTS','SGX','RES','CGX','HLX','CGS','AEM','IMR','NAX','JPS','SJA'];

// ── User-adjustable Lane display order ──────────────────────────────────
// Full Schedule (and the route filter dropdown) sort lanes by routeOrderKey.
// The user can override the order via the "Lane Order" modal; the override is
// persisted to localStorage so it survives page reloads.
var LANE_ORDER_KEY = 'cul_movement_laneorder';
var LANE_ORDER_OVERRIDE = null;  // array of route codes, or null (= use ROUTE_ORDER)

function loadLaneOrder(){
  try {
    var raw = localStorage.getItem(LANE_ORDER_KEY);
    if(raw){
      var arr = JSON.parse(raw);
      if(Array.isArray(arr) && arr.length){ LANE_ORDER_OVERRIDE = arr; return; }
    }
  } catch(e){}
  LANE_ORDER_OVERRIDE = null;
}

// Effective lane order: override (if set) merged with ROUTE_ORDER.
function getLaneOrder(){
  return (LANE_ORDER_OVERRIDE && LANE_ORDER_OVERRIDE.length) ? LANE_ORDER_OVERRIDE.slice() : ROUTE_ORDER.slice();
}

function saveLaneOrder(){
  try {
    if(LANE_ORDER_OVERRIDE && LANE_ORDER_OVERRIDE.length) localStorage.setItem(LANE_ORDER_KEY, JSON.stringify(LANE_ORDER_OVERRIDE));
    else localStorage.removeItem(LANE_ORDER_KEY);
  } catch(e){}
}

function routeOrderKey(r){
  var order = getLaneOrder();
  var i = order.indexOf(r);
  return i<0 ? 9999 : i;
}

// Visible columns state (key = viewId '1' or '2', value = Set of colKeys)
var visibleCols = {
  '1': new Set(COLUMN_DEFS_SUMMARY.filter(c=>c.defaultVisible).map(c=>c.key)),
  '2': new Set(COLUMN_DEFS_FULL.filter(c=>c.defaultVisible).map(c=>c.key)),
};

// Persist column-visibility selections across page reloads (localStorage).
// We store, per view, the full set of column keys ("all") plus the subset that was
// explicitly hidden ("hidden"). On load:
//   - a column NOT in "all"  -> brand-new since save -> follow its defaultVisible flag
//   - a column in "all" & in "hidden" -> keep hidden (user's explicit choice)
//   - a column in "all" & not in "hidden" -> keep visible (covers both default-visible
//     and a default-hidden column the user explicitly turned on)
var COLVIS_KEY = 'cul_movement_cols';
function loadColVisibility(){
  try{
    var raw = localStorage.getItem(COLVIS_KEY);
    if(!raw) return;
    var saved = JSON.parse(raw); // {'1':{all:[...],hidden:[...]}, '2':{...}}
    ['1','2'].forEach(function(vid){
      var rec = saved[vid];
      if(!rec || !Array.isArray(rec.all)) return;
      var defs = vid==='1' ? COLUMN_DEFS_SUMMARY : COLUMN_DEFS_FULL;
      var allSet = new Set(rec.all);
      var hiddenSet = new Set(rec.hidden || []);
      var merged = new Set();
      defs.forEach(function(c){
        if(!allSet.has(c.key)) merged.add(c.key);   // new column -> default visible
        else if(hiddenSet.has(c.key)) {}            // explicitly hidden -> keep hidden
        else merged.add(c.key);                     // was visible -> keep visible
      });
      visibleCols[vid] = merged;
    });
  }catch(e){}
}
function saveColVisibility(){
  try{
    var out = {'1':{all:[],hidden:[]}, '2':{all:[],hidden:[]}};
    ['1','2'].forEach(function(vid){
      var defs = vid==='1' ? COLUMN_DEFS_SUMMARY : COLUMN_DEFS_FULL;
      defs.forEach(function(c){
        out[vid].all.push(c.key);
        if(!visibleCols[vid].has(c.key)) out[vid].hidden.push(c.key);
      });
    });
    localStorage.setItem(COLVIS_KEY, JSON.stringify(out));
  }catch(e){}
}
loadColVisibility();

// Multi-select filter state: {route1: Set, pic1: Set, route2: Set, pic2: Set}
// Default: null means "show all" (empty Set also means show all, used after init)
var filterSelections = {route1: null, vessel1: null, pic1: null, route2: null, vessel2: null, pic2: null};

function setAllFilterOptions(type, viewId){
  var key = type + viewId;
  // Build set of all possible values
  var allValues;
  if(viewId==='1'){
    allValues = new Set(summaryData.map(function(r){return r[type];}));
  } else {
    allValues = new Set(fullData.map(function(r){return r[type];}));
  }
  filterSelections[key] = new Set(allValues);
}

function getFilterSelected(type, viewId){
  var key = type + viewId;
  var sel = filterSelections[key];
  if(sel == null || sel.size === 0) return null;             // null / undefined / empty = show all
  return sel;
}

function buildFilterDropdown(type, viewId){
  var dd = document.getElementById('filterDropdown'+type.charAt(0).toUpperCase()+type.slice(1)+viewId);
  dd.innerHTML = '';

  // Get all unique values
  var dataArr = viewId==='1' ? summaryData : fullData;
  var values = [...new Set(dataArr.map(function(r){return r[type];}))];
  if(type==='route'){
    values.sort(function(a,b){return routeOrderKey(a)-routeOrderKey(b);});
  } else {
    values.sort();
  }

  // Select All / Clear All actions
  var actions = document.createElement('div');
  actions.className = 'filter-actions';
  var selAll = document.createElement('button');
  selAll.textContent = 'Select All';
  selAll.onclick = function(e){ e.stopPropagation(); setAllFilterOptions(type, viewId); buildFilterDropdown(type, viewId); updateFilterButton(type, viewId); rerender(viewId); };
  var clrAll = document.createElement('button');
  clrAll.textContent = 'None';
  clrAll.onclick = function(e){ e.stopPropagation(); filterSelections[type+viewId]=new Set(['\x00NONE\x00']); buildFilterDropdown(type, viewId); updateFilterButton(type, viewId); rerender(viewId); };
  actions.appendChild(selAll);
  actions.appendChild(clrAll);
  dd.appendChild(actions);

  var sel = getFilterSelected(type, viewId);
  var isAll = sel===null;
  values.forEach(function(v){
    var label = document.createElement('label');
    var cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.checked = isAll || sel.has(v);
    cb.onchange = function(e){
      e.stopPropagation();
      if(!filterSelections[type+viewId]) filterSelections[type+viewId] = new Set();
      if(isAll){
        // Transition from "all" to explicit set
        filterSelections[type+viewId] = new Set(values);
        isAll = false;
        sel = filterSelections[type+viewId];
      }
      if(this.checked) filterSelections[type+viewId].add(v);
      else filterSelections[type+viewId].delete(v);
      // If all are deselected, show nothing (not "show all")
      if(filterSelections[type+viewId].size===0) filterSelections[type+viewId]=new Set(['\x00NONE\x00']);
      updateFilterButton(type, viewId);
      rerender(viewId);
    };
    label.appendChild(cb);
    label.appendChild(document.createTextNode(' '+v));
    dd.appendChild(label);
  });
}

function updateFilterButton(type, viewId){
  var btn = document.getElementById('filter'+type.charAt(0).toUpperCase()+type.slice(1)+'Btn'+viewId);
  var key = type + viewId;
  var fs = filterSelections[key];
  var dataArr = viewId==='1' ? summaryData : fullData;
  var totalValues = new Set(dataArr.map(function(r){return r[type];})).size;
  var isNone = fs && fs.has('\x00NONE\x00');
  if(fs===null || (!isNone && fs.size===totalValues)){
    btn.textContent = type==='route' ? 'All Routes' : type==='vessel' ? 'All Vessels' : 'All PIC';
    btn.classList.remove('has-selection');
  } else if(isNone){
    btn.textContent = 'None';
    btn.classList.add('has-selection');
  } else {
    btn.textContent = fs.size + ' selected';
    btn.classList.add('has-selection');
  }
}

function toggleFilterDropdown(type, viewId){
  var ddId = 'filterDropdown'+type.charAt(0).toUpperCase()+type.slice(1)+viewId;
  var dd = document.getElementById(ddId);
  // Rebuild to ensure checkboxes reflect current state
  buildFilterDropdown(type, viewId);
  var isOpen = dd.classList.contains('open');
  document.querySelectorAll('.col-dropdown,.filter-dropdown').forEach(function(d){d.classList.remove('open');});
  if(!isOpen) dd.classList.add('open');
}

function rerender(viewId){
  if(viewId==='1') renderSummary();
  else renderFullSchedule();
}

function loadSnapshots(){try{const s=localStorage.getItem('cul_movement_history');if(s)Object.assign(SNAPSHOTS,JSON.parse(s));}catch(e){}}
function saveSnapshot(d){SNAPSHOTS[d.date]=d.vessels;try{localStorage.setItem('cul_movement_history',JSON.stringify(SNAPSHOTS));}catch(e){}}

/* ── Column toggle dropdown ──────────────────────────────────────────── */
function buildColDropdown(viewId){
  var defs = viewId==='1' ? COLUMN_DEFS_SUMMARY : COLUMN_DEFS_FULL;
  var dd = document.getElementById('colDropdown'+viewId);
  dd.innerHTML = '';
  defs.forEach(function(col){
    var label = document.createElement('label');
    var cb = document.createElement('input');
    cb.type = 'checkbox';
    cb.checked = visibleCols[viewId].has(col.key);
    cb.onchange = function(){
      if(this.checked) visibleCols[viewId].add(col.key);
      else visibleCols[viewId].delete(col.key);
      saveColVisibility();
      if(viewId==='1') renderSummary();
      else renderFullSchedule();
    };
    label.appendChild(cb);
    label.appendChild(document.createTextNode(' '+col.label));
    dd.appendChild(label);
  });
}
function toggleColDropdown(viewId){
  var dd = document.getElementById('colDropdown'+viewId);
  var isOpen = dd.classList.contains('open');
  // Close all dropdowns first
  document.querySelectorAll('.col-dropdown').forEach(function(d){d.classList.remove('open');});
  if(!isOpen) dd.classList.add('open');
}
// Close dropdowns when clicking outside
document.addEventListener('click', function(e){
  if(!e.target.closest('.col-toggle-btn') && !e.target.closest('.col-dropdown') && !e.target.closest('.filter-btn') && !e.target.closest('.filter-dropdown')){
    document.querySelectorAll('.col-dropdown,.filter-dropdown').forEach(function(d){d.classList.remove('open');});
  }
});

/* ── Build table header row (respecting visibleCols) ───────────────── */
function buildTableHeader(viewId, sortState){
  // sortState = {col: idx, dir: 1/-1} or null
  var defs = viewId==='1' ? COLUMN_DEFS_SUMMARY : COLUMN_DEFS_FULL;
  var html = '';
  defs.forEach(function(col, idx){
    if(!visibleCols[viewId].has(col.key)) return;
    var cls = '';
    if(sortState && sortState.col===idx){
      cls = sortState.dir===1 ? 'sort-asc' : 'sort-desc';
    }
    // Index column (#) is not sortable
    if(col.key==='_idx'){
      html += '<th style="width:40px">'+col.label+'</th>';
    } else {
      var onclick = 'onclick="sort' + (viewId==='1'?'Summary':'Full') + '(' + idx + ')"';
      html += '<th class="'+cls+'" '+onclick+'>'+col.label+'<span class="sort-arrow"></span></th>';
    }
  });
  return html;
}

/* ── Tab switching ─────────────────────────────────────────────────────── */
function switchTab(viewId, btn){
  document.querySelectorAll('.tab-content').forEach(el=>el.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(el=>el.classList.remove('active'));
  document.getElementById(viewId).classList.add('active');
  btn.classList.add('active');
  if(viewId==='fullScheduleView'){
    document.getElementById('btnExport').textContent='\u2193 Export Full Schedule';
    document.getElementById('btnExport').style.display='';
  } else if(viewId==='summaryView'){
    document.getElementById('btnExport').textContent='\u2193 Export Excel';
    document.getElementById('btnExport').style.display='';
  } else {
    document.getElementById('btnExport').style.display='none';
  }
  // Close any open column/filter dropdown
  document.querySelectorAll('.col-dropdown,.filter-dropdown').forEach(function(d){d.classList.remove('open');});
  updateCtrlH();
}

/* ═══════════════════════════════════════════════════════════════════
   DECOMMISSIONED VESSELS (已下线)
   源 Excel 中已退租/下线的船，船名带 "-已下线" 后缀。
   Summary 与 Full Schedule 默认隐藏这些船（含其筛选下拉、导出、统计数）；
   需要查看历史时勾选工具栏的 "Show 已下线" 即可临时显示。
   ═══════════════════════════════════════════════════════════════════ */
var DECOM_MARK = '已下线';
var SHOW_DECOM = false;
try{ SHOW_DECOM = sessionStorage.getItem('showDecom')==='1'; }catch(e){ SHOW_DECOM=false; }

function isDecommissioned(v){ return typeof v==='string' && v.indexOf(DECOM_MARK)>=0; }

/* 按当前 SHOW_DECOM 状态重建 Summary / Full Schedule 的数据源 */
function applyDecomFilter(){
  summaryData = (TODAY_DATA.vessels||[]).filter(function(r){ return SHOW_DECOM || !isDecommissioned(r.vessel); });
  fullData    = (TODAY_DATA.fullSchedule||[]).filter(function(r){ return SHOW_DECOM || !isDecommissioned(r.vessel); });
}

/* 工具栏 "Show 已下线" 勾选框 —— 两个视图共用，状态存 sessionStorage */
function onDecomToggle(cb){
  SHOW_DECOM = !!(cb && cb.checked);
  try{ sessionStorage.setItem('showDecom', SHOW_DECOM?'1':'0'); }catch(e){}
  applyDecomFilter();
  // 重建 vessel 分组底色（Full Schedule 交替色带）
  let seen={}, gi=0;
  fullData.forEach(function(r){ if(!(r.vessel in seen)){ seen[r.vessel]=gi%2; gi++; } });
  vesselGroupMap = seen;
  // 同步两个工具栏的勾选框 + 视觉指示
  var boxes = document.querySelectorAll('.decom-cb');
  var onFlag = SHOW_DECOM ? 'on' : '';
  var dotText = SHOW_DECOM ? '✓' : '+';
  document.querySelectorAll('.decom-toggle').forEach(function(lbl){
    lbl.classList.toggle('on', SHOW_DECOM);
    var d=lbl.querySelector('.decom-dot'); if(d) d.textContent=dotText;
  });
  boxes.forEach(function(b){ if(b!==cb) b.checked = SHOW_DECOM; });
  ['route','vessel','pic'].forEach(function(t){
    updateFilterButton(t,'1'); updateFilterButton(t,'2');
  });
  renderSummary();
  renderFullSchedule();
}

/* ═══════════════════════════════════════════════════════════════════
   SUMMARY VIEW
   ═══════════════════════════════════════════════════════════════════ */
let summaryData=[], summarySortCol=-1, summarySortDir=1;

function initSummary(){
  applyDecomFilter();
  buildColDropdown('1');
  updateFilterButton('route', '1');
  updateFilterButton('vessel', '1');
  updateFilterButton('pic', '1');
  renderSummary();
}

function getFilteredSummary(){
  const q=document.getElementById('searchBox').value.toLowerCase();
  const selRoute = getFilterSelected('route', '1');
  const selVessel = getFilterSelected('vessel', '1');
  const selPic = getFilterSelected('pic', '1');
  const delay=document.getElementById('filterDelay').value;
  let data=summaryData.filter(r=>{
    if(selRoute && !selRoute.has(r.route)) return false;
    if(selVessel && !selVessel.has(r.vessel)) return false;
    if(selPic && !selPic.has(r.pic)) return false;
    if(q && !`${r.route} ${r.vessel} ${r.port} ${r.wait} ${r.pic} ${r.code} ${r.voy} ${r.remark}`.toLowerCase().includes(q)) return false;
    if(delay==='delay') return r.etaDelay.toLowerCase().includes('delay');
    if(delay==='ahead') return r.etaDelay.toLowerCase().includes('ahead');
    if(delay==='normal') return !r.etaDelay;
    return true;
  });
  if(summarySortCol>=0){
    const defs = COLUMN_DEFS_SUMMARY;
    const colDef = defs[summarySortCol];
    if(colDef && colDef.key!=='_idx' && visibleCols['1'].has(colDef.key)){
      if(colDef.key==='route'){
        data.sort((a,b)=>(routeOrderKey(a.route)-routeOrderKey(b.route))*summarySortDir);
      } else {
        data.sort((a,b)=>((a[colDef.key]||'').localeCompare(b[colDef.key]||''))*summarySortDir);
      }
    }
  } else {
    data.sort((a,b)=>routeOrderKey(a.route)-routeOrderKey(b.route));
  }
  return data;
}

function delayTag(v){
  if(!v) return '';
  if(v.toLowerCase().startsWith('ahead')) return '<span class="ahead-tag">'+v+'</span>';
  if(v.toLowerCase().startsWith('delay')) return '<span class="delay-tag">'+v+'</span>';
  return v;
}

function dateWithWeekday(dateStr){
  if(dateStr == null || dateStr === '') return '';
  var s = String(dateStr).trim();
  // Excel serial number (e.g. 46210.2291666667)
  var num = parseFloat(s);
  if(!isNaN(num) && num > 40000 && num < 100000) {
    var d = new Date((num - 25569) * 86400000);
    var days = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
    return '<span class="weekday-tag">' + days[d.getUTCDay()] + '</span>';
  }
  // Format: "07/14 07:00"
  var m = s.match(/(\d{1,2})\/(\d{1,2})\s+(\d{2}:\d{2})/);
  if(m) {
    var d = new Date(new Date().getFullYear(), parseInt(m[1])-1, parseInt(m[2]), parseInt(m[3].split(':')[0]), parseInt(m[3].split(':')[1]));
    var days = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'];
    return '<span class="weekday-tag">' + days[d.getDay()] + '</span>';
  }
  // Already weekday name: "Mon", "Thu", "TUE0800"
  var wd = s.match(/^(MON|TUE|WED|THU|FRI|SAT|SUN)/i);
  if(wd) return '<span class="weekday-tag">' + wd[1].toUpperCase() + '</span>';
  return '';
}

function renderSummary(){
  const data=getFilteredSummary();
  // Build header
  document.getElementById('summaryThead').innerHTML =
    buildTableHeader('1', {col: summarySortCol, dir: summarySortDir});

  const tbody=document.getElementById('summaryTbody');
  if(!data.length){tbody.innerHTML='<tr><td colspan="'+visibleCols['1'].size+'" class="no-data">No matching records found.</td></tr>';document.getElementById('statTotal').textContent='0 vessels';document.getElementById('statDelay').textContent='0 delayed';return;}
  const dc=data.filter(r=>r.etaDelay && r.etaDelay.toLowerCase().includes('delay'));
  document.getElementById('statTotal').textContent=new Set(data.map(r=>r.vessel)).size+' vessels';
  document.getElementById('statDelay').textContent=new Set(dc.map(r=>r.vessel)).size+' delayed';

  const defs = COLUMN_DEFS_SUMMARY;
  tbody.innerHTML = data.map(function(r, rowIdx){
    var cells = defs.map(function(col){
      if(!visibleCols['1'].has(col.key)) return null;
      // Row number column
      if(col.key==='_idx') return '<td class="td-center td-idx">'+(rowIdx+1)+'</td>';
      var v = r[col.key] || '';
      // Special formatting per column
      if(col.key==='route')     return '<td class="td-center"><span class="badge-route">'+v+'</span></td>';
      if(col.key==='vessel')    return '<td><strong>'+v+'</strong></td>';
      if(col.key==='code')      return '<td class="td-center"><span class="badge-code">'+v+'</span></td>';
      if(col.key==='pic')       return '<td>'+v+'</td>';
      if(col.key==='port')      return '<td class="td-center td-mono"><strong>'+v+'</strong></td>';
      if(col.key==='wait')      return '<td class="td-center">'+v+'</td>';
      if(col.key==='proforma')  return '<td class="proforma-cell">'+v+'</td>';
      if(col.key==='voy')       return '<td class="td-center td-mono">'+v+'</td>';
      if(col.key==='ltmEta' || col.key==='ltmEtd') return '<td class="td-center td-mono" style="font-size:11px">'+v+'</td>';
      if(col.key==='eta' || col.key==='etb' || col.key==='etd') return '<td class="td-center td-mono">'+v+'</td>';
      if(col.key==='portStay')  return '<td class="td-center">'+v+'</td>';
      if(col.key==='etaDelay')  return '<td class="td-center">'+delayTag(v)+'</td>';
      if(col.key==='etdDelay')  return '<td class="td-center">'+delayTag(v)+'</td>';
      return '<td>'+v+'</td>';
    });
    return '<tr>'+cells.filter(c=>c!==null).join('')+'</tr>';
  }).join('');
}

function sortSummary(col){
  const defs = COLUMN_DEFS_SUMMARY;
  if(defs[col].key==='_idx') return;
  if(!visibleCols['1'].has(defs[col].key)) return;
  if(summarySortCol===col){ summarySortDir*=-1; } else { summarySortCol=col; summarySortDir=1; }
  renderSummary();
}

/* ═══════════════════════════════════════════════════════════════════
   FULL SCHEDULE VIEW
   ═══════════════════════════════════════════════════════════════════ */
let fullData=[], fullSortCol=-1, fullSortDir=1;
let vesselGroupMap = {};

function initFullSchedule(){
  applyDecomFilter();
  loadLaneOrder();
  let seen = {}, gi = 0;
  fullData.forEach(function(r){
    if(!(r.vessel in seen)){ seen[r.vessel] = gi%2; gi++; }
  });
  vesselGroupMap = seen;

  // Set default ETA date range
  document.getElementById('etaFrom2').value = TODAY_DATA.defaultEtaFrom || '';
  document.getElementById('etaTo2').value = TODAY_DATA.defaultEtaTo || '';
  document.getElementById('dateFilterType2').value = 'eta';

  buildColDropdown('2');
  updateFilterButton('route', '2');
  updateFilterButton('vessel', '2');
  updateFilterButton('pic', '2');
  renderFullSchedule();
}

function onDateFilterTypeChange(){
  const mode = document.getElementById('dateFilterType2').value;
  const key = mode === 'etb' ? 'defaultEtbFrom' : 'defaultEtaFrom';
  const keyTo = mode === 'etb' ? 'defaultEtbTo' : 'defaultEtaTo';
  document.getElementById('etaFrom2').value = TODAY_DATA[key] || '';
  document.getElementById('etaTo2').value = TODAY_DATA[keyTo] || '';
  renderFullSchedule();
}

function getFilteredFull(){
  const q=document.getElementById('searchBox2').value.toLowerCase();
  const selRoute = getFilterSelected('route', '2');
  const selVessel = getFilterSelected('vessel', '2');
  const selPic = getFilterSelected('pic', '2');
  const etaFrom = document.getElementById('etaFrom2').value;
  const etaTo   = document.getElementById('etaTo2').value;
  const dateMode = document.getElementById('dateFilterType2').value;
  const dateKey = dateMode === 'etb' ? 'etbRaw' : 'etaRaw';
  let data=fullData.filter(r=>{
    if(selRoute && !selRoute.has(r.route)) return false;
    if(selVessel && !selVessel.has(r.vessel)) return false;
    if(selPic && !selPic.has(r.pic)) return false;
    if(etaFrom && r[dateKey] < etaFrom) return false;
    if(etaTo   && r[dateKey] > etaTo)   return false;
    if(q && !`${r.route} ${r.vessel} ${r.port} ${r.wait} ${r.manIn} ${r.pic} ${r.code} ${r.voy} ${r.date} ${r.remark}`.toLowerCase().includes(q)) return false;
    return true;
  });
  if(fullSortCol>=0){
    const defs = COLUMN_DEFS_FULL;
    const colDef = defs[fullSortCol];
    if(colDef && visibleCols['2'].has(colDef.key)){
      if(colDef.key==='route'){
        data.sort((a,b)=>(routeOrderKey(a.route)-routeOrderKey(b.route))*fullSortDir);
      } else {
        data.sort((a,b)=>((a[colDef.key]||'').localeCompare(b[colDef.key]||''))*fullSortDir);
      }
    }
  } else {
    data.sort((a,b)=>routeOrderKey(a.route)-routeOrderKey(b.route));
  }
  return data;
}

function renderFullSchedule(){
  const data=getFilteredFull();
  // Build header
  document.getElementById('fullThead').innerHTML =
    buildTableHeader('2', {col: fullSortCol, dir: fullSortDir});

  const tbody=document.getElementById('fullTbody');
  if(!data.length){tbody.innerHTML='<tr><td colspan="'+visibleCols['2'].size+'" class="no-data">No matching records found.</td></tr>';document.getElementById('statTotal2').textContent='0 rows';return;}
  document.getElementById('statTotal2').textContent=data.length+' rows';

  const defs = COLUMN_DEFS_FULL;
  let prevVessel = '';
  tbody.innerHTML = data.map(function(r, idx){
    var gc = vesselGroupMap[r.vessel]===0 ? 'vessel-group-even' : 'vessel-group-odd';
    var boundaryCls = '';
    if(r.vessel !== prevVessel){
      boundaryCls = ' vessel-group-first';
      if(idx>0) boundaryCls += ' vessel-group-last-prev';
      prevVessel = r.vessel;
    }
    var endCls = '';
    if(idx===data.length-1 || data[idx+1].vessel !== r.vessel){
      endCls = ' vessel-group-last';
    }
    var cls = gc+boundaryCls+endCls;
    if(r.isSummary) cls += ' summary-highlight';

    var cells = defs.map(function(col){
      if(!visibleCols['2'].has(col.key)) return null;
      var v = r[col.key] || '';
      if(col.key==='route')      return '<td class="td-center"><span class="badge-route">'+v+'</span></td>';
      if(col.key==='vessel')     return '<td><span class="vessel-label">'+v+'</span></td>';
      if(col.key==='code')       return '<td class="td-center"><span class="badge-code">'+v+'</span></td>';
      if(col.key==='pic')        return '<td>'+v+'</td>';
      if(col.key==='port')       return '<td class="td-center td-mono"><strong>'+v+'</strong></td>';
      if(col.key==='manIn' || col.key==='wait' || col.key==='run' || col.key==='fspDistance' || col.key==='speed') return '<td class="td-center">'+v+'</td>';
      if(col.key==='proforma')   return '<td class="proforma-cell">'+v+'</td>';
      if(col.key==='voy')        return '<td class="td-center td-mono">'+v+'</td>';
      if(col.key==='ltmEta' || col.key==='ltmEtd') return '<td class="td-center td-mono" style="font-size:11px">'+v+'</td>';
      if(col.key==='date') return '<td class="td-center td-mono" style="font-size:11px">'+dateWithWeekday(r.eta)+'</td>';
      if(col.key==='eta' || col.key==='etb' || col.key==='etd') return '<td class="td-center td-mono">'+v+'</td>';
      if(col.key==='portStay')   return '<td class="td-center">'+v+'</td>';
      if(col.key==='etaDelay')   return '<td class="td-center">'+delayTag(v)+'</td>';
      if(col.key==='etdDelay')   return '<td class="td-center">'+delayTag(v)+'</td>';
      return '<td>'+v+'</td>';
    });
    return '<tr class="'+cls+'">'+cells.filter(c=>c!==null).join('')+'</tr>';
  }).join('');
}

function sortFull(col){
  const defs = COLUMN_DEFS_FULL;
  if(!visibleCols['2'].has(defs[col].key)) return;
  if(fullSortCol===col){ fullSortDir*=-1; } else { fullSortCol=col; fullSortDir=1; }
  renderFullSchedule();
}

/* ═══════════════════════════════════════════════════════════════════
   EXPORT EXCEL (matches Summary Excel format exactly)
   ═══════════════════════════════════════════════════════════════════ */
function exportExcel(){
  var activeTab = document.querySelector('.tab-btn.active').getAttribute('data-tab');
  _loadXlsx(function(){
    if(activeTab==='fullScheduleView'){
      exportFullScheduleExcel();
    } else {
      exportSummaryExcel();
    }
  });
}

function exportSummaryExcel(wb){
  wb = wb || null;
  var newWb = !wb;
  if(newWb) wb = XLSX.utils.book_new();
  var data=getFilteredSummary(), todayStr=TODAY_DATA.date;
  var headers = COLUMN_DEFS_SUMMARY.filter(c=>visibleCols['1'].has(c.key) && c.key!=='_idx').map(c=>c.label);

  function thinBorder(){var s={style:'thin',color:{rgb:'BFBFBF'}};return{top:s,bottom:s,left:s,right:s};}
  var B=thinBorder();
  function F(rgb){return{patternType:'solid',fgColor:{rgb:rgb}};}
  function A(h,v,wrap){var o={horizontal:h,vertical:v};if(wrap)o.wrapText=true;return o;}

  var tS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:14},fill:F('2E75B6'),alignment:A('center','center'),border:B};
  var hS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:10},fill:F('1F4E79'),alignment:A('center','center',true),border:B};

  var sheetData=[];
  var tr=[];
  var numCols = headers.length;
  for(var c=0;c<numCols;c++) tr[c]={v:(c===0?'CUL VESSEL DAILY MOVEMENT SUMMARY  --  As of '+todayStr:''),s:tS};
  sheetData.push(tr);
  sheetData.push(headers.map(function(h){return{v:h,s:hS};}));

  var defs = COLUMN_DEFS_SUMMARY;
  for(var i=0;i<data.length;i++){
    var r=data[i], fc=(i%2===0)?'EBF3FB':'FFFFFF';
    var nS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('left','center')};
    var bS={font:{name:'Arial',bold:true,sz:9},fill:F(fc),border:B,alignment:A('left','center')};
    var cS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('center','center')};
    var rS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('left','center',true)};
    var dS={font:{name:'Arial',bold:true,color:{rgb:'C00000'},sz:9},fill:F(fc),border:B,alignment:A('center','center')};
    function ds(v){if(!v)return cS;return(v.toLowerCase().indexOf('delay')>=0)?dS:cS;}
    var row = [];
    defs.forEach(function(col){
      if(!visibleCols['1'].has(col.key) || col.key==='_idx') return;
      var v = r[col.key]||'';
      if(col.key==='remark') row.push({v:v,s:rS});
      else if(col.key==='vessel') row.push({v:v,s:bS});
      else if(col.key==='etaDelay'||col.key==='etdDelay') row.push({v:v,s:ds(v)});
      else if(['route','port','voy','eta','etb','etd','portStay','proforma','ltmEta','ltmEtd','wait','date','run','fspDistance','speed','manIn'].includes(col.key)) row.push({v:v,s:cS});
      else row.push({v:v,s:nS});
    });
    sheetData.push(row);
  }

  var ws=XLSX.utils.aoa_to_sheet(sheetData);
  ws['!merges']=[{s:{r:0,c:0},e:{r:0,c:numCols-1}}];
  ws['!cols'] = headers.map(function(h){return{wch:h.length+4};});
  ws['!rows']=[{hpt:28},{hpt:30}];for(var j=0;j<data.length;j++)ws['!rows'].push({hpt:18});
  ws['!autofilter']={ref:'A2:'+XLSX.utils.encode_col(numCols-1)+(data.length+2)};

  var wbOut = wb;
  XLSX.utils.book_append_sheet(wbOut,ws,'Daily Movement Summary');
  if(newWb) XLSX.writeFile(wbOut,'CUL Daily Movement Summary '+todayStr+'.xlsx');
  return wbOut;
}

function exportFullScheduleExcel(wb){
  wb = wb || null;
  var newWb = !wb;
  if(newWb) wb = XLSX.utils.book_new();
  var data=getFilteredFull(), todayStr=TODAY_DATA.date;
  // Detail rows respect the UI column visibility: hidden columns are NOT exported.
  // The per-vessel TITLE row (a merged combined cell) always shows the 4 identity fields
  // (Lane/route, Vessel, Code, PIC) regardless of UI hiding.
  var exportKeys = new Set(visibleCols['2']);
  var headers = COLUMN_DEFS_FULL.filter(c=>exportKeys.has(c.key)).map(c=>c.label);

  function thinBorder(){var s={style:'thin',color:{rgb:'BFBFBF'}};return{top:s,bottom:s,left:s,right:s};}
  var B=thinBorder();
  function F(rgb){return{patternType:'solid',fgColor:{rgb:rgb}};}
  function A(h,v,wrap){var o={horizontal:h,vertical:v};if(wrap)o.wrapText=true;return o;}

  var tS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:14},fill:F('2E75B6'),alignment:A('center','center'),border:B};
  var hS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:10},fill:F('1F4E79'),alignment:A('center','center',true),border:B};
  var sepS={font:{name:'Arial',sz:6},fill:F('D9D9D9'),border:B};

  // Sort by route (custom order) then vessel for grouping (copy to avoid affecting UI sort state)
  var exportData = data.slice().sort(function(a,b){
    var dr = routeOrderKey(a.route)-routeOrderKey(b.route);
    if(dr!==0) return dr;
    return (a.vessel||'').localeCompare(b.vessel||'');
  });

  var sheetData=[];
  var numCols = headers.length;
  // Top global title
  var tr=[];
  for(var c=0;c<numCols;c++) tr[c]={v:(c===0?'CUL VESSEL FULL SCHEDULE  --  As of '+todayStr:''),s:tS};
  sheetData.push(tr);

  var defs = COLUMN_DEFS_FULL;
  // Map visible column keys to their exported indices
  var visKeys=[], keyToIdx={};
  defs.forEach(function(col){
    if(!exportKeys.has(col.key)) return;
    keyToIdx[col.key]=visKeys.length;
    visKeys.push(col.key);
  });

  // Per-vessel title row style (matches source file: route / vessel / code / pic)
  var vTitleS={font:{name:'Arial',bold:true,color:{rgb:'1F4E79'},sz:11},fill:F('DDEBF7'),border:B,alignment:A('left','center')};
  // Blank separator row inserted after each vessel's last schedule entry
  var blankS={font:{name:'Arial',sz:6},fill:F('FFFFFF'),border:B,alignment:A('left','center')};

  var prevVessel='';
  // Track merges so the top title and every per-vessel title row are merged across all columns
  var merges=[{s:{r:0,c:0},e:{r:0,c:numCols-1}}];
  for(var i=0;i<exportData.length;i++){
    var r=exportData[i];
    if(r.vessel!==prevVessel){
      // Per-vessel title row: always show the 4 identity fields (Lane/route, Vessel, Code, PIC)
      // as one merged cell, independent of UI column hiding (detail rows still respect hiding)
      var titleText='Lane: '+(r.route||'')+'    Vessel: '+(r.vessel||'')+'    Code: '+(r.code||'')+'    PIC: '+(r.pic||'');
      var titleRow=[];
      for(var c=0;c<numCols;c++) titleRow[c]={v:'',s:vTitleS};
      titleRow[0]={v:titleText,s:vTitleS};
      var tRow=sheetData.length;
      sheetData.push(titleRow);
      merges.push({s:{r:tRow,c:0},e:{r:tRow,c:numCols-1}});
      // Column header row
      sheetData.push(headers.map(function(h){return{v:h,s:hS};}));
      prevVessel=r.vessel;
    }
    var fc='FFFFFF';
    if(r.isSummary) fc='FFF3CD';
    var nS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('left','center')};
    var bS={font:{name:'Arial',bold:true,sz:9},fill:F(fc),border:B,alignment:A('left','center')};
    var cS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('center','center')};
    var dS={font:{name:'Arial',bold:true,color:{rgb:'C00000'},sz:9},fill:F(fc),border:B,alignment:A('center','center')};
    function ds(v){if(!v)return cS;return(v.toLowerCase().indexOf('delay')>=0)?dS:cS;}
    var row = [];
    defs.forEach(function(col){
      if(!exportKeys.has(col.key)) return;
      var v = r[col.key]||'';
      if(col.key==='vessel') row.push({v:v,s:bS});
      else if(col.key==='etaDelay'||col.key==='etdDelay') row.push({v:v,s:ds(v)});
      else if(['route','port','voy','eta','etb','etd','portStay','proforma','ltmEta','ltmEtd','wait','date','run','fspDistance','speed','manIn'].includes(col.key)) row.push({v:v,s:cS});
      else row.push({v:v,s:nS});
    });
    sheetData.push(row);
    // Blank separator row after the last schedule entry of the current vessel
    if(i+1>=exportData.length || exportData[i+1].vessel!==r.vessel){
      var blankRow=[];
      for(var c=0;c<numCols;c++) blankRow[c]={v:'',s:blankS};
      sheetData.push(blankRow);
    }
  }

  var totalRows = sheetData.length;
  var ws=XLSX.utils.aoa_to_sheet(sheetData);
  ws['!merges']=merges;
  ws['!cols'] = headers.map(function(h){return{wch:Math.max(h.length+4, 10)};});
  ws['!rows']=[{hpt:28}];
  for(var j=1;j<totalRows;j++){
    var firstCell = sheetData[j][0];
    if(firstCell && firstCell.s===hS) ws['!rows'].push({hpt:26});
    else if(firstCell && firstCell.s===vTitleS) ws['!rows'].push({hpt:22});
    else if(firstCell && firstCell.s===blankS) ws['!rows'].push({hpt:8});
    else ws['!rows'].push({hpt:16});
  }
  // No global autofilter because header rows are repeated per vessel

  var wbOut = wb;
  XLSX.utils.book_append_sheet(wbOut,ws,'Full Schedule');
  if(newWb) XLSX.writeFile(wbOut,'CUL Daily Movement Full Schedule '+todayStr+'.xlsx');
  return wbOut;
}

/* ═══════════════════════════════════════════════════════════════════
   MULTI-SELECT EXPORT  (choose tables → one workbook, one sheet each)
   ═══════════════════════════════════════════════════════════════════ */
// Each exportable table: id = DOM <table> id; sheet = Excel sheet name;
// rich = use the dedicated formatted builder; skipFirst = drop leading
// expander # column (Port Wait triangle).
var EXPORT_TABLES = [
  {group:'Summary', items:[
    {id:'summaryTable', label:'Daily Movement Summary', sheet:'Daily Movement Summary', rich:'summary'},
  ]},
  {group:'Full Schedule', items:[
    {id:'fullTable', label:'Full Schedule', sheet:'Full Schedule', rich:'full'},
  ]},
  {group:'Port Wait', items:[
    {id:'portWaitTable', label:'Port Wait Time Analysis', sheet:'Port Wait', skipFirst:true},
    {id:'portWaitRegionTable', label:'Wait by Port Region', sheet:'Wait by Region'},
    {id:'callCountLaneTable', label:'Port Call Count by Lane', sheet:'Call Count by Lane'},
    {id:'callCountRegionTable', label:'Port Call Count by Region', sheet:'Call Count by Region'},
  ]},
  {group:'BOA', items:[
    {id:'boaTblTrade', label:'BOA by Trade', sheet:'BOA by Trade'},
    {id:'boaTblLane', label:'BOA by Trade → Lane', sheet:'BOA by Lane'},
    {id:'boaTblRegion', label:'BOA by Port Region', sheet:'BOA by Region'},
    {id:'boaTblPort', label:'BOA by Port', sheet:'BOA by Port'},
  ]},
  {group:'Speed', items:[
    {id:'speedTable', label:'Vessel Speed', sheet:'Vessel Speed'},
  ]},
  {group:'Maintenance', items:[
    {id:'maintPortTable', label:'Maintenance by Port (all-call detail)', sheet:'Maint by Port'},
    {id:'maintPlogNoTable', label:'Port Log Unmaintained Detail', sheet:'Maint PLog No'},
    {id:'maintVsNoTable', label:'Vessel Schedule Unmaintained Detail', sheet:'Maint VS No'},
    {id:'maintMonthTable', label:'Maintenance Monthly Trend', sheet:'Maint Monthly'},
  ]},
];

function openExportTablesModal(){
  var box=document.getElementById('exportTableChecks');
  var html='';
  EXPORT_TABLES.forEach(function(g){
    html+='<div style="margin:8px 0 2px;font-weight:700;font-size:11px;color:#1F4E79;">'+g.group+'</div>';
    g.items.forEach(function(it){
      html+='<label style="display:flex;align-items:center;gap:8px;padding:3px 4px;font-size:13px;cursor:pointer;">'+
            '<input type="checkbox" class="exp-chk" data-id="'+it.id+'" checked> '+it.label+
            '<span style="color:#9aa7b6;font-size:10px;">('+it.sheet+')</span></label>';
    });
  });
  box.innerHTML=html;
  document.getElementById('exportTablesModal').classList.add('open');
}
function closeExportTablesModal(){ document.getElementById('exportTablesModal').classList.remove('open'); }
function exportSelToggleAll(on){
  document.querySelectorAll('#exportTablesModal .exp-chk').forEach(function(c){ c.checked=on; });
}

// Read a rendered table's DOM into a styled sheet (title row + header + body
// rows). Numbers are parsed to numeric so Excel can sort/filter them.
function _tabSheetFromDom(tableId, title, sheetName, wb, skipFirst){
  var tbl=document.getElementById(tableId);
  if(!tbl) return false;
  var thead=tbl.querySelector('thead'), tbody=tbl.querySelector('tbody');
  var headTr=thead?thead.querySelector('tr'):null;
  if(!headTr) return false;
  var ths=headTr.querySelectorAll('th');
  var sf=skipFirst?1:0;
  var headers=[];
  for(var c=sf;c<ths.length;c++){
    var t=(ths[c].textContent||'').replace(/[▲▼↑↓→]/g,'').replace(/\s+/g,' ').trim();
    headers.push(t);
  }
  if(!headers.length) return false;
  var numCols=headers.length;

  function thinBorder(){var s={style:'thin',color:{rgb:'BFBFBF'}};return{top:s,bottom:s,left:s,right:s};}
  var B=thinBorder();
  function F(rgb){return{patternType:'solid',fgColor:{rgb:rgb}};}
  function A(h,v,wrap){var o={horizontal:h,vertical:v};if(wrap)o.wrapText=true;return o;}
  var tS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:14},fill:F('2E75B6'),alignment:A('center','center'),border:B};
  var hS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:10},fill:F('1F4E79'),alignment:A('center','center',true),border:B};

  var sheetData=[];
  var tr=[]; for(var c=0;c<numCols;c++) tr[c]={v:(c===0?title:''),s:tS};
  sheetData.push(tr);
  sheetData.push(headers.map(function(h){return{v:h,s:hS};}));

  // Only iterate DIRECT child rows (excludes nested tables inside expandable
  // detail sub-rows). Skip the Port Wait "detail-wrap" rows entirely — the user
  // wants the Port Wait export to contain the summary rows only, not the per-call detail.
  var rows=tbody?tbody.children:[];
  for(var i=0;i<rows.length;i++){
    if(rows[i].classList && rows[i].classList.contains('detail-wrap')) continue;
    var tds=rows[i].querySelectorAll('td');
    if(tds.length<=sf) continue;
    var fc=(i%2===0)?'EBF3FB':'FFFFFF';
    var nS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('left','center')};
    var cS={font:{name:'Arial',sz:9},fill:F(fc),border:B,alignment:A('center','center')};
    var row=[];
    for(var c=sf;c<tds.length;c++){
      var raw=(tds[c].textContent||'').trim();
      var clean=raw.replace(/,/g,'');
      var isNum=clean!=='' && /^-?\d+(\.\d+)?$/.test(clean);
      row.push({v:isNum?Number(clean):raw, s:isNum?cS:nS});
    }
    sheetData.push(row);
  }

  var totalRows=sheetData.length;
  var ws=XLSX.utils.aoa_to_sheet(sheetData);
  ws['!merges']=[{s:{r:0,c:0},e:{r:0,c:numCols-1}}];
  ws['!cols']=headers.map(function(h){return{wch:Math.max(h.length+4,10)};});
  ws['!rows']=[{hpt:26},{hpt:28}]; for(var j=2;j<totalRows;j++) ws['!rows'].push({hpt:16});
  if(totalRows>2) ws['!autofilter']={ref:'A2:'+XLSX.utils.encode_col(numCols-1)+totalRows};
  XLSX.utils.book_append_sheet(wb,ws,sheetName);
  return true;
}

function exportSelectedTables(){
  _loadXlsx(function(){
    var sel={};
    document.querySelectorAll('#exportTablesModal .exp-chk').forEach(function(c){ if(c.checked) sel[c.getAttribute('data-id')]=true; });
    var flat=[];
    EXPORT_TABLES.forEach(function(g){ g.items.forEach(function(it){ flat.push(it); }); });
    var wb=XLSX.utils.book_new();
    var todayStr=TODAY_DATA.date;
    var count=0;
    flat.forEach(function(it){
      if(!sel[it.id]) return;            // unselected → skip
      if(it.rich==='summary'){ exportSummaryExcel(wb); count++; }
      else if(it.rich==='full'){ exportFullScheduleExcel(wb); count++; }
      else { if(_tabSheetFromDom(it.id, 'CUL DAILY MOVEMENT — '+it.label+'  —  As of '+todayStr, it.sheet, wb, it.skipFirst)) count++; }
    });
    if(count===0){ alert('请至少勾选一张表再导出。'); return; }
    XLSX.writeFile(wb,'CUL Daily Movement '+todayStr+'.xlsx');
    closeExportTablesModal();
  });
}


/* ═══════════════════════════════════════════════════════════════════
   LANE ORDER MODAL (reorder Full Schedule lanes)
   ═══════════════════════════════════════════════════════════════════ */

// Build the full ordered lane list shown in the modal: only routes present in
// the current data, in the effective (override or ROUTE_ORDER) order, with any
// new data routes not yet in that order appended at the end (sorted by ROUTE_ORDER).
function _buildLaneList(){
  var dataRoutes = [...new Set((fullData||[]).map(function(r){ return r.route; }))].filter(Boolean);
  var present = {}; dataRoutes.forEach(function(r){ present[r]=true; });
  var order = getLaneOrder().filter(function(r){ return present[r]; });   // effective order, data routes only
  var inOrder = {}; order.forEach(function(r){ inOrder[r]=true; });
  var extra = dataRoutes.filter(function(r){ return !inOrder[r]; });
  extra.sort(function(a,b){
    var ia = ROUTE_ORDER.indexOf(a), ib = ROUTE_ORDER.indexOf(b);
    return ((ia<0?9999:ia) - (ib<0?9999:ib));
  });
  return order.concat(extra);
}

function openLaneOrderModal(){
  renderLaneOrderList();
  document.getElementById('laneOrderModal').classList.add('open');
}

function closeLaneOrderModal(){
  document.getElementById('laneOrderModal').classList.remove('open');
}

function renderLaneOrderList(){
  var list = _buildLaneList();
  var box = document.getElementById('laneOrderList');
  box.innerHTML = list.map(function(rt, idx){
    var upDisabled = idx===0 ? 'disabled' : '';
    var downDisabled = idx===list.length-1 ? 'disabled' : '';
    return '<div class="lane-row">'
      + '<span class="lane-idx">'+(idx+1)+'</span>'
      + '<span class="lane-code">'+escapeHtml(rt)+'</span>'
      + '<button class="lane-move" '+upDisabled+' onclick="moveLane('+idx+',-1)">&#9650;</button>'
      + '<button class="lane-move" '+downDisabled+' onclick="moveLane('+idx+',1)">&#9660;</button>'
      + '</div>';
  }).join('');
}

function moveLane(idx, dir){
  var list = _buildLaneList();
  var ni = idx + dir;
  if(ni < 0 || ni >= list.length) return;
  var tmp = list[idx]; list[idx] = list[ni]; list[ni] = tmp;
  LANE_ORDER_OVERRIDE = list;
  saveLaneOrder();
  renderLaneOrderList();          // re-render modal with new order
  renderFullSchedule();           // live-update the table below
}

function resetLaneOrder(){
  LANE_ORDER_OVERRIDE = null;
  saveLaneOrder();
  renderLaneOrderList();
  renderFullSchedule();
}


/* ═══════════════════════════════════════════════════════════════════
   PORT & SPEED ANALYTICS
   ═══════════════════════════════════════════════════════════════════ */

function escapeHtml(s){
  if(!s) return '';
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

// Remark classification: keyword -> category mapping
var REMARK_CATEGORIES = [
  {key:'congestion',  label:'Port Congestion',    keywords:['congestion','塞港','congestion delay','congested','拥堵','拥挤','congesiton']},
  {key:'weather',     label:'Weather',            keywords:['typhoon','台风','避台','大风浪','weather','storm','swell','fog','雾','monsoon']},
  {key:'bunker',      label:'Bunker',             keywords:['bunker','加油','bunkering','fuel','LSFO','MT','BUNKER']},
  {key:'phase',       label:'Phase In/Out',       keywords:['phase in','phase out','slide','eco speed','rotation','P/O','P/I','omit','OMIT','改靠','shifted','suspension']},
  {key:'msa',         label:'MSA / Regulatory',   keywords:['msa','regulatory','MSA']},
  {key:'adhoc',       label:'Ad Hoc Call',        keywords:['ad hoc','adhoc','extra call','add call','private call','ADD CALL']},
  {key:'cargo',       label:'Cargo / Trade Balance',  keywords:['balance','load balance','connection','trade','备货','等货','wait cargo']},
  {key:'other',       label:'Other',              keywords:[]}  // fallback
];

function classifyRemark(remark){
  if(!remark) return null;
  var r=remark.toLowerCase();
  for(var i=0;i<REMARK_CATEGORIES.length-1;i++){
    var kw=REMARK_CATEGORIES[i].keywords;
    for(var j=0;j<kw.length;j++){
      if(r.indexOf(kw[j])>=0) return REMARK_CATEGORIES[i].key;
    }
  }
  return 'other';
}

// Selected remark categories filter (null = show all)
var selRemarkCats = null;

function buildRemarkFilterDropdown(){
  var dd=document.getElementById('remarkFilterDropdown');
  var usedCats={};
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var cat=classifyRemark(sr.remark);
    if(cat) usedCats[cat]=true;
  });

  var html='';
  REMARK_CATEGORIES.forEach(function(cat){
    if(!usedCats[cat.key] && cat.key!=='other') return;
    var checked = !selRemarkCats || selRemarkCats.indexOf(cat.key)>=0;
    html+='<label style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;">';
    html+='<input type="checkbox" value="'+cat.key+'" '+(checked?'checked':'')+' onchange="onRemarkCatChange()">';
    html+=cat.label+'</label>';
  });
  dd.innerHTML=html;
}

function toggleRemarkFilter(){
  var dd=document.getElementById('remarkFilterDropdown');
  if(!dd.classList.contains('open')){ buildRemarkFilterDropdown(); dd.classList.add('open'); }
  else dd.classList.remove('open');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){if(d!==dd) d.classList.remove('open');});
}

function onRemarkCatChange(){
  var checks=document.querySelectorAll('#remarkFilterDropdown input[type=checkbox]');
  var sel=[];
  checks.forEach(function(cb){if(cb.checked) sel.push(cb.value);});
  // Denominator = actual visible checkboxes, not the full category list
  selRemarkCats = sel.length===checks.length ? null : sel;
  // Update button text
  var btn=document.getElementById('remarkFilterBtn');
  if(selRemarkCats===null){
    btn.textContent='All Remarks';
    btn.style.background='';
  } else {
    btn.textContent=sel.length+'/'+checks.length+' categories';
    btn.style.background='#fff3e0';
    btn.style.borderColor='#e67e22';
  }
  renderPortWaitAll();
}
var selPortFilter = null;  // null = show all ports
var selPortSearch = '';    // free-text port name search (AND with checkbox filter)

function onPortSearch(){
  selPortSearch = document.getElementById('portSearch').value.trim();
  renderPortWaitAll();
}

function buildPortFilterDropdown(){
  var dd=document.getElementById('portFilterDropdown');
  var allPorts={};
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var p=normalizePort(sr.port);
    if(p && !isBunkeringPort(sr.port)) allPorts[p]=true;
  });
  var sorted=Object.keys(allPorts).sort();
  var kw = (document.getElementById('portFilterSearchBox') ? document.getElementById('portFilterSearchBox').value : '').toLowerCase();
  var html='<div style="padding:2px 8px 6px;border-bottom:1px solid #e4ecf5;margin-bottom:4px;">';
  html+='<input type="text" id="portFilterSearchBox" placeholder="&#128269; Search ports..." value="'+kw+'" oninput="onPortFilterSearchBox()" style="width:100%;padding:4px 8px;border:1px solid #c9d5e2;border-radius:4px;font-size:12px;box-sizing:border-box;">';
  html+='</div>';
  html+='<label style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;">';
  html+='<input type="checkbox" value="__all__" '+(selPortFilter===null?'checked':'')+' onchange="onPortFilterAllChange(this)">';
  html+='<b>All Ports</b></label>';
  sorted.forEach(function(p){
    var checked = selPortFilter===null || selPortFilter.indexOf(p)>=0;
    var match = !kw || p.toLowerCase().indexOf(kw)>=0;
    html+='<label data-p="'+p+'" style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;'+(match?'':'display:none;')+'">';
    html+='<input type="checkbox" value="'+p+'" '+(checked?'checked':'')+' onchange="onPortFilterItemChange(this)">';
    html+=p+'</label>';
  });
  dd.innerHTML=html;
  updatePortFilterButton();
}

function onPortFilterSearchBox(){
  var kw=document.getElementById('portFilterSearchBox').value.toLowerCase();
  document.querySelectorAll('#portFilterDropdown label[data-p]').forEach(function(lb){
    var p=lb.getAttribute('data-p');
    lb.style.display = (!kw || p.toLowerCase().indexOf(kw)>=0) ? 'flex' : 'none';
  });
}

function togglePortFilter(){
  var dd=document.getElementById('portFilterDropdown');
  if(!dd.classList.contains('open')){ buildPortFilterDropdown(); dd.classList.add('open'); }
  else dd.classList.remove('open');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){if(d!==dd) d.classList.remove('open');});
}

function updatePortFilterButton(){
  var checks=document.querySelectorAll('#portFilterDropdown input[type=checkbox]:not([value="__all__"])');
  var total=0, sel=[];
  checks.forEach(function(cb){total++; if(cb.checked) sel.push(cb.value);});
  var btn=document.getElementById('portFilterBtn');
  if(sel.length===total){
    btn.textContent='All Ports';
  } else {
    btn.textContent=sel.length+' of '+total+' ports';
  }
}

function applyPortFilterFromDOM(){
  var allCb=document.querySelector('#portFilterDropdown input[value="__all__"]');
  var checks=document.querySelectorAll('#portFilterDropdown input[type=checkbox]:not([value="__all__"])');
  var total=0, sel=[];
  checks.forEach(function(cb){total++; if(cb.checked) sel.push(cb.value);});
  if(sel.length===total){
    selPortFilter=null;
    if(allCb) allCb.checked=true;
  } else {
    selPortFilter=sel;
    if(allCb) allCb.checked=false;
  }
  updatePortFilterButton();
  renderPortWaitAll();
}

function onPortFilterAllChange(allCb){
  var checks=document.querySelectorAll('#portFilterDropdown input[type=checkbox]:not([value="__all__"])');
  checks.forEach(function(cb){ cb.checked = allCb.checked; });
  if(allCb.checked){
    selPortFilter=null;
  } else {
    selPortFilter=[];
  }
  updatePortFilterButton();
  renderPortWaitAll();
}

function onPortFilterItemChange(cb){
  applyPortFilterFromDOM();
}

// ── Date Range Filter ───────────────────────────────────────────────
var selPortDateFrom=null, selPortDateTo=null, selPortDateBasis='eta';

// Raw date string for the selected filter basis (ETA / ETB / ETD)
function portWaitDateRaw(sr){
  if(selPortDateBasis==='etb') return sr.etbRaw||'';
  if(selPortDateBasis==='etd') return sr.etdRaw||'';
  return sr.etaRaw||'';
}

function onPortDateChange(){
  selPortDateFrom=document.getElementById('portDateFrom').value||null;
  selPortDateTo=document.getElementById('portDateTo').value||null;
  selPortDateBasis=document.getElementById('portDateBasis').value||'eta';
  renderPortWaitAll();
  // BOA follows the same date range (and basis) as Port Wait
  BOA_CALLS=buildBoaCalls();
  boaRefresh();
}

// ── Speed Vessel Filter (independent filter for Speed tab) ──────────
var selVesselFilter = null;  // null = show all vessels
var selVesselSearch = '';    // free-text vessel name search (AND with checkbox filter)

function onSpeedVesselSearch(){
  selVesselSearch = document.getElementById('speedVesselSearch').value.trim();
  renderSpeedTable();
}

function buildSpeedVesselFilterDropdown(){
  var dd=document.getElementById('speedVesselFilterDropdown');
  var btn=document.getElementById('speedVesselFilterBtn');
  var html='<label><input type="checkbox" value="__all__" onchange="onSpeedVesselFilterChange()"'+(selVesselFilter===null?' checked':'')+'> All Vessels</label>';
  var vessels=[]; var seen=new Set();
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var v=sr.vessel||'';
    if(v && !seen.has(v)){ seen.add(v); vessels.push(v); }
  });
  var selArr=selVesselFilter||[];
  vessels.sort(function(a,b){return a.localeCompare(b);});
  vessels.forEach(function(v){
    html+='<label><input type="checkbox" value="'+escapeHtml(v)+'" onchange="onSpeedVesselFilterChange()"'+(selVesselFilter===null||selArr.indexOf(v)>=0?' checked':'')+'> '+escapeHtml(v)+'</label>';
  });
  dd.innerHTML=html;
  btn.textContent=selVesselFilter===null?'All Vessels':selVesselFilter.length+' of '+vessels.length+' vessels';
}

function toggleSpeedVesselFilter(){
  var dd=document.getElementById('speedVesselFilterDropdown');
  if(!dd.classList.contains('open')){ buildSpeedVesselFilterDropdown(); dd.classList.add('open'); }
  else dd.classList.remove('open');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){if(d!==dd) d.classList.remove('open');});
}

function onSpeedVesselFilterChange(){
  var allCb=document.querySelector('#speedVesselFilterDropdown input[value="__all__"]');
  var checks=document.querySelectorAll('#speedVesselFilterDropdown input[type=checkbox]:not([value="__all__"])');
  var sel=[];
  var total=0;
  checks.forEach(function(cb){total++; if(cb.checked) sel.push(cb.value);});
  if(allCb && allCb.checked){
    selVesselFilter=null;
    checks.forEach(function(cb){cb.checked=true;});
  } else {
    selVesselFilter = sel.length===total ? null : sel;
  }
  if(allCb) allCb.checked = (selVesselFilter===null);
  var btn=document.getElementById('speedVesselFilterBtn');
  btn.textContent=selVesselFilter===null?'All Vessels':selVesselFilter.length+' of '+total+' vessels';
  renderSpeedTable();
}

// ── Port Wait Analysis ────────────────────────────────────────────────

var portWaitData=[];
var portWaitSortCol=-1, portWaitSortDir=1;
var remarkCatTotals={};
var monthlyTrendData=[];

// Port Wait over-range selector (3 categories, mirrors BOA):
//   all = wait > standard (≥6h, SHA/CNNGB≥12h); 24+ = wait ≥24h; 48+ = wait ≥48h
// Berth threshold: Shanghai (CNSHA) and Ningbo (CNNGB) use 12h, all other ports 6h.
function berthThreshold(port){ return (port==='CNSHA' || port==='CNNGB') ? 12 : 6; }
var PORT_OVER_RANGE='all';
function portOverInRange(wait, port){
  if(wait <= berthThreshold(port)) return false;  // berth, not over
  switch(PORT_OVER_RANGE){
    case 'all': return true;
    case '24+': return wait >= 24;
    case '48+': return wait >= 48;
  }
  return false;
}

// Merge port name variants: AEJEA(T1)→AEJEA, CNSHK-CCT→CNSHK, DJJIB(DMP)→DJJIB,
// MYPKG (1st CALL)→MYPKG, THLCH (ESCO)→THLCH, SGSIN(Bunkering)→SGSIN(bunker) etc.
function normalizePort(p){
  if(!p) return '';
  var s=p.trim();
  // Strip anything in parentheses/brackets: (T1), (DMP), (SGTD), (RSGT), (ESCO), (TIPS), (1st CALL), (2nd CALL), (Bunkering)
  s=s.replace(/\s*[\(\（][^)\）]*[\)\）]/g,'');
  // Strip "-suffix": -Shipyard, -CCT, -MCT, -DPW
  s=s.replace(/\s*-\s*[A-Za-z0-9]+$/,'');
  // Strip " anchorage" / "anchoage" (typo)
  s=s.replace(/\s*anchorage/i,'').replace(/\s*anchoage/i,'');
  // Strip " 1st CALL" / " 2nd CALL"
  s=s.replace(/\s*\d+st\s*CALL/i,'').replace(/\s*\d+nd\s*CALL/i,'');
  // Special: "CJK & NGB anchoage" → "CNNGB" (merge into Ningbo)
  if(/CJK.*NGB/i.test(s)) s='CNNGB';
  // ── Final UN/LOCODE aliases (AFTER all stripping, so variants like
  //    "JED-DPW"→"JED" also merge) ──────────────────────────────
  if(s==='CNNSA') s='CNNAS';      // Nansha
  if(s==='JED')  s='SAJED';       // Jeddah
  return s.trim();
}

// Is this a bunkering-only call?
function isBunkeringPort(p){
  return /bunker/i.test(p||'');
}

var totalExcludedByRemark=0;
function buildPortWaitData(){
  var byPort={};
  totalExcludedByRemark=0;
  var todayStr=TODAY_DATA.date;

  // Pre-compute date range bounds
  var df=selPortDateFrom||'', dt=selPortDateTo||todayStr;

  TODAY_DATA.fullSchedule.forEach(function(sr){
    var rawPort=sr.port||'';
    if(!rawPort) return;
    // Skip bunkering-only calls
    if(isBunkeringPort(rawPort)) return;
    var port=normalizePort(rawPort);
    if(!port) return;

    // Date range filter (by selected basis: ETA / ETB / ETD)
    var era=portWaitDateRaw(sr);
    if(!era) return;
    if(df && era < df) return;
    if(dt && era > dt) return;

    // Apply port filter
    if(selPortFilter && selPortFilter.indexOf(port)<0) return;
    // Apply free-text port search (AND with checkbox filter)
    if(selPortSearch && port.toLowerCase().indexOf(selPortSearch.toLowerCase())<0) return;

    var wait=parseFloat(sr.wait)||0;
    var remark=sr.remark||'';
    var cat=classifyRemark(remark)||'other';

    // If remark filter is active, ONLY include calls whose remark matches selected categories
    if(selRemarkCats){
      var match=false;
      for(var i=0;i<selRemarkCats.length;i++){
        if(cat===selRemarkCats[i]){match=true;break;}
        if(selRemarkCats[i]==='other' && !remark){match=true;break;}
      }
      if(!match){
        // Ensure port record exists so excluded calls can still be inspected
        if(!byPort[port]){
          byPort[port]={port:port, region:BOA_PORT_REGION[port]||'Unknown',
                        calls:[], totalWait:0, maxWait:0, longWaitCalls:0, overCalls:0, berthedCalls:0,
                        remarks:{}, excludedCalls:[]};
        }
        byPort[port].excludedCalls.push({wait:wait, remark:remark, cat:cat, vessel:sr.vessel, voy:sr.voy, eta:sr.eta, etb:sr.etb, etd:sr.etd, rawPort:rawPort});
        totalExcludedByRemark++;
        return;  // excluded calls do not count in berth rate numerator or denominator
      }
    }

    // Initialize port record if needed
    if(!byPort[port]){
      byPort[port]={port:port, region:BOA_PORT_REGION[port]||'Unknown',
                    calls:[], totalWait:0, maxWait:0, longWaitCalls:0, overCalls:0, berthedCalls:0,
                    remarks:{}, excludedCalls:[]};
    }
    var rec=byPort[port];

    // Filtered call — add to stats
    rec.calls.push({wait:wait, remark:remark, cat:cat, vessel:sr.vessel, voy:sr.voy, eta:sr.eta, etb:sr.etb, etd:sr.etd, etbRaw:sr.etbRaw||'', dateKey:era, rawPort:rawPort, port:port});
    rec.totalWait+=wait;
    if(wait>rec.maxWait) rec.maxWait=wait;
    if(wait>=24) rec.longWaitCalls++;
    if(portOverInRange(wait, port))     rec.overCalls++;
    if(wait <= berthThreshold(port)) rec.berthedCalls++;  // berthed = waited within threshold (SHA/NGB 12h, others 6h)
    if(remark){
      if(!rec.remarks[cat]) rec.remarks[cat]=[];
      rec.remarks[cat].push(remark);
    }
  });

  var result=[];
  // Collect all calls for monthly aggregation
  var allFilteredCalls=[];
  for(var p in byPort){
    var rec=byPort[p];
    rec.avgWait=rec.calls.length>0 ? (rec.totalWait/rec.calls.length) : 0;
    // Berth rate = berthed calls (with ETB) / total filtered calls; remark filter applies to both
    rec.berthRate=rec.calls.length>0 ? Math.round(rec.berthedCalls/rec.calls.length*100) : 0;
    rec.catLabels=Object.keys(rec.remarks).map(function(k){
      var found=REMARK_CATEGORIES.find(function(c){return c.key===k;});
      return found ? found.label : k;
    }).join(', ');
    result.push(rec);
    // Collect for monthly trend
    rec.calls.forEach(function(cl){ allFilteredCalls.push(cl); });
  }

  // Default sort: berth rate descending (best ports first)
  if(portWaitSortCol<0){
    result.sort(function(a,b){return b.berthRate-a.berthRate || a.avgWait-b.avgWait;});
  }

  // Aggregate wait time by remark category
  var catTotals={};
  result.forEach(function(rec){
    rec.calls.forEach(function(cl){
      var c=cl.cat||'other';
      if(!catTotals[c]) catTotals[c]={wait:0, calls:0};
      catTotals[c].wait+=cl.wait;
      catTotals[c].calls++;
    });
  });

  portWaitData=result;
  remarkCatTotals=catTotals;

  // ── Monthly Trend Aggregation ──────────────────────────────────────
  // Per-month berth rate = (calls berthed within threshold) / (total filtered calls in month)
  var byMonth={};
  allFilteredCalls.forEach(function(cl){
    // dateKey is the selected date-basis raw ("YYYY-MM-DD" format)
    var dk=cl.dateKey||'';
    if(!dk || dk.length<7) return;
    var m=dk.substring(0,7); // "YYYY-MM"
    if(!byMonth[m]) byMonth[m]={month:m, totalWait:0, count:0, maxWait:0, berthedCalls:0, calls:[]};
    var mr=byMonth[m];
    mr.totalWait+=cl.wait;
    mr.count++;
    if(cl.wait>mr.maxWait) mr.maxWait=cl.wait;
    if(cl.wait <= berthThreshold(cl.port||'')) mr.berthedCalls++;  // berthed = waited within threshold
    mr.calls.push(cl);
  });
  // Sort months
  var months=Object.keys(byMonth).sort();
  monthlyTrendData=months.map(function(m){
    var mr=byMonth[m];
    return {month:m, totalWait:mr.totalWait, count:mr.count, avgWait:mr.totalWait/mr.count,
            maxWait:mr.maxWait, berthCalls:mr.berthedCalls,
            berthRate: mr.count>0 ? Math.round(mr.berthedCalls/mr.count*100) : 0};
  });
}

var PORT_WAIT_COLS=[
  {key:'rank',      label:'#'},
  {key:'port',      label:'Port'},
  {key:'region',    label:'Region'},
  {key:'calls',     label:'Calls'},
  {key:'totalWait', label:'Total Wait (hrs)'},
  {key:'avgWait',   label:'Avg Wait (hrs)'},
  {key:'maxWait',   label:'Max Wait (hrs)'},
  {key:'overCalls', label:'Over (range)'},
  {key:'berthRate', label:'Berth Rate'},
  {key:'catLabels', label:'Remark Categories'},
];

function berthRateColor(rate){
  if(rate>=80) return '#27ae60';  // green = good
  if(rate>=50) return '#e67e22';  // orange = medium
  return '#c0392b';               // red = bad
}

function renderPortWaitTable(){
  buildPortWaitData();
  var data=portWaitData;
  var thead=document.getElementById('portWaitThead');
  var tbody=document.getElementById('portWaitTbody');

  // Header
  var h='<th style="width:22px;"></th>';
  PORT_WAIT_COLS.forEach(function(col,i){
    var arrow='';
    if(portWaitSortCol===i) arrow=portWaitSortDir===1?' \u25B2':' \u25BC';
    h+='<th style="cursor:pointer" onclick="sortPortWait('+i+')">'+col.label+arrow+'</th>';
  });
  thead.innerHTML='<tr>'+h+'</tr>';

  // Sort
  if(portWaitSortCol>=0){
    var def=PORT_WAIT_COLS[portWaitSortCol];
    var k=def.key; var d=portWaitSortDir;
    data=data.slice().sort(function(a,b){
      if(k==='port' || k==='catLabels' || k==='region') return (a[k]||'').localeCompare(b[k]||'')*d;
      if(k==='rank') return 0;  // rank is display-only, re-sort by berthRate
      var va=a[k]||0, vb=b[k]||0;
      return (va-vb)*d;
    });
  }

  // Body
  var rows='';
  data.forEach(function(r,idx){
    var rate=r.berthRate;
    var c=berthRateColor(rate);
    var bar='<div style="display:inline-flex;align-items:center;gap:6px;">'+
      '<div style="width:80px;height:14px;background:#e8e8e8;border-radius:7px;overflow:hidden;">'+
      '<div style="width:'+rate+'%;height:100%;background:'+c+';border-radius:7px;"></div></div>'+
      '<b style="color:'+c+';min-width:36px;">'+rate+'%</b></div>';
    var cat=(r.catLabels||'') ? '<span style="color:#C00000;font-size:11px;">'+r.catLabels+'</span>' : '<span style="color:#8a9bb0;">—</span>';
    var rowBg=rate<50 ? ' style="background:#fff5f5;"' : (rate>=80 ? ' style="background:#f0faf3;"' : '');
    var pid='pw'+idx;
    // Build call detail rows (hidden by default)
    var callRows='';
    r.calls.forEach(function(cl,i2){
      var w=cl.wait;
      var wClass=w>=24?'delay':(w<6?'ontime':'');
      var wDisp=w.toFixed(1);
      if(wClass==='delay') wDisp='<b>'+wDisp+'</b>';
      callRows+='<tr class="detail-row" style="background:#fff;">'+
        '<td style="color:#8a9bb0;font-size:11px;">'+(i2+1)+'</td>'+
        '<td style="font-size:12px;">'+escapeHtml(cl.vessel||'')+'</td>'+
        '<td style="font-size:11px;color:#6a7b8d;">'+escapeHtml(cl.voy||'')+'</td>'+
        '<td style="font-size:11px;color:#6a7b8d;">'+escapeHtml(cl.eta||'')+'</td>'+
        '<td style="font-size:11px;color:#6a7b8d;">'+escapeHtml(cl.etb||'')+'</td>'+
        '<td style="font-size:11px;color:#6a7b8d;">'+escapeHtml(cl.etd||'')+'</td>'+
        '<td class="center '+wClass+'" style="font-size:12px;">'+wDisp+'</td>'+
        '<td style="font-size:11px;'+(cl.remark?'color:#C00000;':'color:#8a9bb0;')+'">'+(cl.remark?escapeHtml(cl.remark):'—')+'</td>'+
        '</tr>';
    });
    // Show excluded calls if remark filter is active
    var excludedRows='';
    if(selRemarkCats && r.excludedCalls && r.excludedCalls.length>0){
      excludedRows='<tr><td colspan="8" style="padding:6px 8px;color:#999;font-size:11px;border-top:1px dashed #e0e0e0;">'+
        '<span style="color:#e67e22;">&#9888;</span> Filtered out ('+r.excludedCalls.length+' calls excluded by remark filter):</td></tr>';
      r.excludedCalls.forEach(function(cl){
        var catLabel='';
        var found=REMARK_CATEGORIES.find(function(c){return c.key===cl.cat;});
        if(found) catLabel=found.label;
        excludedRows+='<tr class="excluded-row" style="background:#fafafa;color:#b0b0b0;text-decoration:line-through;">'+
          '<td style="color:#ccc;font-size:11px;">—</td>'+
          '<td style="font-size:11px;">'+escapeHtml(cl.vessel||'')+'</td>'+
          '<td style="font-size:10px;">'+escapeHtml(cl.voy||'')+'</td>'+
          '<td style="font-size:10px;">'+escapeHtml(cl.eta||'')+'</td>'+
          '<td style="font-size:10px;color:#6a7b8d;">'+escapeHtml(cl.etb||'')+'</td>'+
          '<td style="font-size:10px;color:#6a7b8d;">'+escapeHtml(cl.etd||'')+'</td>'+
          '<td class="center" style="font-size:11px;">'+(cl.wait||0).toFixed(1)+'</td>'+
          '<td style="font-size:10px;"><span style="background:#f5f5f5;color:#999;padding:1px 5px;border-radius:3px;">'+catLabel+'</span> '+(cl.remark?escapeHtml(cl.remark):'—')+'</td>'+
          '</tr>';
      });
    }
    var detailHtml='';
    if(callRows || excludedRows){
      detailHtml='<tr id="'+pid+'-detail" class="detail-wrap" style="display:none;"><td></td><td colspan="10" style="padding:0;">'+
        '<div style="padding:4px 0;">'+
        '<table style="width:100%;font-size:12px;border-collapse:collapse;">'+
        '<thead><tr style="background:#eef3f7;color:#5a697a;font-size:11px;">'+
        '<th style="width:24px;padding:3px 6px;">#</th>'+
        '<th style="padding:3px 6px;text-align:left;">Vessel</th>'+
        '<th style="padding:3px 6px;text-align:left;">Voy</th>'+
        '<th style="padding:3px 6px;text-align:left;">ETA</th>'+
        '<th style="padding:3px 6px;text-align:left;">ETB</th>'+
        '<th style="padding:3px 6px;text-align:left;">ETD</th>'+
        '<th style="padding:3px 6px;text-align:center;">Wait (hrs)</th>'+
        '<th style="padding:3px 6px;text-align:left;">Remark</th>'+
        '</tr></thead>'+
        '<tbody>'+callRows+excludedRows+'</tbody>'+
        '</table></div></td></tr>';
    }
    var overCell = r.overCalls>0 ? '<b>'+r.overCalls+'</b>' : '0';
    rows+='<tr'+rowBg+' class="port-row" onclick="togglePortWaitDetail(\''+pid+'\')" style="cursor:pointer;">'+
      '<td class="center" style="font-size:12px;color:#8a9bb0;">'+(callRows?'<span id="'+pid+'-icon">&#9654;</span>':'')+'</td>'+
      '<td class="center" style="color:#8a9bb0;font-size:12px;">'+(idx+1)+'</td>'+
      '<td><strong>'+r.port+'</strong></td>'+
      '<td class="center">'+r.region+'</td>'+
      '<td class="center">'+(selRemarkCats?r.calls.length+' <span style="font-size:10px;color:#999;">/ '+r.allCalls+'</span>':r.calls.length)+'</td>'+
      '<td class="center">'+r.totalWait.toFixed(1)+'</td>'+
      '<td class="center">'+r.avgWait.toFixed(1)+'</td>'+
      '<td class="center">'+r.maxWait.toFixed(1)+'</td>'+
      '<td class="center'+(r.overCalls>0?' delay':'')+'">'+overCell+'</td>'+
      '<td>'+bar+'</td>'+
      '<td>'+cat+'</td>'+
      '</tr>'+detailHtml;
  });
  tbody.innerHTML=rows;
  var statText=data.length+' ports';
  // Compute total calls across all ports (unfiltered by remark)
  var totalAllCalls=0;
  data.forEach(function(r){totalAllCalls+=r.allCalls;});
  if(selRemarkCats && totalExcludedByRemark>0){
    statText+=' · <span style="color:#e67e22;">'+totalExcludedByRemark+' calls filtered</span>';
    statText+=' · '+totalAllCalls+' total calls in range';
  } else {
    statText+=' · '+totalAllCalls+' calls in range';
  }
  statText+=' · berth rate = wait < threshold (CNSHA & CNNGB <12h, others <6h) / total calls · Over (range) follows current over-range selector';
  document.getElementById('statPortWait').innerHTML=statText;
}

function togglePortWaitDetail(pid){
  var detail=document.getElementById(pid+'-detail');
  var icon=document.getElementById(pid+'-icon');
  if(!detail) return;
  if(detail.style.display==='none'){
    detail.style.display='';
    if(icon) icon.innerHTML='&#9660;';
  }else{
    detail.style.display='none';
    if(icon) icon.innerHTML='&#9654;';
  }
}

function sortPortWait(col){
  // col is PORT_WAIT_COLS index: 0=rank(not sortable), 1=port, 2=calls, ...
  if(col===0) return;
  if(portWaitSortCol===col) portWaitSortDir*=-1;
  else{portWaitSortCol=col;portWaitSortDir=1;}
  renderPortWaitAll();
}

// ── Wait by Port Region (Port→Region matching, mirrors BOA) ────────────────
var PORT_WAIT_REGION_COLS=[
  {key:'region',    label:'Port Region'},
  {key:'calls',     label:'Calls'},
  {key:'totalWait', label:'Total Wait (hrs)'},
  {key:'avgWait',   label:'Avg Wait (hrs)'},
  {key:'overCalls', label:'Over (range)'},
  {key:'overRate',  label:'Over Rate'},
];
function renderPortWaitByRegion(){
  var byRegion={};
  portWaitData.forEach(function(r){
    var reg=r.region||'Unknown';
    if(!byRegion[reg]) byRegion[reg]={region:reg, calls:0, totalWait:0, over:0, maxWait:0};
    var a=byRegion[reg];
    a.calls+=r.calls.length;
    a.totalWait+=r.totalWait;
    a.over+=r.overCalls;
    if(r.maxWait>a.maxWait) a.maxWait=r.maxWait;
  });
  var rows=Object.keys(byRegion).map(function(k){
    var a=byRegion[k];
    return {region:k, calls:a.calls, totalWait:a.totalWait,
            avgWait:a.calls?a.totalWait/a.calls:0,
            overCalls:a.over, overRate:a.calls?a.over/a.calls:0, maxWait:a.maxWait};
  });
  rows.sort(function(a,b){return b.calls-a.calls;});
  var thead=document.getElementById('portWaitRegionThead');
  var tbody=document.getElementById('portWaitRegionTbody');
  thead.innerHTML='<tr>'+PORT_WAIT_REGION_COLS.map(function(c){return '<th>'+c.label+'</th>';}).join('')+'</tr>';
  if(!rows.length){ tbody.innerHTML='<tr><td colspan="6" class="no-data">No data in range</td></tr>'; return; }
  var totalCalls=0,totalOver=0,totalWait=0;
  var body=rows.map(function(r){
    totalCalls+=r.calls; totalOver+=r.overCalls; totalWait+=r.totalWait;
    return '<tr>'+
      '<td><strong>'+r.region+'</strong></td>'+
      '<td class="center">'+r.calls+'</td>'+
      '<td class="center">'+r.totalWait.toFixed(1)+'</td>'+
      '<td class="center">'+r.avgWait.toFixed(1)+'</td>'+
      '<td class="center'+(r.overCalls>0?' delay':'')+'">'+(r.overCalls>0?'<b>'+r.overCalls+'</b>':'0')+'</td>'+
      '<td class="center">'+Math.round(r.overRate*100)+'%</td>'+
      '</tr>';
  }).join('');
  body+='<tr style="font-weight:700;background:#eef4fa;">'+
    '<td>Total</td>'+
    '<td class="center">'+totalCalls+'</td>'+
    '<td class="center">'+totalWait.toFixed(1)+'</td>'+
    '<td class="center">'+(totalCalls?(totalWait/totalCalls).toFixed(1):'0.0')+'</td>'+
    '<td class="center">'+totalOver+'</td>'+
    '<td class="center">'+(totalCalls?Math.round(totalOver/totalCalls*100)+'%':'0%')+'</td>'+
    '</tr>';
  tbody.innerHTML=body;
}

// ── Unified filtered call list (date+port+search+remark), port-normalized ──
function getFilteredCalls(){
  var todayStr=TODAY_DATA.date;
  var df=selPortDateFrom||'', dt=selPortDateTo||todayStr;
  var out=[];
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var rawPort=sr.port||'';
    if(!rawPort) return;
    if(isBunkeringPort(rawPort)) return;
    var port=normalizePort(rawPort);
    if(!port) return;
    var era=portWaitDateRaw(sr);
    if(!era) return;
    if(df && era<df) return;
    if(dt && era>dt) return;
    if(selPortFilter && selPortFilter.indexOf(port)<0) return;
    if(selPortSearch && port.toLowerCase().indexOf(selPortSearch.toLowerCase())<0) return;
    var wait=parseFloat(sr.wait)||0;
    var remark=sr.remark||'';
    var cat=classifyRemark(remark)||'other';
    if(selRemarkCats){
      var match=false;
      for(var i=0;i<selRemarkCats.length;i++){
        if(cat===selRemarkCats[i]){match=true;break;}
        if(selRemarkCats[i]==='other' && !remark){match=true;break;}
      }
      if(!match) return;
    }
    var route=sr.route||'(blank)';
    out.push({port:port, region:BOA_PORT_REGION[port]||'Unknown', route:route, trade:BOA_LANE_TRADE[route]||'Unknown', wait:wait});
  });
  return out;
}

// ── Port Call Count by Lane (Trade→Lane), Excel summary "BOA by Lane" form ──
function renderPortCallCountByLane(){
  var calls=getFilteredCalls();
  var byTrade={};
  calls.forEach(function(c){
    if(!byTrade[c.trade]) byTrade[c.trade]={trade:c.trade, lanes:{}, total:0};
    var t=byTrade[c.trade];
    if(!t.lanes[c.route]) t.lanes[c.route]=0;
    t.lanes[c.route]++;
    t.total++;
  });
  var trades=Object.keys(byTrade).sort();
  var thead=document.getElementById('callCountLaneThead');
  var tbody=document.getElementById('callCountLaneTbody');
  thead.innerHTML='<tr><th>Lane Trade</th><th>Lane</th><th>Calls</th></tr>';
  if(!trades.length){ tbody.innerHTML='<tr><td colspan="3" class="no-data">No data in range</td></tr>'; return; }
  var html='';
  trades.forEach(function(tk){
    var t=byTrade[tk];
    var lanes=Object.keys(t.lanes).sort();
    lanes.forEach(function(lk){
      html+='<tr class="boa-subrow"><td class="boa-grp">'+t.trade+'</td><td>'+lk+'</td><td class="center">'+t.lanes[lk]+'</td></tr>';
    });
    html+='<tr class="boa-sumrow"><td class="boa-grp" colspan="2">'+t.trade+' Total</td><td class="center">'+t.total+'</td></tr>';
  });
  html+='<tr class="boa-grand"><td class="boa-grp" colspan="2">Total</td><td class="center">'+calls.length+'</td></tr>';
  tbody.innerHTML=html;
}

// ── Port Call Count by Region (Region→Port), Excel summary "BOA by Port" form ──
function renderPortCallCountByRegion(){
  var calls=getFilteredCalls();
  var byRegion={};
  calls.forEach(function(c){
    if(!byRegion[c.region]) byRegion[c.region]={region:c.region, ports:{}, total:0};
    var r=byRegion[c.region];
    if(!r.ports[c.port]) r.ports[c.port]=0;
    r.ports[c.port]++;
    r.total++;
  });
  var regions=Object.keys(byRegion).sort();
  var thead=document.getElementById('callCountRegionThead');
  var tbody=document.getElementById('callCountRegionTbody');
  thead.innerHTML='<tr><th>Port Region</th><th>Port</th><th>Calls</th></tr>';
  if(!regions.length){ tbody.innerHTML='<tr><td colspan="3" class="no-data">No data in range</td></tr>'; return; }
  var html='';
  regions.forEach(function(rk){
    var r=byRegion[rk];
    var ports=Object.keys(r.ports).sort();
    ports.forEach(function(pk){
      html+='<tr class="boa-subrow"><td class="boa-grp">'+r.region+'</td><td>'+pk+'</td><td class="center">'+r.ports[pk]+'</td></tr>';
    });
    html+='<tr class="boa-sumrow"><td class="boa-grp" colspan="2">'+r.region+' Total</td><td class="center">'+r.total+'</td></tr>';
  });
  html+='<tr class="boa-grand"><td class="boa-grp" colspan="2">Total</td><td class="center">'+calls.length+'</td></tr>';
  tbody.innerHTML=html;
}

function renderRemarkSummary(){
  var totals=remarkCatTotals;
  var totalWait=0, totalCalls=0;
  for(var k in totals){ totalWait+=totals[k].wait; totalCalls+=totals[k].calls; }

  var cont=document.getElementById('remarkSummary');
  if(totalWait===0){
    cont.innerHTML='<div style=\"color:#8a9bb0;font-size:12px;padding:12px;\">No wait data matches the selected remark filter.</div>';
    return;
  }

  // Build sorted list by wait desc
  var items=[];
  REMARK_CATEGORIES.forEach(function(cat){
    var t=totals[cat.key]||{wait:0,calls:0};
    if(t.calls>0 || cat.key==='other') items.push({key:cat.key, label:cat.label, wait:t.wait, calls:t.calls});
  });
  items.sort(function(a,b){return b.wait-a.wait;});

  // Color palette per category
  var colors={congestion:'#c0392b', weather:'#2980b9', bunker:'#8e44ad', phase:'#d35400', msa:'#e67e22', adhoc:'#f39c12', cargo:'#16a085', other:'#7f8c8d'};

  var html='';
  // Summary header
  html+='<div style="display:flex;align-items:center;gap:16px;margin-bottom:8px;font-size:12px;">';
  html+='<span style="font-weight:600;">Total Wait: <b style="color:#c00000;">'+totalWait.toFixed(1)+'h</b></span>';
  html+='<span>Avg per call: <b>'+ (totalCalls>0?(totalWait/totalCalls).toFixed(1):'0') +'h</b></span>';
  html+='<span>Calls: <b>'+totalCalls+'</b></span>';
  html+='</div>';

  // Bars
  var maxW=totalWait||1;
  items.forEach(function(it){
    var pct=Math.round(it.wait/maxW*100);
    var c=colors[it.key]||'#7f8c8d';
    if(pct===0){ html+='<div style="display:flex;align-items:center;gap:8px;font-size:12px;color:#8a9bb0;height:20px;">'+it.label+' <span>— 0h (0 calls)</span></div>'; return; }
    html+='<div style="display:flex;align-items:center;gap:8px;font-size:12px;">';
    html+='<span style="width:160px;text-align:right;font-weight:600;flex-shrink:0;">'+it.label+'</span>';
    html+='<div style="flex:1;height:20px;background:#e8e8e8;border-radius:4px;overflow:hidden;min-width:80px;">';
    html+='<div style="width:'+(pct||1)+'%;height:100%;background:'+c+';border-radius:4px;display:flex;align-items:center;justify-content:flex-end;padding-right:6px;font-size:10px;color:#fff;font-weight:600;min-width:'+(pct>0?pct*0.5:1)+'px;">'+(pct>=8?pct+'%':'')+'</div>';
    html+='</div>';
    html+='<span style="font-weight:700;color:'+c+';width:56px;text-align:right;flex-shrink:0;">'+it.wait.toFixed(1)+'h</span>';
    html+='<span style="color:#8a9bb0;width:48px;text-align:right;flex-shrink:0;">'+it.calls+' call'+(it.calls>1?'s':'')+'</span>';
    html+='</div>';
  });

  cont.innerHTML=html;
}

// ── Monthly Port Wait Trend ──────────────────────────────────────────
function buildMonthlyTrend(){
  // Called by buildPortWaitData — monthlyTrendData is already populated
  // This is a no-op since data is built inline. Re-call buildPortWaitData if needed.
}
function renderMonthlyTrend(){
  buildPortWaitData();
  var data=monthlyTrendData;
  var cont=document.getElementById('monthlyTrend');
  if(!data || data.length===0){
    cont.innerHTML='<div style="color:#8a9bb0;font-size:12px;padding:12px;">No monthly data available for the selected range.</div>';
    return;
  }

  var totalCalls=0, totalBerth=0;
  data.forEach(function(m){ totalCalls+=m.count; totalBerth+=m.berthCalls; });
  var overallRate = totalCalls>0 ? Math.round(totalBerth/totalCalls*100) : 0;

  // Summary header
  var html='<div style="display:flex;align-items:center;gap:16px;margin-bottom:8px;font-size:12px;">';
  html+='<span>Months: <b>'+data.length+'</b></span>';
  html+='<span>Total Calls: <b>'+totalCalls+'</b></span>';
  html+='<span>Berth Rate: <b style="color:'+berthRateColor(overallRate)+';">'+overallRate+'%</b></span>';
  html+='<span>Total Wait: <b style="color:#c00000;">'+data.reduce(function(s,m){return s+m.totalWait;},0).toFixed(1)+'h</b></span>';
  html+='</div>';

  // Bars — width = monthly berth rate (berthed-within-threshold calls / total calls); color follows berthRateColor
  data.forEach(function(m){
    var pct=m.berthRate; // 0-100
    var avg=m.avgWait.toFixed(1);
    var c=berthRateColor(pct);
    html+='<div style="display:flex;align-items:center;gap:8px;font-size:12px;">';
    html+='<span style="width:80px;text-align:right;font-weight:600;flex-shrink:0;">'+m.month+'</span>';
    html+='<div style="flex:1;height:22px;background:#e8e8e8;border-radius:4px;overflow:hidden;min-width:80px;">';
    html+='<div style="width:'+(pct||1)+'%;height:100%;background:'+c+';border-radius:4px;display:flex;align-items:center;justify-content:flex-end;padding-right:6px;font-size:10px;color:#fff;font-weight:600;min-width:'+(pct>0?Math.max(pct*0.5,18):1)+'px;">'+pct+'%</div>';
    html+='</div>';
    html+='<span style="font-weight:700;color:'+c+';width:44px;text-align:right;flex-shrink:0;">'+pct+'%</span>';
    html+='<span style="color:#8a9bb0;width:84px;text-align:right;flex-shrink:0;">'+m.berthCalls+'/'+m.count+' berthed</span>';
    html+='<span style="color:#6a7b8d;width:64px;text-align:right;flex-shrink:0;">avg '+avg+'h</span>';
    html+='</div>';
  });

  cont.innerHTML=html;
}

// ── Vessel Speed Analysis ─────────────────────────────────────────────

var vesselSpeedData=[];
var speedSortCol=-1, speedSortDir=1;

function buildVesselSpeedData(){
  var byVessel={};
  var search=(selVesselSearch||'').toLowerCase();
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var v=sr.vessel||'';
    if(selVesselFilter && selVesselFilter.indexOf(v)<0) return;  // vessel checkbox filter
    if(search && v.toLowerCase().indexOf(search)<0) return;      // vessel name search
    var spd=parseFloat(sr.speed);
    if(!v || isNaN(spd) || spd<=0 || spd>20) return;  // exclude unrealistic: ≤0 or >20kn
    if(!byVessel[v]) byVessel[v]={vessel:v, route:sr.route, speeds:[], sum:0, min:spd, max:spd};
    var rec=byVessel[v];
    rec.speeds.push(spd);
    rec.sum+=spd;
    if(spd<rec.min) rec.min=spd;
    if(spd>rec.max) rec.max=spd;
  });
  var result=[];
  for(var v in byVessel){
    var rec=byVessel[v];
    rec.avg=rec.speeds.length>0 ? rec.sum/rec.speeds.length : 0;
    rec.legs=rec.speeds.length;
    result.push(rec);
  }
  vesselSpeedData=result;
}

var SPEED_COLS=[
  {key:'vessel', label:'Vessel'},
  {key:'route',  label:'Route'},
  {key:'legs',   label:'Legs'},
  {key:'avg',    label:'Avg Speed (kn)'},
  {key:'min',    label:'Min Speed (kn)'},
  {key:'max',    label:'Max Speed (kn)'},
];

function renderSpeedTable(){
  buildVesselSpeedData();
  var data=vesselSpeedData;
  var thead=document.getElementById('speedThead');
  var tbody=document.getElementById('speedTbody');

  var h='';
  SPEED_COLS.forEach(function(col,i){
    var arrow='';
    if(speedSortCol===i) arrow=speedSortDir===1?' \u25B2':' \u25BC';
    h+='<th style="cursor:pointer" onclick="sortSpeed('+i+')">'+col.label+arrow+'</th>';
  });
  thead.innerHTML='<tr>'+h+'</tr>';

  if(speedSortCol>=0){
    var def=SPEED_COLS[speedSortCol];
    var k=def.key; var d=speedSortDir;
    data=data.slice().sort(function(a,b){
      if(k==='vessel' || k==='route') return (a[k]||'').localeCompare(b[k]||'')*d;
      var va=a[k]||0, vb=b[k]||0;
      return (va-vb)*d;
    });
  }

  var rows='';
  data.forEach(function(r){
    rows+='<tr>'+
      '<td><strong>'+r.vessel+'</strong></td>'+
      '<td class="center">'+r.route+'</td>'+
      '<td class="center">'+r.legs+'</td>'+
      '<td class="center"><b>'+r.avg.toFixed(1)+'</b></td>'+
      '<td class="center">'+r.min.toFixed(1)+'</td>'+
      '<td class="center">'+r.max.toFixed(1)+'</td>'+
      '</tr>';
  });
  tbody.innerHTML=rows;
  document.getElementById('statSpeed').textContent=data.length+' vessels';
}

function sortSpeed(col){
  if(speedSortCol===col) speedSortDir*=-1;
  else{speedSortCol=col;speedSortDir=1;}
  renderSpeedTable();
}

function exportSpeedExcel(){
  buildVesselSpeedData();
  var data=vesselSpeedData;
  var todayStr=TODAY_DATA.date;
  var headers=SPEED_COLS.map(function(c){return c.label;});
  var numCols=headers.length;

  function thinBorder(){var s={style:'thin',color:{rgb:'BFBFBF'}};return{top:s,bottom:s,left:s,right:s};}
  var B=thinBorder();
  function F(rgb){return{patternType:'solid',fgColor:{rgb:rgb}};}
  var tS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:14},fill:F('2E75B6'),alignment:{horizontal:'center',vertical:'center'},border:B};
  var hS={font:{name:'Arial',bold:true,color:{rgb:'FFFFFF'},sz:10},fill:F('1F4E79'),alignment:{horizontal:'center',vertical:'center',wrapText:true},border:B};

  var sheetData=[];
  var tr=[];
  for(var c=0;c<numCols;c++) tr[c]={v:(c===0?'CUL VESSEL SPEED ANALYSIS  --  As of '+todayStr:''),s:tS};
  sheetData.push(tr);
  sheetData.push(headers.map(function(h){return{v:h,s:hS};}));

  data.forEach(function(r){
    var nS={font:{name:'Arial',sz:9},fill:F('FFFFFF'),border:B,alignment:{horizontal:'left',vertical:'center'}};
    var cS={font:{name:'Arial',sz:9},fill:F('FFFFFF'),border:B,alignment:{horizontal:'center',vertical:'center'}};
    var bS={font:{name:'Arial',bold:true,sz:9},fill:F('FFFFFF'),border:B,alignment:{horizontal:'left',vertical:'center'}};
    sheetData.push([
      {v:r.vessel,s:bS},
      {v:r.route,s:cS},
      {v:r.legs,s:cS},
      {v:r.avg.toFixed(1),s:{font:{name:'Arial',bold:true,sz:9},fill:F('FFFFFF'),border:B,alignment:{horizontal:'center',vertical:'center'}}},
      {v:r.min.toFixed(1),s:cS},
      {v:r.max.toFixed(1),s:cS},
    ]);
  });

  var ws=XLSX.utils.aoa_to_sheet(sheetData);
  ws['!merges']=[{s:{r:0,c:0},e:{r:0,c:numCols-1}}];
  ws['!cols']=headers.map(function(h){return{wch:Math.max(h.length+4,10)};});
  var totalRows=sheetData.length;
  ws['!rows']=[{hpt:28},{hpt:30}];for(var j=2;j<totalRows;j++)ws['!rows'].push({hpt:16});

  var wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb,ws,'Vessel Speed');
  XLSX.writeFile(wb,'CUL Vessel Speed '+todayStr+'.xlsx');
}

// Render ALL Port Wait tab tables (re-run on any filter/date/range change)
function renderPortWaitAll(){
  renderPortWaitTable();
  renderRemarkSummary();
  renderMonthlyTrend();
  renderPortWaitByRegion();
  renderPortCallCountByLane();
  renderPortCallCountByRegion();
}

// ── Init ──────────────────────────────────────────────────────────────

function initPortView(){
  // Set default date range: Jan 1 of current year → today
  var todayStr=TODAY_DATA.date;
  var yearStart = todayStr.slice(0,4) + '-01-01';
  document.getElementById('portDateFrom').value = selPortDateFrom || yearStart;
  document.getElementById('portDateTo').value = selPortDateTo || todayStr;
  selPortDateFrom = selPortDateFrom || yearStart;
  selPortDateTo = selPortDateTo || todayStr;
  buildPortFilterDropdown();
  buildRemarkFilterDropdown();
  renderPortWaitAll();
}

function initSpeedView(){
  buildSpeedVesselFilterDropdown();
  renderSpeedTable();
}

/* ═══════════════════════════════════════════════════════════════════
   HISTORY MODAL
   ═══════════════════════════════════════════════════════════════════ */
function openHistory(){
  const dates=Object.keys(SNAPSHOTS).sort().reverse();
  const listEl=document.getElementById('historyDateList');
  listEl.innerHTML='';
  dates.forEach((d,i)=>{
    const btn=document.createElement('button');
    btn.className='history-date-btn'+(i===0?' active':'');
    btn.textContent=d;
    btn.onclick=function(){document.querySelectorAll('.history-date-btn').forEach(b=>b.classList.remove('active'));btn.classList.add('active');renderHistoryTable(d);};
    listEl.appendChild(btn);
  });
  if(dates.length) renderHistoryTable(dates[0]);
  document.getElementById('historyModal').classList.add('open');
}
function renderHistoryTable(date){
  const vessels=SNAPSHOTS[date]||[];
  if(!vessels.length){document.getElementById('historyTableWrap').innerHTML='<p style="color:#8a9bb0;padding:20px">No data.</p>';return;}
  const headers=['Route','Vessel','Code','Port','Wait','Voy','ETA','ETB','ETD','Port Stay','ETA Delay','ETD Delay','PIC','Remark'];
  const keys=['route','vessel','code','port','wait','voy','eta','etb','etd','portStay','etaDelay','etdDelay','pic','remark'];
  document.getElementById('historyTableWrap').innerHTML='<table><thead><tr>'+headers.map(h=>'<th style="font-size:11px;padding:7px 8px">'+h+'</th>').join('')+'</tr></thead><tbody>'+vessels.map(r=>'<tr>'+keys.map(k=>'<td style="padding:6px 8px;border-bottom:1px solid #eee;font-size:11px">'+(r[k]||'')+'</td>').join('')+'</tr>').join('')+'</tbody></table>';
}
function closeHistory(){document.getElementById('historyModal').classList.remove('open');}
document.getElementById('historyModal').addEventListener('click',function(e){if(e.target===document.getElementById('historyModal'))closeHistory();});

/* ═══════════════════════════════════════════════════════════════════
   BOA berth-on-arrival stats (merged into Port Wait tab)
   Data rows computed from TODAY_DATA.fullSchedule, following the same
   date filter as Port Wait. Lane→Trade / Port→Region from embedded maps.
   ═══════════════════════════════════════════════════════════════════ */
var BOA_LANE_TRADE  = TODAY_DATA.laneTradeMap  || {};
var BOA_PORT_REGION = TODAY_DATA.portRegionMap || {};
var BOA_CALLS = [];
var BOA_RANGE  = 'all';
function buildBoaCalls(){
  var calls=[];
  var todayStr=TODAY_DATA.date;
  var df=selPortDateFrom||'', dt=selPortDateTo||todayStr;
  TODAY_DATA.fullSchedule.forEach(function(sr){
    var rawPort=sr.port||'';
    if(!rawPort) return;
    if(isBunkeringPort(rawPort)) return;
    var port=normalizePort(rawPort);
    if(!port) return;
    var era=portWaitDateRaw(sr);
    if(!era) return;
    if(df && era<df) return;
    if(dt && era>dt) return;
    var wait=parseFloat(sr.wait);
    if(isNaN(wait)) return;   // same rule as before: only numeric WAIT counts
    var route=sr.route||'';
    calls.push({
      t: BOA_LANE_TRADE[route] || 'Unknown',
      r: BOA_PORT_REGION[port] || 'Unknown',
      l: route || '(blank)',
      p: port,
      w: wait,
      etbRaw: sr.etbRaw || ''
    });
  });
  return calls;
}
var BOA_HEADERS = {
  boaTblTrade:  ['Trade', 'Berth', 'Over', 'Total', 'Rate'],
  boaTblLane:   ['Trade', 'Lane', 'Berth', 'Over', 'Total', 'Rate'],
  boaTblRegion: ['Port Region', 'Berth', 'Over', 'Total', 'Rate'],
  boaTblPort:   ['Port Region', 'Port', 'Berth', 'Over', 'Total', 'Rate']
};
function boaIsBerth(c){ return c.w <= berthThreshold(c.p); }  // berthed = waited within threshold (SHA/NGB 12h, others 6h)
function boaOverInRange(c){
  if(boaIsBerth(c)) return false;
  switch(BOA_RANGE){
    case 'all':   return true;            // All over = WAIT > threshold (≥6h, SHA≥12h)
    case '24+':   return c.w >= 24;       // 24h+
    case '48+':   return c.w >= 48;       // 48h+
  }
  return false;
}
function boaAgg(items){
  var berth=0, over=0;
  items.forEach(function(c){
    if(boaIsBerth(c)) berth++; else if(boaOverInRange(c)) over++;
  });
  var total = items.length;  // total calls in selected range
  return {berth:berth, over:over, total:total, rate: total>0 ? berth/total : 0};
}
function boaRateCls(r){ return r>=0.8 ? 'rate-good' : (r>=0.5 ? 'rate-mid' : 'rate-bad'); }
function boaPct(r){ return (r*100).toFixed(1)+'%'; }
function boaRowHtml(cells, cls){ return '<tr'+(cls?(' class="'+cls+'"'):'')+'>'+cells.join('')+'</tr>'; }
function boaTdNum(v){ return '<td class="num">'+v+'</td>'; }
function boaTdDim(v, cls){ return '<td class="'+(cls||'')+'">'+v+'</td>'; }
function boaTdRate(r){ return '<td class="num '+boaRateCls(r)+'">'+boaPct(r)+'</td>'; }

var boaSortState = {};
function boaBuildHeader(id){
  return '<tr>'+BOA_HEADERS[id].map(function(h){return '<th>'+h+'</th>';}).join('')+'</tr>';
}
// Render thead with current sort state (active column highlighted + arrow)
function boaRenderHeader(id){
  var s = boaSortState[id] || {col:-1, dir:1};
  var html = '<tr>' + BOA_HEADERS[id].map(function(h,i){
    if(s.col===i){
      var arrow = (s.dir===1) ? ' &#9650;' : ' &#9660;';
      return '<th style="background:#d3e4f5;color:#1F4E79;cursor:pointer;">' + h + arrow + '</th>';
    }
    return '<th style="cursor:pointer;">' + h + '</th>';
  }).join('') + '</tr>';
  document.getElementById(id).querySelector('thead').innerHTML = html;
}
function boaBuildLeaf(group, dim, row){
  var cells=[];
  if(group!==null) cells.push(boaTdDim(group,'boa-grp'));
  cells.push(boaTdDim(dim,'boa-grp'));
  cells.push(boaTdNum(row.berth), boaTdNum(row.over), boaTdNum(row.total), boaTdRate(row.rate));
  return cells;
}
function boaGrandRow(grouped){
  var cells=[];
  if(grouped) cells.push(boaTdDim('Total','boa-grp'), boaTdDim('',''));
  else cells.push(boaTdDim('Total','boa-grp'));
  cells.push(boaTdNum(BOA_STAT.berth), boaTdNum(BOA_STAT.over), boaTdNum(BOA_STAT.total), boaTdRate(BOA_STAT.rate));
  return boaRowHtml(cells, 'boa-grand');
}
function boaValAt(aggRow, col, hasGroup){
  var arr=[aggRow.berth, aggRow.over, aggRow.total, aggRow.rate];
  var idx = hasGroup ? col-2 : col-1;
  return arr[idx];
}
function boaMakeSortable(tblId, renderFn){
  var rateIdx = BOA_HEADERS[tblId].indexOf('Rate');
  boaSortState[tblId] = {col: rateIdx, dir: -1};   // default: Rate descending
  var thead = document.getElementById(tblId).querySelector('thead');
  thead.onclick = function(ev){
    var th = ev.target && ev.target.closest ? ev.target.closest('th') : null;
    if(!th) return;
    var ths = Array.prototype.slice.call(thead.querySelectorAll('th'));
    var i = ths.indexOf(th);
    if(i < 0) return;
    var s = boaSortState[tblId];
    if(s.col===i) s.dir = -s.dir; else { s.col=i; s.dir=1; }
    renderFn();
  };
}
function boaRenderTrade(){
  boaRenderHeader('boaTblTrade');
  var s = boaSortState['boaTblTrade'];
  var map = {};
  BOA_CALLS.forEach(function(c){
    var k = c.t || '(blank)';
    if(!map[k]) map[k]=[];
    map[k].push(c);
  });
  var rows = Object.keys(map).map(function(k){ return {dim:k, agg:boaAgg(map[k])}; });
  rows.sort(function(a,b){
    if(s.col===0) return (a.dim<b.dim?-1:a.dim>b.dim?1:0) * s.dir;
    var va=boaValAt(a.agg,s.col,false), vb=boaValAt(b.agg,s.col,false);
    return (va<vb?-1:va>vb?1:0) * s.dir;
  });
  var tb = document.getElementById('boaTblTrade').querySelector('tbody');
  var html = rows.map(function(r){ return boaRowHtml(boaBuildLeaf(null, r.dim, r.agg)); }).join('');
  html += boaGrandRow(false);
  tb.innerHTML = html;
}
function boaRenderRegion(){
  boaRenderHeader('boaTblRegion');
  var s = boaSortState['boaTblRegion'];
  var map = {};
  BOA_CALLS.forEach(function(c){
    var k = c.r || '(blank)';
    if(!map[k]) map[k]=[];
    map[k].push(c);
  });
  var rows = Object.keys(map).map(function(k){ return {dim:k, agg:boaAgg(map[k])}; });
  rows.sort(function(a,b){
    if(s.col===0) return (a.dim<b.dim?-1:a.dim>b.dim?1:0) * s.dir;
    var va=boaValAt(a.agg,s.col,false), vb=boaValAt(b.agg,s.col,false);
    return (va<vb?-1:va>vb?1:0) * s.dir;
  });
  var tb = document.getElementById('boaTblRegion').querySelector('tbody');
  var html = rows.map(function(r){ return boaRowHtml(boaBuildLeaf(null, r.dim, r.agg)); }).join('');
  html += boaGrandRow(false);
  tb.innerHTML = html;
}
function boaRenderGrouped(tblId, groupField, dimField){
  boaRenderHeader(tblId);
  var s = boaSortState[tblId];
  var map = {};
  BOA_CALLS.forEach(function(c){
    var g = c[groupField] || '(blank)';
    var d = c[dimField]   || '(blank)';
    var k = g+'\u0001'+d;
    if(!map[k]) map[k]=[];
    map[k].push(c);
  });
  var rows = Object.keys(map).map(function(k){
    var parts = k.split('\u0001');
    return {g:parts[0], d:parts[1], agg:boaAgg(map[k])};
  });
  var groups = []; var gSet = {};
  rows.forEach(function(r){ if(!gSet[r.g]){ gSet[r.g]=1; groups.push(r.g); } });
  var gAgg = {};
  groups.forEach(function(g){
    var items=[];
    rows.forEach(function(r){ if(r.g===g) items = items.concat(map[g+'\u0001'+r.d]); });
    gAgg[g] = boaAgg(items);
  });
  groups.sort(function(a,b){
    if(s.col===0) return (a<b?-1:a>b?1:0) * s.dir;
    if(s.col===1) return (gAgg[a].total>gAgg[b].total?-1:gAgg[a].total<gAgg[b].total?1:0) * s.dir;
    var va=boaValAt(gAgg[a],s.col,true), vb=boaValAt(gAgg[b],s.col,true);
    return (va<vb?-1:va>vb?1:0) * s.dir;
  });
  var tb = document.getElementById(tblId).querySelector('tbody');
  var html = '';
  groups.forEach(function(g){
    var leaves = rows.filter(function(r){ return r.g===g; });
    leaves.sort(function(a,b){
      if(s.col<=1) return (a.d<b.d?-1:a.d>b.d?1:0) * s.dir;
      var va=boaValAt(a.agg,s.col,true), vb=boaValAt(b.agg,s.col,true);
      return (va<vb?-1:va>vb?1:0) * s.dir;
    });
    leaves.forEach(function(r){
      html += boaRowHtml(boaBuildLeaf(g, r.d, r.agg), 'boa-subrow');
    });
    var ga = gAgg[g];
    html += boaRowHtml([
      boaTdDim(g+' Total','boa-grp'), boaTdDim('',''),
      boaTdNum(ga.berth), boaTdNum(ga.over), boaTdNum(ga.total), boaTdRate(ga.rate)
    ], 'boa-sumrow');
  });
  html += boaGrandRow(true);
  tb.innerHTML = html;
}
function boaRenderLane(){ boaRenderGrouped('boaTblLane', 't', 'l'); }
function boaRenderPort(){ boaRenderGrouped('boaTblPort', 'r', 'p'); }

var BOA_STAT = {berth:0, over:0, total:0, rate:0};
function boaRefresh(){
  BOA_STAT = boaAgg(BOA_CALLS);
  document.getElementById('boaChipTotal').textContent = BOA_CALLS.length;
  document.getElementById('boaChipBerth').textContent = BOA_STAT.berth;
  document.getElementById('boaChipOver').textContent  = BOA_STAT.over;
  document.getElementById('boaChipRate').textContent  = boaPct(BOA_STAT.rate);
  boaRenderTrade(); boaRenderRegion(); boaRenderLane(); boaRenderPort();
}
function boaInit(){
  BOA_CALLS = buildBoaCalls();
  var lab = document.getElementById('boaLabelSpan');
  if(lab) lab.textContent = '· computed from CUL Daily Movement (follows Port Wait date range)';
  Object.keys(BOA_HEADERS).forEach(function(id){
    boaSortState[id] = {col: BOA_HEADERS[id].indexOf('Rate'), dir: -1};
    boaRenderHeader(id);
  });
  boaMakeSortable('boaTblTrade', boaRenderTrade);
  boaMakeSortable('boaTblRegion', boaRenderRegion);
  boaMakeSortable('boaTblLane', boaRenderLane);
  boaMakeSortable('boaTblPort', boaRenderPort);
  document.querySelectorAll('input[name=boarange]').forEach(function(rb){
    rb.addEventListener('change', function(){
      BOA_RANGE = this.value;
      document.querySelectorAll('.range-opt').forEach(function(l){
        var inp = l.querySelector('input');
        if(inp && inp.name==='boarange') l.classList.toggle('sel', inp.checked);
      });
      boaRefresh();
    });
  });
  // Port Wait over-range selector (3 categories)
  document.querySelectorAll('input[name=portrange]').forEach(function(rb){
    rb.addEventListener('change', function(){
      PORT_OVER_RANGE = this.value;
      document.querySelectorAll('.range-opt').forEach(function(l){
        var inp = l.querySelector('input');
        if(inp && inp.name==='portrange') l.classList.toggle('sel', inp.checked);
      });
      renderPortWaitAll();
    });
  });
  boaRefresh();
}

/* ═══════════════════════════════════════════════════════════════════
   LOGIN
   ═══════════════════════════════════════════════════════════════════ */
var LOGIN_PWD='CUL1234';
function doLogin(){
  var v=document.getElementById('loginPwd').value.trim();
  if(v===LOGIN_PWD){
    sessionStorage.setItem('cul_auth','1');
    document.getElementById('loginOverlay').classList.add('hidden');
    init();
  } else {
    document.getElementById('loginErr').textContent='Incorrect password';
    document.getElementById('loginPwd').value='';
  }
}
// Check if already authenticated this session — invoked at the END of the
// script (see bottom) so all top-level data (MAINT_RECORDS, TODAY_DATA, …)
// is assigned before init() runs. Running init() too early (parse time)
// left MAINT_RECORDS undefined on reload and broke the Maintenance view.

/* ═══════════════════════════════════════════════════════════════════
   Maintenance Rate (Vessel Schedule / Port Log 维护率)
   MAINT_DATA.records[] = {service, vessel, voyage, dir, operator,
      port, etd, plog('Y'/'N'), vsched(0/1)}
   Port Log 维护率 = plog=='Y' / total
   Vessel Schedule (Actual Schedule) 维护率 = vsched==1 ("Maintain timely") / total
   ═══════════════════════════════════════════════════════════════════ */
var MAINT_RECORDS = ((typeof MAINT_DATA !== 'undefined') && MAINT_DATA && MAINT_DATA.records) || [];
// 上游是否仍导出 Vessel Schedule(Maintain timely) 列; 2026-08 起该列停供时, vsched 相关统计显示 '—'
var VSCHED_AVAILABLE = !(MAINT_DATA && MAINT_DATA.vschedAvailable === false);
  document.getElementById('maintSourceTs').textContent = (MAINT_DATA && MAINT_DATA.sourceMtime) ? MAINT_DATA.sourceMtime : '—';
var maintOpFilter = 'CUL';      // 默认 CUL（用户要求）
var selMaintFrom = '', selMaintTo = '';

// ── Maintenance Vessel / Service multi-select filters (mirror Port Wait port filter) ──
var MAINT_FILTER_DEFS = {
  service: { btn:'maintServiceFilterBtn', dd:'maintServiceFilterDropdown', search:'maintServiceFilterSearchBox', lsKey:'cul_movement_maint_service', nullText:'All Services', field:'service' },
  vessel:  { btn:'maintVesselFilterBtn',  dd:'maintVesselFilterDropdown',  search:'maintVesselFilterSearchBox',  lsKey:'cul_movement_maint_vessel',  nullText:'All Vessels',  field:'vessel' }
};
var selMaintServiceFilter = null;  // null = all services
var selMaintVesselFilter  = null;  // null = all vessels

function getMaintFilterVal(which){ return which==='service' ? selMaintServiceFilter : selMaintVesselFilter; }
function setMaintFilterVal(which, v){ if(which==='service') selMaintServiceFilter = v; else selMaintVesselFilter = v; }

function buildMaintFilterDropdown(which){
  var def = MAINT_FILTER_DEFS[which];
  var dd = document.getElementById(def.dd);
  var values = {};
  MAINT_RECORDS.forEach(function(r){ var v=(r[def.field]||'').trim(); if(v) values[v]=true; });
  var sorted = Object.keys(values).sort();
  var sel = getMaintFilterVal(which);
  var kw = (document.getElementById(def.search) ? document.getElementById(def.search).value : '').toLowerCase();
  var html = '<div style="padding:2px 8px 6px;border-bottom:1px solid #e4ecf5;margin-bottom:4px;">';
  html += '<input type="text" id="'+def.search+'" placeholder="&#128269; Search..." value="'+kw+'" oninput="onMaintFilterSearchBox(\''+which+'\')" style="width:100%;padding:4px 8px;border:1px solid #c9d5e2;border-radius:4px;font-size:12px;box-sizing:border-box;">';
  html += '</div>';
  html += '<label style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;">';
  html += '<input type="checkbox" value="__all__" '+(sel===null?'checked':'')+' onchange="onMaintFilterAllChange(\''+which+'\',this)">';
  html += '<b>'+def.nullText+'</b></label>';
  sorted.forEach(function(v){
    var checked = sel===null || sel.indexOf(v)>=0;
    var match = !kw || v.toLowerCase().indexOf(kw)>=0;
    html += '<label data-v="'+v+'" style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;'+(match?'':'display:none;')+'">';
    html += '<input type="checkbox" value="'+v+'" '+(checked?'checked':'')+' onchange="onMaintFilterItemChange(\''+which+'\',this)">';
    html += v+'</label>';
  });
  dd.innerHTML = html;
  updateMaintFilterButton(which);
}

function onMaintFilterSearchBox(which){
  var def = MAINT_FILTER_DEFS[which];
  var kw = (document.getElementById(def.search).value||'').toLowerCase();
  document.querySelectorAll('#'+def.dd+' label[data-v]').forEach(function(lb){
    var v = lb.getAttribute('data-v');
    lb.style.display = (!kw || v.toLowerCase().indexOf(kw)>=0) ? 'flex' : 'none';
  });
}

function toggleMaintFilter(which){
  var def = MAINT_FILTER_DEFS[which];
  var dd = document.getElementById(def.dd);
  if(!dd.classList.contains('open')){ buildMaintFilterDropdown(which); dd.classList.add('open'); }
  else dd.classList.remove('open');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){ if(d!==dd) d.classList.remove('open'); });
}
function toggleMaintServiceFilter(){ toggleMaintFilter('service'); }
function toggleMaintVesselFilter(){ toggleMaintFilter('vessel'); }

function updateMaintFilterButton(which){
  var def = MAINT_FILTER_DEFS[which];
  var checks = document.querySelectorAll('#'+def.dd+' input[type=checkbox]:not([value="__all__"])');
  var total=0, sel=[];
  checks.forEach(function(cb){ total++; if(cb.checked) sel.push(cb.value); });
  var btn = document.getElementById(def.btn);
  if(sel.length===total){ btn.textContent = def.nullText; btn.classList.remove('has-selection'); }
  else { btn.textContent = sel.length + ' selected'; btn.classList.add('has-selection'); }
}

function applyMaintFilterFromDOM(which){
  var def = MAINT_FILTER_DEFS[which];
  var allCb = document.querySelector('#'+def.dd+' input[value="__all__"]');
  var checks = document.querySelectorAll('#'+def.dd+' input[type=checkbox]:not([value="__all__"])');
  var total=0, sel=[];
  checks.forEach(function(cb){ total++; if(cb.checked) sel.push(cb.value); });
  if(sel.length===total){ setMaintFilterVal(which, null); if(allCb) allCb.checked=true; }
  else { setMaintFilterVal(which, sel); if(allCb) allCb.checked=false; }
  updateMaintFilterButton(which);
  saveMaintFilter(which);
  renderMaint();
}

function onMaintFilterAllChange(which, allCb){
  var def = MAINT_FILTER_DEFS[which];
  var checks = document.querySelectorAll('#'+def.dd+' input[type=checkbox]:not([value="__all__"])');
  checks.forEach(function(cb){ cb.checked = allCb.checked; });
  if(allCb.checked){ setMaintFilterVal(which, null); }
  else { setMaintFilterVal(which, []); }
  updateMaintFilterButton(which);
  saveMaintFilter(which);
  renderMaint();
}

function onMaintFilterItemChange(which, cb){
  applyMaintFilterFromDOM(which);
}

// ── Maintenance Port filter (mirrors Port Wait: checkbox dropdown + free-text search) ──
var selMaintPortFilter = null;  // null = show all ports
var selMaintPortSearch = '';    // free-text port search (AND with checkbox filter)

function onMaintPortSearch(){
  selMaintPortSearch = (document.getElementById('maintPortSearch').value || '').trim();
  renderMaint();
}

function buildMaintPortFilterDropdown(){
  var dd = document.getElementById('maintPortFilterDropdown');
  var allPorts = {};
  MAINT_RECORDS.forEach(function(r){ var p = (r.port || '').trim(); if(p) allPorts[p] = true; });
  var sorted = Object.keys(allPorts).sort();
  var kw = (document.getElementById('maintPortFilterSearchBox') ? document.getElementById('maintPortFilterSearchBox').value : '').toLowerCase();
  var html = '<div style="padding:2px 8px 6px;border-bottom:1px solid #e4ecf5;margin-bottom:4px;">';
  html += '<input type="text" id="maintPortFilterSearchBox" placeholder="&#128269; Search ports..." value="'+kw+'" oninput="onMaintPortFilterSearchBox()" style="width:100%;padding:4px 8px;border:1px solid #c9d5e2;border-radius:4px;font-size:12px;box-sizing:border-box;">';
  html += '</div>';
  html += '<label style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;">';
  html += '<input type="checkbox" value="__all__" '+(selMaintPortFilter===null?'checked':'')+' onchange="onMaintPortFilterAllChange(this)">';
  html += '<b>All Ports</b></label>';
  sorted.forEach(function(p){
    var checked = selMaintPortFilter===null || selMaintPortFilter.indexOf(p)>=0;
    var match = !kw || p.toLowerCase().indexOf(kw)>=0;
    html += '<label data-p="'+p+'" style="display:flex;align-items:center;gap:6px;padding:4px 8px;cursor:pointer;white-space:nowrap;'+(match?'':'display:none;')+'">';
    html += '<input type="checkbox" value="'+p+'" '+(checked?'checked':'')+' onchange="onMaintPortFilterItemChange(this)">';
    html += p+'</label>';
  });
  dd.innerHTML = html;
  updateMaintPortFilterButton();
}

function onMaintPortFilterSearchBox(){
  var kw = (document.getElementById('maintPortFilterSearchBox').value || '').toLowerCase();
  document.querySelectorAll('#maintPortFilterDropdown label[data-p]').forEach(function(lb){
    var p = lb.getAttribute('data-p');
    lb.style.display = (!kw || p.toLowerCase().indexOf(kw)>=0) ? 'flex' : 'none';
  });
}

function toggleMaintPortFilter(){
  var dd = document.getElementById('maintPortFilterDropdown');
  if(!dd.classList.contains('open')){ buildMaintPortFilterDropdown(); dd.classList.add('open'); }
  else dd.classList.remove('open');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){ if(d!==dd) d.classList.remove('open'); });
}

function updateMaintPortFilterButton(){
  var checks = document.querySelectorAll('#maintPortFilterDropdown input[type=checkbox]:not([value="__all__"])');
  var total = 0, sel = [];
  checks.forEach(function(cb){ total++; if(cb.checked) sel.push(cb.value); });
  var btn = document.getElementById('maintPortFilterBtn');
  if(sel.length === total){ btn.textContent = 'All Ports'; btn.classList.remove('has-selection'); }
  else { btn.textContent = sel.length + ' of ' + total + ' ports'; btn.classList.add('has-selection'); }
}

function applyMaintPortFilterFromDOM(){
  var allCb = document.querySelector('#maintPortFilterDropdown input[value="__all__"]');
  var checks = document.querySelectorAll('#maintPortFilterDropdown input[type=checkbox]:not([value="__all__"])');
  var total = 0, sel = [];
  checks.forEach(function(cb){ total++; if(cb.checked) sel.push(cb.value); });
  if(sel.length === total){ selMaintPortFilter = null; if(allCb) allCb.checked = true; }
  else { selMaintPortFilter = sel; if(allCb) allCb.checked = false; }
  updateMaintPortFilterButton();
  renderMaint();
}

function onMaintPortFilterAllChange(allCb){
  var checks = document.querySelectorAll('#maintPortFilterDropdown input[type=checkbox]:not([value="__all__"])');
  checks.forEach(function(cb){ cb.checked = allCb.checked; });
  selMaintPortFilter = allCb.checked ? null : [];
  updateMaintPortFilterButton();
  renderMaint();
}

function onMaintPortFilterItemChange(cb){
  applyMaintPortFilterFromDOM();
}

function saveMaintFilter(which){
  var def = MAINT_FILTER_DEFS[which];
  try { localStorage.setItem(def.lsKey, JSON.stringify(getMaintFilterVal(which))); } catch(e){}
}
function loadMaintFilter(which){
  var def = MAINT_FILTER_DEFS[which];
  if(!MAINT_RECORDS) return;
  try {
    var raw = localStorage.getItem(def.lsKey);
    if(raw === null) return;
    var v = JSON.parse(raw);
    var avail = {};
    MAINT_RECORDS.forEach(function(r){ var x=(r[def.field]||'').trim(); if(x) avail[x]=true; });
    if(v === null){ setMaintFilterVal(which, null); }
    else if(Array.isArray(v)){
      var cleaned = v.filter(function(x){ return avail[x]; });
      setMaintFilterVal(which, cleaned.length ? cleaned : []);
    } else { setMaintFilterVal(which, null); }
  } catch(e){ setMaintFilterVal(which, null); }
}
function refreshMaintFilterButtons(){
  ['service','vessel'].forEach(function(which){
    var def = MAINT_FILTER_DEFS[which];
    var sel = getMaintFilterVal(which);
    var btn = document.getElementById(def.btn);
    if(sel === null){ btn.textContent = def.nullText; btn.classList.remove('has-selection'); }
    else { btn.textContent = sel.length + ' selected'; btn.classList.add('has-selection'); }
  });
}

function maintFiltered(){
  var f = selMaintFrom || '', t = selMaintTo || '';
  return MAINT_RECORDS.filter(function(r){
    if(maintOpFilter && maintOpFilter !== 'ALL' && r.operator !== maintOpFilter) return false;
    var e = r.etd || '';
    if(f && e && e < f) return false;
    if(t && e && e > t) return false;
    if(selMaintServiceFilter && selMaintServiceFilter.indexOf((r.service||'').trim()) < 0) return false;
    if(selMaintVesselFilter && selMaintVesselFilter.indexOf((r.vessel||'').trim()) < 0) return false;
    var _mp = (r.port || '').trim();
    if(selMaintPortFilter && selMaintPortFilter.indexOf(_mp) < 0) return false;
    if(selMaintPortSearch && _mp.toLowerCase().indexOf(selMaintPortSearch.toLowerCase()) < 0) return false;
    return true;
  });
}
function maintPct(n, d){ return d > 0 ? (n / d * 100) : 0; }
function maintRateCls(r){ return r >= 80 ? 'rate-good' : (r >= 50 ? 'rate-mid' : 'rate-bad'); }
function maintRateCell(pct){ return '<td class="num ' + maintRateCls(pct) + '">' + pct.toFixed(1) + '%</td>'; }
function maintNumCell(v){ return '<td class="num">' + v + '</td>'; }

function renderMaintChips(recs){
  var total = recs.length, plog = 0, vs = 0;
  recs.forEach(function(r){ if(r.plog === 'Y') plog++; if(r.vsched === 1) vs++; });
  var plogPct = maintPct(plog, total), vsPct = maintPct(vs, total);
  var plogNo = total - plog, vsNo = total - vs;
  document.getElementById('maintChipTotal').textContent = total.toLocaleString();
  document.getElementById('maintChipPortLog').textContent = plog.toLocaleString() + ' (' + plogPct.toFixed(1) + '%)';
  document.getElementById('maintChipPortLogNo').textContent = plogNo.toLocaleString() + ' (' + maintPct(plogNo, total).toFixed(1) + '%)';
  document.getElementById('maintChipVSched').textContent = VSCHED_AVAILABLE
    ? vs.toLocaleString() + ' (' + vsPct.toFixed(1) + '%)' : '— (列已停供)';
  document.getElementById('maintChipVSchedNo').textContent = VSCHED_AVAILABLE
    ? vsNo.toLocaleString() + ' (' + maintPct(vsNo, total).toFixed(1) + '%)' : '—';
  document.getElementById('statMaint').textContent = total.toLocaleString() + ' calls';
}

function maintPortId(port){ return 'maintPortDetail-' + String(port).replace(/[^A-Za-z0-9_-]/g,'_'); }

function buildMaintDetailBlocks(g){
  // Show ALL calls for this port, with Port Log / Vessel Sched status per row
  var arr = g.all.slice().sort(function(a,b){
    return (a.etd||'').localeCompare(b.etd||'') || (a.service||'').localeCompare(b.service||'') || (a.vessel||'').localeCompare(b.vessel||'');
  });
  var h = '<div style="font-size:11px;font-weight:600;color:#1F4E79;margin:8px 0 2px;">All calls (' + arr.length + ') &mdash; ' +
    'Port Log: <b style="color:#2E7D32;">Y</b> maintained / <b style="color:#C0392B;">N</b> not maintained; ' +
    (VSCHED_AVAILABLE
      ? 'Vessel Sched: <b style="color:#2E7D32;">Maintain timely</b> / <b style="color:#C0392B;">Not maintained</b>'
      : 'Vessel Sched: <b style="color:#8a9bb0;">— 列已停供(上游不再导出)</b>') + '</div>' +
    '<table style="width:100%;border-collapse:collapse;font-size:11px;margin:2px 0 8px;">' +
    '<thead><tr style="background:#e8eff6;color:#1F4E79;"><th style="padding:3px 6px;text-align:left;">Service</th><th style="padding:3px 6px;text-align:left;">Vessel</th><th style="padding:3px 6px;text-align:left;">Voyage</th><th style="padding:3px 6px;text-align:left;">Dir</th><th style="padding:3px 6px;text-align:left;">Operator</th><th style="padding:3px 6px;text-align:left;">Port</th><th style="padding:3px 6px;text-align:left;">ETD</th><th style="padding:3px 6px;text-align:left;">Port Log</th><th style="padding:3px 6px;text-align:left;">Vessel Sched</th></tr></thead><tbody>';
  arr.forEach(function(r, i){
    var bg = (i % 2 === 0) ? '' : ' style="background:#f0f6fc;"';
    var plogCls = (r.plog === 'Y') ? 'color:#2E7D32;' : 'color:#C0392B;font-weight:600;';
    var vsCls   = !VSCHED_AVAILABLE ? 'color:#8a9bb0;' : ((r.vsched === 1) ? 'color:#2E7D32;' : 'color:#C0392B;font-weight:600;');
    h += '<tr' + bg + '>' +
      '<td style="padding:3px 6px;">' + (r.service||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.vessel||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.voyage||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.dir||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.operator||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.port||'') + '</td>' +
      '<td style="padding:3px 6px;">' + (r.etd||'') + '</td>' +
      '<td style="padding:3px 6px;' + plogCls + '">' + (r.plog==='Y' ? 'Y' : 'N') + '</td>' +
      '<td style="padding:3px 6px;' + vsCls + '">' + (!VSCHED_AVAILABLE ? '—' : (r.vsched===1 ? 'Maintain timely' : 'Not maintained')) + '</td></tr>';
  });
  return h + '</tbody></table>';
}

function toggleMaintPort(port){
  var row = document.getElementById(maintPortId(port));
  if(!row) return;
  var open = (row.style.display === 'none');
  row.style.display = open ? '' : 'none';
  var tr = row.previousElementSibling;
  if(tr){ var sp = tr.querySelector('.maint-toggle-arrow'); if(sp) sp.innerHTML = open ? '&#9652;' : '&#9656;'; }
}

function expandAllMaintPort(){
  document.querySelectorAll('#maintPortTbody tr.detail-wrap').forEach(function(row){
    row.style.display = '';
    var sp = row.previousElementSibling ? row.previousElementSibling.querySelector('.maint-toggle-arrow') : null;
    if(sp) sp.innerHTML = '&#9652;';
  });
}
function collapseAllMaintPort(){
  document.querySelectorAll('#maintPortTbody tr.detail-wrap').forEach(function(row){
    row.style.display = 'none';
    var sp = row.previousElementSibling ? row.previousElementSibling.querySelector('.maint-toggle-arrow') : null;
    if(sp) sp.innerHTML = '&#9656;';
  });
}

function renderMaintByPort(){
  var recs = maintFiltered();
  var groups = {};
  recs.forEach(function(r){
    var k = r.port || '(blank)';
    if(!groups[k]) groups[k] = {key:k, calls:0, plog:0, vs:0, unmaintained:[], all:[]};
    var g = groups[k];
    g.calls++;
    g.all.push(r);
    if(r.plog === 'Y') g.plog++;
    if(r.vsched === 1) g.vs++;
    if(r.plog !== 'Y' || (VSCHED_AVAILABLE && r.vsched !== 1)) g.unmaintained.push(r);
  });
  var rows = Object.keys(groups).map(function(k){ return groups[k]; });
  rows.sort(function(a,b){ return b.calls - a.calls; });

  document.getElementById('maintPortThead').innerHTML =
    '<tr><th></th><th>Port</th><th class="num">Calls</th><th class="num">Port Log Maint</th>' +
    '<th class="num">Port Log Rate</th><th class="num">Vessel Sched Maint</th><th class="num">Vessel Sched Rate</th><th class="num">Unmaintained</th></tr>';

  var tbody = document.getElementById('maintPortTbody');
  if(!rows.length){
    tbody.innerHTML = '<tr><td colspan="8" style="text-align:center;color:#8a9bb0;padding:14px;">No data for current filters.</td></tr>';
    return;
  }
  var html = '';
  rows.forEach(function(g, i){
    var plogPct = maintPct(g.plog, g.calls), vsPct = maintPct(g.vs, g.calls);
    var styleAttr = (i % 2 === 0) ? 'background:#EBF3FB;' : '';
    var hasDetail = g.calls > 0;
    if(hasDetail) styleAttr += 'cursor:pointer;';
    var bg = styleAttr ? ' style="' + styleAttr + '"' : '';
    var arrow = hasDetail ? '<span class="maint-toggle-arrow" style="color:#1F4E79;">&#9656;</span>' : '';
    var click = hasDetail ? ' onclick="toggleMaintPort(\'' + String(g.key).replace(/'/g,"\\'") + '\')"' : '';
    html += '<tr' + bg + click + '>' +
      '<td style="text-align:center;width:24px;">' + arrow + '</td>' +
      '<td>' + g.key + '</td>' +
      maintNumCell(g.calls) + maintNumCell(g.plog) + maintRateCell(plogPct) +
      (VSCHED_AVAILABLE ? maintNumCell(g.vs) + maintRateCell(vsPct)
                        : '<td class="num" style="color:#8a9bb0;">—</td><td class="num" style="color:#8a9bb0;">—</td>') +
      maintNumCell(g.unmaintained.length) + '</tr>';
    if(hasDetail){
      html += '<tr class="detail-wrap" id="' + maintPortId(g.key) + '" style="display:none;"><td colspan="8" style="padding:4px 8px 4px 28px;background:#fafcff;">' +
        buildMaintDetailBlocks(g) + '</td></tr>';
    }
  });
  tbody.innerHTML = html;
}

// ── Per-Port Maintenance Rate (dedicated rate summary) ──────────────────
var maintRateSort = {key:'plogRate', dir:1};  // default: Port Log Rate ascending (worst first)

function maintRateBar(pct){
  var cls = maintRateCls(pct);
  var color = pct >= 80 ? '#2e8b57' : (pct >= 50 ? '#e67e22' : '#c0392b');
  var w = Math.max(1, Math.min(100, pct));
  return '<td class="num ' + cls + '">' +
    '<div style="display:inline-block;width:72px;height:8px;background:#e9eef3;border-radius:4px;vertical-align:middle;overflow:hidden;margin-right:7px;">' +
    '<div style="height:100%;width:' + w + '%;background:' + color + ';"></div></div>' +
    pct.toFixed(1) + '%</td>';
}

function maintRateSortTable(key){
  if(maintRateSort.key === key){ maintRateSort.dir = -maintRateSort.dir; }
  else { maintRateSort.key = key; maintRateSort.dir = (key === 'port') ? 1 : -1; }
  renderMaintPortRate();
}

function renderMaintPortRate(){
  var recs = maintFiltered();
  var groups = {};
  recs.forEach(function(r){
    var k = r.port || '(blank)';
    if(!groups[k]) groups[k] = {key:k, calls:0, plog:0, vs:0};
    groups[k].calls++;
    if(r.plog === 'Y') groups[k].plog++;
    if(r.vsched === 1) groups[k].vs++;
  });
  var rows = Object.keys(groups).map(function(k){
    var g = groups[k];
    return {key:k, calls:g.calls, plog:g.plog, vs:g.vs,
            plogRate: maintPct(g.plog, g.calls), vsRate: maintPct(g.vs, g.calls)};
  });
  var s = maintRateSort;
  rows.sort(function(a,b){
    if(s.key === 'port'){ var va=a.key, vb=b.key; return (va<vb?-1:va>vb?1:0)*s.dir; }
    var va = a[s.key], vb = b[s.key];
    return (va<vb?-1:va>vb?1:0)*s.dir;
  });

  var arrow = function(col){ return s.key === col ? (s.dir > 0 ? ' ▲' : ' ▼') : ''; };
  document.getElementById('maintRateThead').innerHTML =
    '<tr>' +
    '<th onclick="maintRateSortTable(\'port\')">Port' + arrow('port') + '</th>' +
    '<th class="num" onclick="maintRateSortTable(\'calls\')">Calls' + arrow('calls') + '</th>' +
    '<th class="num" onclick="maintRateSortTable(\'plogRate\')">Port Log Rate' + arrow('plogRate') + '</th>' +
    '<th class="num" onclick="maintRateSortTable(\'vsRate\')">Vessel Sched Rate' + arrow('vsRate') + '</th>' +
    '</tr>';

  var tbody = document.getElementById('maintRateTbody');
  if(!rows.length){
    tbody.innerHTML = '<tr><td colspan="4" style="text-align:center;color:#8a9bb0;padding:14px;">No data for current filters.</td></tr>';
    return;
  }
  var html = '';
  rows.forEach(function(g, i){
    var bg = (i % 2 === 0) ? ' style="background:#EBF3FB;"' : '';
    html += '<tr' + bg + '>' +
      '<td>' + g.key + '</td>' +
      maintNumCell(g.calls) +
      maintRateBar(g.plogRate) +
      (VSCHED_AVAILABLE ? maintRateBar(g.vsRate) : '<td class="num" style="color:#8a9bb0;">—</td>') + '</tr>';
  });
  tbody.innerHTML = html;
}

var maintSort = { plogNo: {key:'etd', dir:1}, vsNo: {key:'etd', dir:1} };

function maintSortTable(which, key){
  var st = maintSort[which];
  if(st.key === key){ st.dir = -st.dir; } else { st.key = key; st.dir = 1; }
  renderMaintUnmaintained();
}

function renderMaintUnmaintained(){
  var recs = maintFiltered();
  var plogNo = [], vsNo = [];
  recs.forEach(function(r){
    if(r.plog !== 'Y') plogNo.push(r);
    if(VSCHED_AVAILABLE && r.vsched !== 1) vsNo.push(r);
  });

  // Search filters (match any text column, case-insensitive)
  var qPlog = (document.getElementById('maintPlogNoSearch').value || '').trim().toLowerCase();
  var qVs   = (document.getElementById('maintVsNoSearch').value || '').trim().toLowerCase();
  function match(r, q){
    if(!q) return true;
    return ['service','vessel','voyage','dir','operator','port','etd'].some(function(k){
      return String(r[k]||'').toLowerCase().indexOf(q) !== -1;
    });
  }
  if(qPlog) plogNo = plogNo.filter(function(r){ return match(r, qPlog); });
  if(qVs)   vsNo   = vsNo.filter(function(r){ return match(r, qVs); });

  // Sort by current column/direction
  function sortArr(arr, st){
    arr.sort(function(a,b){
      return String(a[st.key]||'').localeCompare(String(b[st.key]||''), 'zh', {numeric:true}) * st.dir;
    });
  }
  sortArr(plogNo, maintSort.plogNo);
  sortArr(vsNo, maintSort.vsNo);

  var COLS = [['service','Service'],['vessel','Vessel'],['voyage','Voyage'],['dir','Dir'],
              ['operator','Operator'],['port','Port'],['etd','ETD'],['agent','Agent / OP']];
  function agentCellHtml(port){
    var list = AGENT_BY_PORT[port] || [];
    if(!list.length) return '<td style="color:#8a9bb0;font-size:11px;">—</td>';
    var inner = list.map(function(c){
      var bits = ['<b>'+escapeHtml(c.name||'')+'</b>'];
      if(c.email)  bits.push('<a href="mailto:'+escapeHtml(c.email)+'" style="color:#1F4E79;">'+escapeHtml(c.email)+'</a>');
      if(c.tel)    bits.push('Tel: '+escapeHtml(c.tel));
      if(c.mobile) bits.push('Mob: '+escapeHtml(c.mobile));
      return '<div style="line-height:1.4;margin:2px 0;">'+bits.join('<br>')+'</div>';
    }).join('');
    return '<td style="font-size:11px;vertical-align:top;">'+inner+'</td>';
  }
  function headHtml(which){
    var st = maintSort[which], h = '<tr>';
    COLS.forEach(function(c){
      var arrow = (st.key === c[0]) ? (st.dir === 1 ? ' &#9650;' : ' &#9660;') : '';
      h += '<th style="cursor:pointer;white-space:nowrap;user-select:none;" onclick="maintSortTable(\'' + which + '\',\'' + c[0] + '\')">' + c[1] + arrow + '</th>';
    });
    h += '<th>Status</th></tr>';
    return h;
  }
  document.getElementById('maintPlogNoThead').innerHTML = headHtml('plogNo');
  document.getElementById('maintVsNoThead').innerHTML = headHtml('vsNo');

  function rowsHtml(arr, statusTxt, isFiltered){
    if(!arr.length){
      var msg = isFiltered
        ? 'No rows match the current search filter.'
        : 'All maintained &#8212; no unmaintained calls for current filters.';
      return '<tr><td colspan="9" style="text-align:center;color:#8a9bb0;padding:12px;">' + msg + '</td></tr>';
    }
    var h = '';
    arr.forEach(function(r, i){
      var bg = (i % 2 === 0) ? ' style="background:#fdf2f2;"' : '';
      h += '<tr' + bg + '>' +
        '<td>' + (r.service||'') + '</td>' +
        '<td>' + (r.vessel||'') + '</td>' +
        '<td>' + (r.voyage||'') + '</td>' +
        '<td>' + (r.dir||'') + '</td>' +
        '<td>' + (r.operator||'') + '</td>' +
        '<td>' + (r.port||'') + '</td>' +
        '<td>' + (r.etd||'') + '</td>' +
        agentCellHtml(r.port) +
        '<td style="color:#C0392B;font-weight:600;">' + statusTxt + '</td></tr>';
    });
    return h;
  }
  document.getElementById('maintPlogNoTbody').innerHTML = rowsHtml(plogNo, 'N', !!qPlog);
  document.getElementById('maintVsNoTbody').innerHTML = VSCHED_AVAILABLE
    ? rowsHtml(vsNo, 'Not maintained', !!qVs)
    : '<tr><td colspan="9" style="text-align:center;color:#8a9bb0;padding:12px;">Vessel Schedule 列已停供（上游 2026-08 起不再导出），无法统计 Not maintained。</td></tr>';
  document.getElementById('maintPlogNoCount').textContent = '(' + plogNo.length.toLocaleString() + ' calls)';
  document.getElementById('maintVsNoCount').textContent = VSCHED_AVAILABLE ? '(' + vsNo.length.toLocaleString() + ' calls)' : '—';
}

function renderMaintMonth(){
  var recs = maintFiltered();
  var byMonth = {};
  recs.forEach(function(r){
    var m = (r.etd || '').substring(0,7);
    if(!m) return;
    if(!byMonth[m]) byMonth[m] = {month:m, calls:0, plog:0, vs:0};
    var g = byMonth[m];
    g.calls++;
    if(r.plog === 'Y') g.plog++;
    if(r.vsched === 1) g.vs++;
  });
  var months = Object.keys(byMonth).sort();

  var barsHtml = '';
  months.forEach(function(m){
    var g = byMonth[m];
    var plogPct = maintPct(g.plog, g.calls), vsPct = maintPct(g.vs, g.calls);
    barsHtml +=
      '<div style="font-size:11px;font-weight:600;color:#555;margin-top:2px;">' + m + ' <span style="color:#8a9bb0;font-weight:400;">(' + g.calls + ' calls)</span></div>' +
      '<div style="display:flex;align-items:center;gap:6px;margin:2px 0;">' +
        '<span style="display:inline-block;width:92px;font-size:10px;color:#1F4E79;text-align:right;">Port Log</span>' +
        '<div style="flex:1;background:#eef2f7;border-radius:3px;height:14px;overflow:hidden;"><div style="width:' + plogPct.toFixed(1) + '%;background:#2E75B6;height:100%;"></div></div>' +
        '<span style="width:50px;font-size:10px;color:#2E75B6;">' + plogPct.toFixed(1) + '%</span></div>' +
      (VSCHED_AVAILABLE
        ? '<div style="display:flex;align-items:center;gap:6px;margin:2px 0 6px;">' +
          '<span style="display:inline-block;width:92px;font-size:10px;color:#C55A11;text-align:right;">Vessel Sched</span>' +
          '<div style="flex:1;background:#eef2f7;border-radius:3px;height:14px;overflow:hidden;"><div style="width:' + vsPct.toFixed(1) + '%;background:#ED7D31;height:100%;"></div></div>' +
          '<span style="width:50px;font-size:10px;color:#C55A11;">' + vsPct.toFixed(1) + '%</span></div>'
        : '');
  });
  document.getElementById('maintMonthBars').innerHTML = barsHtml ||
    '<div style="color:#8a9bb0;font-size:11px;">No data for current filters.</div>';

  document.getElementById('maintMonthThead').innerHTML =
    '<tr><th>Month</th><th class="num">Calls</th><th class="num">Port Log Rate</th><th class="num">Vessel Sched Rate</th></tr>';
  var tbody = document.getElementById('maintMonthTbody');
  if(!months.length){
    tbody.innerHTML = '<tr><td colspan="4" style="text-align:center;color:#8a9bb0;padding:14px;">No data for current filters.</td></tr>';
    return;
  }
  var html = '';
  months.forEach(function(m, i){
    var g = byMonth[m];
    var plogPct = maintPct(g.plog, g.calls), vsPct = maintPct(g.vs, g.calls);
    var bg = (i % 2 === 0) ? ' style="background:#EBF3FB;"' : '';
    html += '<tr' + bg + '><td>' + m + '</td>' + maintNumCell(g.calls) + maintRateCell(plogPct) +
      (VSCHED_AVAILABLE ? maintRateCell(vsPct) : '<td class="num" style="color:#8a9bb0;">—</td>') + '</tr>';
  });
  tbody.innerHTML = html;
}

function renderMaint(){
  renderMaintChips(maintFiltered());
  renderMaintPortRate();
  renderMaintByPort();
  renderMaintUnmaintained();
  renderMaintMonth();
}

function onMaintChange(){
  maintOpFilter = document.getElementById('maintOpFilter').value || 'ALL';
  selMaintFrom = document.getElementById('maintFrom').value || '';
  selMaintTo = document.getElementById('maintTo').value || '';
  renderMaint();
}

function initMaintView(){
  if(!MAINT_RECORDS || !MAINT_RECORDS.length){
    document.getElementById('statMaint').textContent = 'Maintenance data unavailable (' +
      (MAINT_DATA ? MAINT_DATA.source : 'no source') + ').';
    return;
  }
  // Operator dropdown: ALL + distinct operators (CUL default)
  var ops = {};
  MAINT_RECORDS.forEach(function(r){ if(r.operator) ops[r.operator] = (ops[r.operator]||0)+1; });
  var opList = Object.keys(ops).sort(function(a,b){ return ops[b]-ops[a]; });
  var sel = document.getElementById('maintOpFilter');
  sel.innerHTML = '<option value="ALL">All Operators</option>' +
    opList.map(function(o){ return '<option value="'+o+'"'+(o==='CUL'?' selected':'')+'>'+o+' ('+ops[o]+')</option>'; }).join('');
  // ETD range defaults = year-to-date (Jan 1 of current year → today)
  var _now = new Date();
  var _ymd = function(d){ var y=d.getFullYear(),m=('0'+(d.getMonth()+1)).slice(-2),dd=('0'+d.getDate()).slice(-2); return y+'-'+m+'-'+dd; };
  var lo = _now.getFullYear() + '-01-01';
  var hi = _ymd(_now);
  document.getElementById('maintFrom').value = selMaintFrom || lo;
  document.getElementById('maintTo').value = selMaintTo || hi;
  selMaintFrom = selMaintFrom || lo;
  selMaintTo = selMaintTo || hi;
  maintOpFilter = 'CUL';
  sel.value = 'CUL';
  loadMaintFilter('service');
  loadMaintFilter('vessel');
  refreshMaintFilterButtons();
  renderMaint();
}

/* ═══════════════════════════════════════════════════════════════════
   INIT
   ═══════════════════════════════════════════════════════════════════ */
// Keep sticky table headers below the pinned top bar (header+tabs) + controls
function updateCtrlH(){
  var top=document.querySelector('.top-pinned');
  var topH=top?top.offsetHeight:105;
  document.documentElement.style.setProperty('--top-h', topH+'px');
  var v=document.querySelector('.tab-content.active .controls');
  var h=v?v.offsetHeight:62;
  document.documentElement.style.setProperty('--ctrl-h', h+'px');
}
window.addEventListener('resize', updateCtrlH);

function init(){
  loadSnapshots(); saveSnapshot(TODAY_DATA);
  // 同步 "Show 已下线" 勾选框的初始视觉与状态
  document.querySelectorAll('.decom-cb').forEach(function(b){ b.checked = SHOW_DECOM; });
  document.querySelectorAll('.decom-toggle').forEach(function(lbl){
    lbl.classList.toggle('on', SHOW_DECOM);
    var d=lbl.querySelector('.decom-dot'); if(d) d.textContent = SHOW_DECOM ? '✓' : '+';
  });
  var _hdrV = (TODAY_DATA.vessels||[]).filter(function(r){return !isDecommissioned(r.vessel);});
  var _hdrF = (TODAY_DATA.fullSchedule||[]).filter(function(r){return !isDecommissioned(r.vessel);});
  document.getElementById('headerDate').textContent='Data as of '+TODAY_DATA.date+'  |  Updated '+TODAY_DATA.generatedAt+'  |  '+_hdrV.length+' vessels  |  '+_hdrF.length+' schedule rows';
  document.getElementById('footerTs').textContent='Data updated: '+TODAY_DATA.generatedAt;
  initSummary();
  initFullSchedule();
  initPortView();
  initSpeedView();
  boaInit();
  initMaintView();
  updateCtrlH();
}
// init() only called after login success (see LOGIN section above)

// ── Bootstrap: if already authenticated this session, render immediately ──
// Placed at the very end so every top-level data assignment (MAINT_RECORDS,
// TODAY_DATA, COLUMN_DEFS_*, …) has executed before init() runs. Running it
// earlier (at parse time) left MAINT_RECORDS undefined on reload.
if(sessionStorage.getItem('cul_auth') === '1'){
  var _lo = document.getElementById('loginOverlay');
  if(_lo) _lo.classList.add('hidden');
  init();
}
</script>
</body>
</html>"""

# ── Main ──────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--excel', default=DEFAULT_EXCEL)
    parser.add_argument('--out',   default=DEFAULT_HTML)
    args = parser.parse_args()

    excel_path = args.excel
    out_path   = args.out

    if not os.path.exists(os.path.dirname(excel_path)):
        excel_path = os.path.join(SCRIPT_DIR, 'CUL DAILY MOVEMENT.xlsx')
    if not os.path.exists(os.path.dirname(out_path)):
        out_path = os.path.join(SCRIPT_DIR, 'cul_daily_movement.html')

    print(f'Reading: {excel_path}')
    data = extract(excel_path)
    print(f'  -> {len(data["vessels"])} vessels (summary)')
    print(f'  -> {len(data["fullSchedule"])} schedule rows (full)')
    print(f'  -> date={data["date"]}')

    boa_lane_trade, boa_port_region = load_boa_mappings()
    data['laneTradeMap']  = boa_lane_trade
    data['portRegionMap'] = boa_port_region

    maint = resolve_maint(out_path)
    agent = load_agent_contacts()

    # 生成后自检：MAINT 有数据但 port 列超一半是时间戳 → 列映射错位，禁止发布
    _maint_recs = maint.get('records') or []
    if _maint_recs:
        _ports = [str(r.get('port') or '') for r in _maint_recs]
        _ts = [p for p in _ports if re.match(r'^\d{4}-\d{2}-\d{2}', p)]
        if _ts and len(_ts) / max(1, len(_ports)) >= 0.5:
            print('  [MAINT] ERROR: port column misaligned (%d/%d values are timestamps) — '
                  'aborting publish to avoid shipping a broken dashboard'
                  % (len(_ts), len(_ports)), file=sys.stderr, flush=True)
            sys.exit(1)

    # Build column defs JSON for JS
    col_defs_summary = [{"key":c[0],"label":c[1],"defaultVisible":c[2]} for c in SUMMARY_COLUMNS]
    col_defs_full    = [{"key":c[0],"label":c[1],"defaultVisible":c[3]} for c in FULL_COLUMNS]

    html = HTML_TEMPLATE
    html = html.replace('__TODAY_DATA__',       json.dumps(data, ensure_ascii=False))
    html = html.replace('__COLUMN_DEFS_SUMMARY__', json.dumps(col_defs_summary, ensure_ascii=False))
    html = html.replace('__COLUMN_DEFS_FULL__',    json.dumps(col_defs_full,    ensure_ascii=False))
    html = html.replace('__MAINT_DATA__',          json.dumps(maint, ensure_ascii=False))
    html = html.replace('__AGENT_BY_PORT__',       json.dumps(agent, ensure_ascii=False))

    os.makedirs(os.path.dirname(out_path) if os.path.dirname(out_path) else '.', exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'Saved : {out_path}')

if __name__ == '__main__':
    main()
