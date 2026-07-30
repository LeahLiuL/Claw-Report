import openpyxl
from collections import Counter
from datetime import datetime

F = r'P:/04 上海操作中心/01 船期管理科/船期管理/VSL Daily Movement/更新/CUL DAILY MOVEMENT.rebuilt.xlsx'
wb = openpyxl.load_workbook(F)
ws = wb.active

def norm(s):
    return (s or '').strip().upper()

# 找所有船块: 块头行 r (C1=航线码, 且 r+1 行 C1='PORT')
blocks = []
r = 1
while r < ws.max_row:
    c1 = ws.cell(r, 1).value
    nxt = ws.cell(r + 1, 1).value
    if c1 and norm(c1) != 'PORT' and nxt and norm(nxt) == 'PORT':
        blocks.append(r)
    r += 1

# ASR 块校验
print('=== ASR 块 (列映射修正) ===')
asr = next((b for b in blocks if norm(ws.cell(b, 1).value) == 'ASR'), None)
if asr:
    print('块头 C1/C4/C9/C16:', ws.cell(asr,1).value, '|', ws.cell(asr,4).value, '|', ws.cell(asr,9).value, '|', ws.cell(asr,16).value)
    print('表头:', [ws.cell(asr+1, c).value for c in range(1, 12)])
    for k in range(2, 5):
        rr = asr + 1 + k
        print(f'  行{rr} C1..C11:',
              [ (str(ws.cell(rr,c).value)[:16] if ws.cell(rr,c).value is not None else None) for c in range(1,12)])
else:
    print('  未找到 ASR 块')

# 全局: 港口数据行的 Voy(C7)/Date(C8)/ETB(C10) 类型
print('\n=== 全局港口数据行日期/航次类型 ===')
voy_empty = 0; voy_total = 0
c8_serial = 0; c8_dt = 0; c8_other = 0
c10_dt = 0; c10_other = 0
for b in blocks:
    rr = b + 2  # 第一个数据行
    while rr <= ws.max_row:
        c1 = ws.cell(rr, 1).value
        if c1 is None:
            rr += 1; continue
        if norm(c1) == 'PORT':
            break  # 下一块表头
        if str(c1).strip().startswith('Remark'):
            break
        # 这是港口数据行
        voy_total += 1
        v = ws.cell(rr, 7).value
        if not v or str(v).strip() == '':
            voy_empty += 1
        # C8 Date
        d8 = ws.cell(rr, 8).value
        if isinstance(d8, datetime): c8_dt += 1
        elif isinstance(d8, (int, float)): c8_serial += 1
        else: c8_other += 1
        # C10 ETB
        d10 = ws.cell(rr, 10).value
        if isinstance(d10, datetime): c10_dt += 1
        else: c10_other += 1
        rr += 1

print(f'港口数据行总数: {voy_total}')
print(f'Voy.No 空值: {voy_empty}  (向上就近后应趋近0)')
print(f'Date(C8): datetime={c8_dt}  serial残留={c8_serial}  其他(文本标记)={c8_other}')
print(f'ETB(C10): datetime={c10_dt}  其他={c10_other}')

# 列出仍空 Voy.No 的港口行(抽样)
print('\n=== 仍为空 Voy.No 的港口行(前10) ===')
n = 0
for b in blocks:
    rr = b + 2
    while rr <= ws.max_row:
        c1 = ws.cell(rr, 1).value
        if c1 is None:
            rr += 1; continue
        if norm(c1) == 'PORT': break
        if str(c1).strip().startswith('Remark'): break
        v = ws.cell(rr, 7).value
        if not v or str(v).strip() == '':
            print(f'  {ws.cell(b,4).value} | 港口={c1} | 块行{rr}')
            n += 1
            if n >= 10: break
        rr += 1
    if n >= 10: break
