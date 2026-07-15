#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Full processing for ALL sheets of Vessel Bapfile.xlsx.

Two modes:
  * Default (full rebuild): drops bapfile.db and reloads everything from the
    xlsx. Used for initial setup / backfill.
  * --append (incremental): keeps the existing cumulative db and only inserts
    rows that are NOT an exact duplicate of an existing row. The FTP source is
    a rolling window that overlaps previous data, so we ACCUMULATE, never
    replace. "Exact duplicate" = all 17 data columns identical (any field
    different is kept as a new row).

Uses a chunked regex (C-backed) over the decompressed worksheet XML.
Column mapping is dynamic: the header row (row 1) is read first to build a
col_index -> col_name map, so files with different column layouts (e.g. SFTP
source with POL_ETB between POL_CD and POD_CD) are handled correctly.
"""
import argparse
import json
import os
import re
import sqlite3
import datetime
import zipfile
from collections import Counter

SRC = os.environ.get("BAPFILE_LOCAL") or os.path.join(os.path.dirname(os.path.abspath(__file__)), "Vessel Bapfile.xlsx")
HERE = os.path.dirname(os.path.abspath(__file__))
OUT_AGG = os.path.join(HERE, "agg.json")
OUT_DB = os.path.join(HERE, "bapfile.db")

COLS = ["VSL_CD","SCH_VOY_NR","SCH_DIR_CD","VVD","REVENUE_MONTH","LANE","CONT_NR",
        "TARGET_PORT","CARRIER_ID","FE_FLG","POR_CD","POL_CD","POD_CD","DEL_CD",
        "CONT_TP_SIZE_CD","CONT_WT","EDI_COC_SOC","BKG_COC_SOC","FIXED_FLG","AWK_FLG",
        "DG_FLG","RF_FLG","BB_FLG","TEU","BL_NO","UNIT","SLOT_OWN_PTR_ID","CONT_OPR_PTR_ID"]

# The 17 columns actually stored in bapfile.db. An "exact duplicate" is judged
# on ALL of these (any difference => keep as a new row).
DB_COLS = ["vvd","lane","container_no","fe","pol","pod","type_size","weight",
           "awk","dg","rf","bb","slot_opr","cont_opr","rev_month","sheet","target_port"]

COL_DEFS = (
    "id INTEGER PRIMARY KEY, vvd TEXT, lane TEXT, container_no TEXT, fe TEXT, "
    "pol TEXT, pod TEXT, type_size TEXT, weight REAL, awk TEXT, dg TEXT, rf TEXT, "
    "bb TEXT, slot_opr TEXT, cont_opr TEXT, rev_month TEXT, sheet INTEGER, target_port TEXT"
)
INSERT_SQL = (
    "INSERT OR IGNORE INTO bapfile ("
    + ",".join(DB_COLS)
    + ") VALUES (" + ",".join(["?"] * len(DB_COLS)) + ")"
)
UNIQ_COLS = ",".join(DB_COLS)


def col_to_idx(letters):
    idx = 0
    for ch in letters:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


EXCEL_EPOCH = datetime.datetime(1899, 12, 30)
_MONTH_RE = re.compile(r'^(\d{4})[-/](\d{1,2})')

def build_column_maps(xlsx_path, sheet_names):
    """Read header row (row 1) of each sheet and return a list of
    {col_index: col_name} dicts plus the actual sheet names.

    Uses openpyxl to read just the header row. This makes the parser
    robust against column layout differences between source files
    (e.g. SFTP file has POL_ETB between POL_CD and POD_CD, while the
    local 近4月 file puts POL_ETB at the far right).

    Header names are normalised so that source files with the same
    meaning but different spelling/typos (e.g. TAGERT_PORT) map to a
    canonical key (e.g. TARGET_PORT) used by process_row.
    """
    import openpyxl as _opxl
    wb = _opxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    col_maps = []
    actual_names = []
    normalise = {
        "TAGERT_PORT": "TARGET_PORT",
    }
    for i, sname in enumerate(sheet_names):
        if i < len(wb.sheetnames):
            actual_names.append(wb.sheetnames[i])
        else:
            actual_names.append(sname)
        if i < len(wb.sheetnames):
            ws = wb[wb.sheetnames[i]]
            headers = [cell.value for cell in ws[1]]
            col_map = {idx: normalise.get(str(h), str(h)) for idx, h in enumerate(headers) if h is not None}
        else:
            col_map = {}
        col_maps.append(col_map)
    wb.close()
    return actual_names, col_maps
def to_month(val):
    if not val:
        return None
    val = str(val).strip()
    m = _MONTH_RE.match(val)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}"
    try:
        d = EXCEL_EPOCH + datetime.timedelta(days=int(round(float(val))))
        return d.strftime("%Y-%m")
    except Exception:
        return None


# ---- aggregate accumulators (used by full mode) ----
total_rows = 0
rows_by_sheet = {}
month_rows = Counter(); month_teu = Counter(); month_wt = Counter(); month_unit = Counter()
lane_rows = Counter(); lane_teu = Counter(); lane_wt = Counter()
carrier_rows = Counter(); carrier_teu = Counter()
vessel_rows = Counter(); vessel_teu = Counter()
dir_rows = Counter()
ctype_rows = Counter(); ctype_teu = Counter()
fe_rows = Counter(); fe_teu = Counter()
reefer_rows = 0; dg_rows = 0; awk_rows = 0; bb_rows = 0
edisoc = Counter(); bkgsoc = Counter()
pol_rows = Counter(); pod_rows = Counter()
slot_rows = Counter(); opr_rows = Counter()
tot_teu = 0.0; tot_wt = 0.0; tot_unit = 0.0
n_full = 0; n_empty = 0
distinct_vvd = set()
samples = []

CELL_RE = re.compile(r'<c r="([A-Z]+)\d+"[^>]*?(?:>(?:<is><t>(.*?)</t></is>|<v>(.*?)</v>)?</c>|/>)')
ROW_SPLIT = b'</row>'

# ---- SQLite ----
con = None  # set in main
batch = []
BATCH_SIZE = 100000


def flush_batch():
    global batch
    if batch:
        con.executemany(INSERT_SQL, batch)
        batch = []


def ensure_schema(append):
    global con
    con.execute(f"CREATE TABLE IF NOT EXISTS bapfile ({COL_DEFS})")
    if append:
        has_uniq = con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='index' AND name='ix_row_uniq'"
        ).fetchone()
        if not has_uniq:
            migrate_dedup_existing()
        for sql in (
            "CREATE INDEX IF NOT EXISTS ix_vvd ON bapfile(vvd)",
            "CREATE INDEX IF NOT EXISTS ix_lane ON bapfile(lane)",
            "CREATE INDEX IF NOT EXISTS ix_month ON bapfile(rev_month)",
            "CREATE INDEX IF NOT EXISTS ix_lane_month ON bapfile(lane, rev_month)",
        ):
            con.execute(sql)
        con.commit()


def migrate_dedup_existing():
    """One-time: remove exact-duplicate rows from an existing db that was built
    before the unique index existed. Keeps the earliest id per duplicate group."""
    global con
    before = con.execute("SELECT COUNT(*) FROM bapfile").fetchone()[0]
    con.execute(
        "CREATE TABLE bapfile_tmp AS "
        "SELECT * FROM ("
        f"  SELECT *, ROW_NUMBER() OVER (PARTITION BY {UNIQ_COLS} ORDER BY id) rn "
        "  FROM bapfile) WHERE rn = 1"
    )
    con.execute("DROP TABLE bapfile")
    con.execute("ALTER TABLE bapfile_tmp RENAME TO bapfile")
    con.execute(f"CREATE UNIQUE INDEX ix_row_uniq ON bapfile({UNIQ_COLS})")
    after = con.execute("SELECT COUNT(*) FROM bapfile").fetchone()[0]
    print(f"[DEDUP] removed {before - after} exact-duplicate rows ({before} -> {after})", flush=True)


def process_row(d, sheet_idx, sheet_name):
    global total_rows, reefer_rows, dg_rows, awk_rows, bb_rows, tot_teu, tot_wt, tot_unit
    global n_full, n_empty
    total_rows += 1

    mon = to_month(d.get("REVENUE_MONTH","")) if d.get("REVENUE_MONTH") else None
    try: teu = float(d.get("TEU") or 0)
    except Exception: teu = 0.0
    try: wt = float(d.get("CONT_WT") or 0)
    except Exception: wt = 0.0
    try: unit = float(d.get("UNIT") or 0)
    except Exception: unit = 0.0

    if mon:
        month_rows[mon]+=1; month_teu[mon]+=teu; month_wt[mon]+=wt; month_unit[mon]+=unit

    lane = d.get("LANE") or "(空)"
    lane_rows[lane]+=1; lane_teu[lane]+=teu; lane_wt[lane]+=wt

    carrier = d.get("CARRIER_ID") or "(空)"
    carrier_rows[carrier]+=1; carrier_teu[carrier]+=teu

    vsl = d.get("VSL_CD") or "(空)"
    vessel_rows[vsl]+=1; vessel_teu[vsl]+=teu

    dir_rows[d.get("SCH_DIR_CD") or "(空)"]+=1

    ctype = d.get("CONT_TP_SIZE_CD") or "(空)"
    ctype_rows[ctype]+=1; ctype_teu[ctype]+=teu

    fe = d.get("FE_FLG") or "(空)"
    fe_rows[fe]+=1; fe_teu[fe]+=teu
    if fe=="Full": n_full+=1
    elif fe=="Empty": n_empty+=1

    if (d.get("RF_FLG") or "N")=="Y": reefer_rows+=1
    if (d.get("DG_FLG") or "N")=="Y": dg_rows+=1
    if (d.get("AWK_FLG") or "N")=="Y": awk_rows+=1
    if (d.get("BB_FLG") or "N")=="Y": bb_rows+=1

    edisoc[d.get("EDI_COC_SOC") or "(空)"]+=1
    bkgsoc[d.get("BKG_COC_SOC") or "(空)"]+=1
    pol_rows[d.get("POL_CD") or "(空)"]+=1
    pod_rows[d.get("POD_CD") or "(空)"]+=1
    slot_rows[d.get("SLOT_OWN_PTR_ID") or "(空)"]+=1
    opr_rows[d.get("CONT_OPR_PTR_ID") or "(空)"]+=1

    tot_teu+=teu; tot_wt+=wt; tot_unit+=unit
    vvd = d.get("VVD")
    if vvd: distinct_vvd.add(vvd)
    if sheet_idx == 0 and total_rows <= 200 and len(samples) < 8:
        samples.append({c: d.get(c,"") for c in COLS})

    # detail row
    target_port = d.get("TARGET_PORT") or d.get("TAGERT_PORT") or d.get("VSL_SCH_PORT_CD") or ""

    batch.append((
        vvd, lane, d.get("CONT_NR") or "", fe,
        d.get("POL_CD") or "", d.get("POD_CD") or "", d.get("CONT_TP_SIZE_CD") or "", wt,
        d.get("AWK_FLG") or "", d.get("DG_FLG") or "", d.get("RF_FLG") or "", d.get("BB_FLG") or "",
        d.get("SLOT_OWN_PTR_ID") or "", d.get("CONT_OPR_PTR_ID") or "", mon, sheet_idx + 1,
        target_port
    ))
    if len(batch) >= BATCH_SIZE:
        flush_batch()


def parse_sheet(stream, sheet_idx, sheet_name, col_map):
    global total_rows
    before = total_rows
    pending = b""
    chunk = 1 << 24
    while True:
        data = stream.read(chunk)
        if not data:
            break
        buf = pending + data
        parts = buf.split(ROW_SPLIT)
        pending = parts[-1]
        for seg in parts[:-1]:
            seg = seg + ROW_SPLIT
            if b"VSL_CD" in seg:
                continue
            try:
                s = seg.decode("utf-8")
            except Exception:
                s = seg.decode("utf-8", "replace")
            cells = []
            for m in CELL_RE.finditer(s):
                col = m.group(1)
                val = m.group(2) if m.group(2) is not None else (m.group(3) or "")
                cells.append((col, val))
            if cells:
                d = {}
                for col, val in cells:
                    idx = col_to_idx(col)
                    col_name = col_map.get(idx)
                    if col_name:
                        d[col_name] = val
                process_row(d, sheet_idx, sheet_name)
    if pending.strip():
        seg = pending + ROW_SPLIT
        if b"VSL_CD" not in seg:
            try:
                s = seg.decode("utf-8")
            except Exception:
                s = seg.decode("utf-8", "replace")
            cells = []
            for m in CELL_RE.finditer(s):
                col = m.group(1)
                val = m.group(2) if m.group(2) is not None else (m.group(3) or "")
                cells.append((col, val))
            if cells:
                d = {}
                for col, val in cells:
                    idx = col_to_idx(col)
                    col_name = col_map.get(idx)
                    if col_name:
                        d[col_name] = val
                process_row(d, sheet_idx, sheet_name)
    rows_by_sheet[sheet_name] = {"data_rows": total_rows - before}
    print(f"[OK] {sheet_name}: {total_rows-before} data rows", flush=True)


def build_agg_from_db():
    """Recompute agg.json from the FULL cumulative db (used in append mode,
    because the parsed xlsx is only a rolling window). Dimensions that are not
    stored in the db (carrier / vessel / dir / teu / unit / coc_soc) are emitted
    empty; run a full rebuild to refresh those."""
    global con
    total = con.execute("SELECT COUNT(*) FROM bapfile").fetchone()[0]
    dvvd = con.execute("SELECT COUNT(DISTINCT vvd) FROM bapfile").fetchone()[0]
    by_sheet = dict(con.execute("SELECT sheet, COUNT(*) FROM bapfile GROUP BY sheet").fetchall())
    rows_by_sheet_local = {f"sheet{k}": {"data_rows": v} for k, v in sorted(by_sheet.items())}

    month = dict(con.execute("SELECT rev_month, COUNT(*) FROM bapfile GROUP BY rev_month").fetchall())
    lane = dict(con.execute("SELECT lane, COUNT(*) FROM bapfile GROUP BY lane").fetchall())
    ctype = dict(con.execute("SELECT type_size, COUNT(*) FROM bapfile GROUP BY type_size").fetchall())
    fe = dict(con.execute("SELECT fe, COUNT(*) FROM bapfile GROUP BY fe").fetchall())
    pol = dict(con.execute("SELECT pol, COUNT(*) FROM bapfile GROUP BY pol ORDER BY COUNT(*) DESC LIMIT 25").fetchall())
    pod = dict(con.execute("SELECT pod, COUNT(*) FROM bapfile GROUP BY pod ORDER BY COUNT(*) DESC LIMIT 25").fetchall())
    slot = dict(con.execute("SELECT slot_opr, COUNT(*) FROM bapfile GROUP BY slot_opr ORDER BY COUNT(*) DESC LIMIT 20").fetchall())
    opr = dict(con.execute("SELECT cont_opr, COUNT(*) FROM bapfile GROUP BY cont_opr ORDER BY COUNT(*) DESC LIMIT 20").fetchall())
    reefer = con.execute("SELECT COUNT(*) FROM bapfile WHERE rf='Y'").fetchone()[0]
    dg = con.execute("SELECT COUNT(*) FROM bapfile WHERE dg='Y'").fetchone()[0]
    awk = con.execute("SELECT COUNT(*) FROM bapfile WHERE awk='Y'").fetchone()[0]
    bb = con.execute("SELECT COUNT(*) FROM bapfile WHERE bb='Y'").fetchone()[0]
    n_full = con.execute("SELECT COUNT(*) FROM bapfile WHERE fe='Full'").fetchone()[0]
    n_empty = con.execute("SELECT COUNT(*) FROM bapfile WHERE fe='Empty'").fetchone()[0]
    total_wt = con.execute("SELECT COALESCE(SUM(weight),0) FROM bapfile").fetchone()[0]
    samp = con.execute(
        "SELECT vvd,lane,container_no,fe,pol,pod,type_size,weight,awk,dg,rf,bb,"
        "slot_opr,cont_opr,rev_month,target_port FROM bapfile LIMIT 8"
    ).fetchall()
    samples_local = [
        {"VVD": r[0],"LANE": r[1],"CONT_NR": r[2],"FE_FLG": r[3],"POL_CD": r[4],"POD_CD": r[5],
         "CONT_TP_SIZE_CD": r[6],"CONT_WT": r[7],"AWK_FLG": r[8],"DG_FLG": r[9],"RF_FLG": r[10],
         "BB_FLG": r[11],"SLOT_OWN_PTR_ID": r[12],"CONT_OPR_PTR_ID": r[13],"REVENUE_MONTH": r[14],
         "TARGET_PORT": r[15]}
        for r in samp
    ]

    result = {
        "total_rows": total,
        "distinct_vvd": dvvd,
        "rows_by_sheet": rows_by_sheet_local,
        "totals": {"teu": 0, "weight_tons": round(total_wt, 1),
                   "unit": 0, "n_full": n_full, "n_empty": n_empty},
        "month": {"rows": month, "teu": {}, "weight_tons": {k: round(v, 1) for k, v in month.items()}, "unit": {}},
        "lane": {"rows": dict(Counter(lane).most_common()), "teu": {}, "weight_tons": {}},
        "carrier": {"rows": {}, "teu": {}},
        "vessel": {"rows": {}, "teu": {}},
        "dir": {},
        "cont_type": {"rows": dict(Counter(ctype).most_common()), "teu": {}},
        "fe": {"rows": fe, "teu": {}},
        "flags": {"reefer": reefer, "dg": dg, "awk": awk, "bb": bb},
        "coc_soc": {"edi": {}, "bkg": {}},
        "ports": {"pol_top": pol, "pod_top": pod},
        "slot_owner": slot,
        "cont_operator": opr,
        "samples": samples_local,
        "_note": "agg.json rebuilt from cumulative db in append mode; carrier/vessel/dir/teu/unit/coc_soc omitted (not stored in db). Run full rebuild to refresh.",
    }
    with open(OUT_AGG, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[AGG] rebuilt agg.json from db: {total} rows", flush=True)


def main():
    global con
    ap = argparse.ArgumentParser()
    ap.add_argument("--append", action="store_true",
                    help="incremental: keep existing db, INSERT OR IGNORE exact-duplicate rows")
    ap.add_argument("--db", default=OUT_DB)
    ap.add_argument("--src", default=SRC)
    args = ap.parse_args()

    con = sqlite3.connect(args.db)
    con.execute("PRAGMA synchronous=OFF")

    if args.append:
        ensure_schema(append=True)
    else:
        con.execute("DROP TABLE IF EXISTS bapfile")
        con.execute(f"CREATE TABLE bapfile ({COL_DEFS})")
        con.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS ix_row_uniq ON bapfile({UNIQ_COLS})")

    zf = zipfile.ZipFile(args.src)
    wb = zf.read("xl/workbook.xml").decode("utf-8")
    names = re.findall(r'<sheet[^>]*name="([^"]+)"', wb)
    # Read header row of every sheet to build dynamic column maps.
    # This handles different column layouts between source files
    # (e.g. SFTP file has POL_ETB between POL_CD and POD_CD).
    actual_names, col_maps = build_column_maps(args.src, names)
    sheet_entries = [n for n in zf.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", n)]
    sheet_entries.sort(key=lambda x: int(re.search(r"sheet(\d+)\.xml$", x).group(1)))
    for i, entry in enumerate(sheet_entries):
        name = actual_names[i] if i < len(actual_names) else entry
        with zf.open(entry) as s:
            parse_sheet(s, i, name, col_maps[i])
    flush_batch()
    con.commit()

    if args.append:
        # Append mode only accumulates rows into the db; gen_static.py rebuilds
        # the public shards from it. The dashboard agg.json is NOT served on
        # GitHub Pages and its dimensions are incomplete in append mode, so we
        # skip rebuilding it here (also avoids the legacy dict() crash).
        pass
    else:
        print("[INDEX] creating indexes ...", flush=True)
        for sql in (
            "CREATE INDEX IF NOT EXISTS ix_vvd ON bapfile(vvd)",
            "CREATE INDEX IF NOT EXISTS ix_lane ON bapfile(lane)",
            "CREATE INDEX IF NOT EXISTS ix_month ON bapfile(rev_month)",
            "CREATE INDEX IF NOT EXISTS ix_lane_month ON bapfile(lane, rev_month)",
        ):
            con.execute(sql)
        con.commit()
        write_agg_from_accumulators()

    con.close()
    print(f"\n[DONE] total_rows={total_rows} distinct_vvd={len(distinct_vvd)} -> {args.db}", flush=True)


def write_agg_from_accumulators():
    result = {
        "total_rows": total_rows,
        "distinct_vvd": len(distinct_vvd),
        "rows_by_sheet": rows_by_sheet,
        "totals": {"teu": round(tot_teu,1), "weight_tons": round(tot_wt,1),
                   "unit": round(tot_unit,1), "n_full": n_full, "n_empty": n_empty},
        "month": {"rows": dict(month_rows), "teu": {k: round(v,1) for k,v in month_teu.items()},
                  "weight_tons": {k: round(v,1) for k,v in month_wt.items()},
                  "unit": {k: round(v,1) for k,v in month_unit.items()}},
        "lane": {"rows": dict(lane_rows.most_common()), "teu": {k: round(v,1) for k,v in lane_teu.most_common()},
                 "weight_tons": {k: round(v,1) for k,v in lane_wt.most_common()}},
        "carrier": {"rows": dict(carrier_rows.most_common()), "teu": {k: round(v,1) for k,v in carrier_teu.most_common()}},
        "vessel": {"rows": dict(vessel_rows.most_common()), "teu": {k: round(v,1) for k,v in vessel_teu.most_common()}},
        "dir": dict(dir_rows),
        "cont_type": {"rows": dict(ctype_rows.most_common()), "teu": {k: round(v,1) for k,v in ctype_teu.most_common()}},
        "fe": {"rows": dict(fe_rows), "teu": {k: round(v,1) for k,v in fe_teu.items()}},
        "flags": {"reefer": reefer_rows, "dg": dg_rows, "awk": awk_rows, "bb": bb_rows},
        "coc_soc": {"edi": dict(edisoc.most_common()), "bkg": dict(bkgsoc.most_common())},
        "ports": {"pol_top": dict(pol_rows.most_common(25)), "pod_top": dict(pod_rows.most_common(25))},
        "slot_owner": dict(slot_rows.most_common(20)),
        "cont_operator": dict(opr_rows.most_common(20)),
        "samples": samples,
    }
    with open(OUT_AGG, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f"[DONE] agg.json -> {OUT_AGG}", flush=True)


if __name__ == "__main__":
    main()
