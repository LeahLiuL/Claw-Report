#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_daily_movement.py
=====================================================================
自动刷新 CUL DAILY MOVEMENT.xlsx（运营用的"大Excel"）。

【策略】沿用现有大Excel作为模板：
  - 保留原文件的分组(#VALUE!段标题)、船顺序、航线码、船代码、PIC、
    列头、备注行、以及每艘船当前挑选的港口行窗口（100%结构一致）
  - 只把每个船块对应的源文件（2026/下每文件夹"更新时间最新"的xlsx
    的"第一个sheet"）里的港口行数值，刷新进模板对应位置

这样既能自动同步最新船期数据，又不会破坏人工编排好的结构/顺序/窗口。

【数据映射】源文件第一个sheet 的 C1..C16 直接对应大Excel 的 C1..C16：
  C1=PORT C2=man in C3=wait C4=Proforma C5=ltm eta C6=ltm etd
  C7=VOY.NO C8=date C9=ETA C10=ETB C11=ETD C12=run
  C13=Port Stay(hr) C14=fsp distance C15=speed C16=ETA Delay/Ahead
（大Excel 丢弃了源的 C17=ETD Delay/Ahead、C18=REMARK，列头只到 C16）

【匹配规则】每个船块用 (PORT, VOY.NO) 作为主键，去源文件 lookup 里找同一条
港口行；找到就刷新 C1..C16，找不到则保留原值（不破坏）。

【船块→源文件夹】用"船名 或 船代码"双匹配定位 2026/ 下的文件夹：
  - 船名直接相等（忽略大小写/空格）
  - 船名别名表（块里写法 vs 文件夹名）
  - 船代码(C9) 相等
（DONG FANG MIN HAI / HONG DA XIN 728 这类拼写差异靠别名+代码兜底）

用法：
  # 这台电脑（P: 盘）
  python build_daily_movement.py --src "P:\\04 上海操作中心\\01 船期管理科\\船期管理\\VSL Daily Movement\\更新\\2026" --template "P:\\04 上海操作中心\\01 船期管理科\\船期管理\\VSL Daily Movement\\更新\\CUL DAILY MOVEMENT.xlsx"

  # culadmin 那台（Z: 盘，默认即 Z:，可直接 python build_daily_movement.py）

  # 先生成到测试文件核对，不覆盖原件：
  python build_daily_movement.py --output "CUL DAILY MOVEMENT.generated.xlsx"

  # 确认无误后覆盖原件（会自动先备份为 .bak_时间戳）：
  python build_daily_movement.py
=====================================================================
"""
import argparse
import os
import glob
import shutil
from datetime import datetime

import openpyxl


# ---- 默认路径（culadmin 那台电脑用 Z: 盘；这台用 --src/--template 覆盖）----
DEFAULT_SRC = r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\2026"
DEFAULT_TEMPLATE = r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\CUL DAILY MOVEMENT.xlsx"

# 块里船名 -> 文件夹名 的别名（拼写/缩写差异兜底）
NAME_ALIASES = {
    "HONG DA XIN 728": "HDX 728",
    "DONG FANG MIN HAI": "DONG FANG MING HAI",
}


def norm(s):
    """归一化：去空格、转大写。"""
    return (s or "").strip().upper()


def norm_voy(s):
    """航次号归一化：去空格、转大写（2627 W -> 2627W）。容错非字符串类型。"""
    if s is None:
        return ""
    return str(s).strip().upper().replace(" ", "")


def latest_xlsx(folder):
    """返回文件夹下 mtime 最新的 xlsx（排除 ~$ 临时锁文件）。"""
    files = [f for f in glob.glob(os.path.join(folder, "*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return files[0]


def read_source_sheet(path):
    """
    读取源文件第一个sheet。
    返回 (code, lookup)。
      code   : R1 C9 的船代码（可能为空）
      lookup : {(port, voy): [c1..c16 值]}  仅港口数据行
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    code = ws.cell(1, 9).value  # R1 C9

    # 找列头行（C1 == 'PORT'）
    header_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if norm(ws.cell(r, 1).value) == "PORT":
            header_row = r
            break
    lookup = {}
    if header_row is None:
        print(f"    [WARN] 源文件无 PORT 列头，跳过数据: {os.path.basename(path)}")
        return code, lookup

    for r in range(header_row + 1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None:
            continue
        s1 = norm(c1)
        if s1 == "PORT":          # 源文件含多段航次表，重复PORT列头需跳过继续读后续数据行
            continue
        if s1 == "" or s1.startswith("REMARK"):
            continue
        voy = ws.cell(r, 7).value
        key = (s1, norm_voy(voy))
        vals = [ws.cell(r, c).value for c in range(1, 17)]  # C1..C16
        lookup[key] = vals
    return code, lookup


def build_folder_index(src_dir):
    """扫描 2026/ 下所有船文件夹，构建 folder -> {code, lookup, file} 索引。"""
    folders = sorted([
        d for d in os.listdir(src_dir)
        if os.path.isdir(os.path.join(src_dir, d)) and d != "已下线船舶"
    ])
    index = {}
    for fol in folders:
        folder = os.path.join(src_dir, fol)
        path = latest_xlsx(folder)
        if not path:
            print(f"  [WARN] 文件夹无 xlsx，跳过: {fol}")
            continue
        code, lookup = read_source_sheet(path)
        index[fol] = {"code": code, "lookup": lookup, "file": path}
    return index


def find_folder(vessel, code, index):
    """用 船名(含别名) 或 船代码 定位源文件夹。返回 (folder_name, info) 或 (None,None)。"""
    bv = norm(vessel)
    bc = norm(code)
    cands = {bv}
    if bv in NAME_ALIASES:
        cands.add(norm(NAME_ALIASES[bv]))
    # 1) 船名（含别名）
    for fol, info in index.items():
        if norm(fol) in cands:
            return fol, info
    # 2) 船代码
    for fol, info in index.items():
        if bc and norm(info["code"]) == bc:
            return fol, info
    return None, None


def is_block_header(ws, r):
    """判断 r 是否为船头行：下一行 C1 == 'PORT'，且本行 C4 是船名（非段标题）。"""
    if r >= ws.max_row:
        return False
    if norm(ws.cell(r + 1, 1).value) != "PORT":
        return False
    c4 = ws.cell(r, 4).value
    if not c4 or "VESSEL" in norm(c4):
        return False
    return True


def main():
    ap = argparse.ArgumentParser(description="自动刷新 CUL DAILY MOVEMENT.xlsx")
    ap.add_argument("--src", default=DEFAULT_SRC, help="2026 源文件夹根目录")
    ap.add_argument("--template", default=DEFAULT_TEMPLATE, help="现有大Excel模板路径")
    ap.add_argument("--output", default=None, help="输出路径，默认覆盖 template（先自动备份）")
    ap.add_argument("--no-backup", action="store_true", help="覆盖原件时不备份")
    args = ap.parse_args()

    src_dir = args.src
    template = args.template
    output = args.output or template

    if not os.path.isdir(src_dir):
        print(f"[ERROR] 源目录不存在: {src_dir}")
        return 1
    if not os.path.isfile(template):
        print(f"[ERROR] 模板文件不存在: {template}")
        return 1

    print("=== 1/3 构建源索引（每文件夹最新xlsx·第一个sheet）===")
    index = build_folder_index(src_dir)
    print(f"  源文件夹数: {len(index)}")

    print("=== 2/3 加载模板并刷新数据（保留格式/结构）===")
    # 双加载：
    #   wb_ro  (data_only)  —— 用于"检测"块边界/读船名，因为列头/船名可能是公式，
    #                           data_only 才能拿到缓存出的 "PORT" / 船名 字面量
    #   wb     (无data_only) —— 用于"写入"，保留原文件公式/样式/列宽/合并
    # 两个版本行号一致，按坐标 (row,col) 对应写入
    wb_ro = openpyxl.load_workbook(template, data_only=True)
    ws_ro = wb_ro[wb_ro.sheetnames[0]]
    wb = openpyxl.load_workbook(template)
    ws = wb[wb.sheetnames[0]]

    # 合并单元格锚点映射：写入合并格时改写到其左上角，避免 read-only 报错且保留合并
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        top, left = rng.min_row, rng.min_col
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merge_map[(row, col)] = (top, left)


    def set_val(r, c, val):
        tgt = merge_map.get((r, c), (r, c))
        ws.cell(tgt[0], tgt[1]).value = val

    blocks = 0
    updated = 0
    kept = 0
    unmatched_blocks = []
    total_rows = ws_ro.max_row
    r = 1
    while r <= total_rows:
        if is_block_header(ws_ro, r):
            vessel = str(ws_ro.cell(r, 4).value).strip()
            code = ws_ro.cell(r, 9).value
            blocks += 1
            fol, info = find_folder(vessel, code, index)
            if not fol:
                unmatched_blocks.append(vessel)
                r += 1
                continue
            lookup = info["lookup"]
            # 列头在 r+1；数据从 r+2 开始（坐标在 wb_ro 与 wb 中一致）
            rr = r + 2
            _bh = 0; _bm = 0
            while rr <= total_rows:
                sc1 = norm(ws_ro.cell(rr, 1).value)
                if sc1 == "PORT" or sc1 == "#VALUE!" or \
                   (ws_ro.cell(rr, 4).value and "VESSEL" in norm(ws_ro.cell(rr, 4).value)) or \
                   sc1.startswith("PIC") or \
                   is_block_header(ws_ro, rr):   # 遇到下一个船头行也停下，避免跳过块
                    break
                if sc1 == "" or sc1.startswith("REMARK"):
                    rr += 1
                    continue
                key = (sc1, norm_voy(ws_ro.cell(rr, 7).value))
                if key in lookup:
                    src = lookup[key]
                    for c in range(1, 17):
                        set_val(rr, c, src[c - 1])
                    updated += 1; _bh += 1
                else:
                    kept += 1; _bm += 1
                rr += 1
            print(f"  [DBG] {vessel:26s} -> {fol:22s} hit={_bh:3d} miss={_bm:3d} lookup={len(lookup)}")
            r = rr
            continue
        r += 1

    print("=== 3/3 保存 ===")
    if output == template and not args.no_backup:
        bak = template + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(template, bak)
        print(f"  已备份原文件 -> {bak}")
    wb.save(output)

    print("---- 摘要 ----")
    print(f"  处理船块数      : {blocks}")
    print(f"  刷新数据行      : {updated}")
    print(f"  保留原值(源无匹配): {kept}")
    print(f"  未找到源文件夹的块: {unmatched_blocks if unmatched_blocks else '无'}")
    print(f"  输出文件         : {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
