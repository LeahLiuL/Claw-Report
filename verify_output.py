#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证 generated 文件：结构完好 + 匹配行确实刷新 + 数据格类型未被破坏。"""
import os
import openpyxl
from datetime import datetime

TEMPLATE = r"P:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\CUL DAILY MOVEMENT.xlsx"
GEN = r"P:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\CUL DAILY MOVEMENT.generated.xlsx"

def norm(s): return (s or "").strip().upper()
def norm_voy(s):
    if s is None: return ""
    return str(s).strip().upper().replace(" ", "")

def block_rows(ws, vessel):
    """返回某船块内数据行号列表（坐标）。"""
    total = ws.max_row
    def is_hdr(ws, r):
        if r >= ws.max_row: return False
        if norm(ws.cell(r + 1, 1).value) != "PORT": return False
        c4 = ws.cell(r, 4).value
        if not c4 or "VESSEL" in norm(c4): return False
        return True
    r = 1
    while r <= total:
        if is_hdr(ws, r) and str(ws.cell(r, 4).value).strip() == vessel:
            rows = []
            rr = r + 2
            while rr <= total:
                sc1 = norm(ws.cell(rr, 1).value)
                if sc1 == "PORT" or sc1 == "#VALUE!" or (ws.cell(rr, 4).value and "VESSEL" in norm(ws.cell(rr, 4).value)) or sc1.startswith("PIC") or is_hdr(ws, rr):
                    break
                if sc1 == "" or sc1.startswith("REMARK"):
                    rr += 1; continue
                rows.append(rr); rr += 1
            return rows
        r += 1
    return []

def main():
    wb_t = openpyxl.load_workbook(TEMPLATE)
    wb_g = openpyxl.load_workbook(GEN)
    ws_t = wb_t[wb_t.sheetnames[0]]
    ws_g = wb_g[wb_g.sheetnames[0]]

    print("=== 结构对比 ===")
    print(f"  模板 sheet名 : {wb_t.sheetnames}")
    print(f"  生成 sheet名 : {wb_g.sheetnames}")
    print(f"  模板 维度    : {ws_t.dimensions} (max_row={ws_t.max_row}, max_col={ws_t.max_column})")
    print(f"  生成 维度    : {ws_g.dimensions} (max_row={ws_g.max_row}, max_col={ws_g.max_column})")
    print(f"  模板 合并区数: {len(ws_t.merged_cells.ranges)}")
    print(f"  生成 合并区数: {len(ws_g.merged_cells.ranges)}")
    print(f"  模板 块数     : {sum(1 for r in range(1, ws_t.max_row+1) if _hdr(ws_t, r))}")
    print(f"  生成 块数     : {sum(1 for r in range(1, ws_g.max_row+1) if _hdr(ws_g, r))}")

    print("\n=== CUL HAIPHONG 匹配行刷新校验 (模板 vs 生成) ===")
    rows = block_rows(ws_t, "CUL HAIPHONG")
    print(f"  块内数据行数: {len(rows)}")
    changed = 0; same = 0; formula_preserved = 0; formula_lost = 0
    for rr in rows:
        # 检查 C1..C16 是否变化 + 模板该格是否为公式
        diff = False
        for c in range(1, 17):
            tv = ws_t.cell(rr, c).value
            gv = ws_g.cell(rr, c).value
            if tv != gv:
                diff = True
            # 模板数据格是否公式
            if ws_t.cell(rr, c).data_type == 'f':
                if ws_g.cell(rr, c).data_type == 'f':
                    formula_preserved += 1
                else:
                    formula_lost += 1
        if diff: changed += 1
        else: same += 1
    print(f"  变化的行数: {changed}  未变化行数: {same}")
    print(f"  模板公式格保留: {formula_preserved}  模板公式格被覆盖: {formula_lost}")

    # 抽样打印前3行 C1/C7/C9 模板 vs 生成
    print("\n  抽样(行号 | 模板 PORT/VOY/ETA | 生成 PORT/VOY/ETA):")
    for rr in rows[:3]:
        tv = [ws_t.cell(rr, c).value for c in (1, 7, 9)]
        gv = [ws_g.cell(rr, c).value for c in (1, 7, 9)]
        print(f"    r{rr}: T={tv}  G={gv}")

    # 全局：扫描生成文件是否出现新公式错误文本
    print("\n=== 公式错误扫描 ===")
    errs = 0
    for row in ws_g.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#") and cell.value.endswith("!"):
                errs += 1
    print(f"  生成文件中形如 #X! 的单元格数: {errs}")

def _hdr(ws, r):
    if r >= ws.max_row: return False
    if norm(ws.cell(r + 1, 1).value) != "PORT": return False
    c4 = ws.cell(r, 4).value
    if not c4 or "VESSEL" in norm(c4): return False
    return True

if __name__ == "__main__":
    main()
