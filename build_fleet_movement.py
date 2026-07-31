#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_fleet_movement.py  —— 重建版 CUL DAILY MOVEMENT 大Excel
=====================================================================
按当前 2026/ 各船文件夹(=当前船队, 增删自动反映)重建大Excel。
【完全不依赖旧大Excel】——所有权威信息来自:
  - vessel.csv   (GitHub 仓库内, 船名<->代码<->显示名; 两机 git pull 同步, 用户手动维护)
  - P盘 PIC汇总.xlsx (人工维护的 PIC 对照表)

规则(用户2026-07-30确认):
  1. 港口行: 仅显示 ETB 在 今日±WINDOW天 窗口内的行 (默认±30)。
  2. 船名代码(code, 块头C9) / 显示名(C4): 从 vessel.csv 取(权威);
     仅当 vessel.csv 缺该船时, code 回退 源R1C9、显示名回退 文件夹名(并报警提示补登)。
  3. PIC(块头C16): 从 P盘 PIC汇总.xlsx 取(按文件夹名); 取不到则留空待补。
  4. 船块顺序: 按 ROUTE_ORDER 常量(固化20组, 可编辑) 分组; 组内按船名(文件夹名)排序;
     新航线追加到末尾(字母序)。
  5. Remark: 取源第一个sheet中 ETB 最接近今日 那行的 C18, 重建成 "Remark:航次 港口 原文";
     若该行为空(无remark)则【整行不写】。

重要事实(已核查):
  - 源 R1 C4(船名) 经常为空/错 -> 显示名以 vessel.csv 的 ship 列为准, 缺则文件夹名。
  - 源 R1 C1(航线码) 与规范航线码大量改名 -> 用 ROUTE_OVERRIDE / ROUTE_ALIAS 校正。
  - 源 R1 C1 可能为空(ZBM) -> 回退 ROUTE_FALLBACK。
  - 源 R1 C9(船代码) 布局不统一(ASR代码在C12) -> 以 vessel.csv 为准。


用法:
  python build_fleet_movement.py --src "P:\\...\\2026" --output "CUL DAILY MOVEMENT.rebuilt.xlsx"
  (culadmin 那台默认 Z: 盘, 直接 python build_fleet_movement.py)
"""
import argparse, os, glob, shutil, csv
from datetime import datetime, date, timedelta
import openpyxl

# 数据更新目录: 本机(leahliu)=P:, 另一台(culadmin)=Z:。自动探测存在的盘符, 两机通用, 无需传参。
_BASES = [
    r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新",
    r"P:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新",
]
UPD_DIR = next((b for b in _BASES if os.path.isdir(b)), _BASES[0])
# 脚本生成的输出统一放这个子文件夹(本机P:/另一台Z: 自动切换), 与源数据 2026/ 分开。
GEN_SUBDIR = "生成结果"
GEN_DIR = os.path.join(UPD_DIR, GEN_SUBDIR)
DEFAULT_SRC = os.path.join(UPD_DIR, "2026")
DEFAULT_OUT = os.path.join(GEN_DIR, "CUL DAILY MOVEMENT.rebuilt.xlsx")
# 船名<->代码<->显示名 权威表(GitHub 仓库内, 两机 git pull 同步); 用户在此手动维护。
VESSEL_CSV = os.path.join(os.path.dirname(os.path.abspath(__file__)), "vessel.csv")
# PIC 权威表: 生成结果子文件夹内的 PIC汇总.xlsx (人工编辑主文件, 仅 xlsx)。两机通用。
DEFAULT_PIC = os.path.join(GEN_DIR, "PIC汇总.xlsx")

# 航线分组顺序(固化常量, 与当前网页/大Excel展示一致; 新增航线追加到末尾)。
# 组内按船名(文件夹名)排序。如要调整顺序, 改这里即可。
ROUTE_ORDER = ["ST3","CHT","HDT","CST","CCT","NP2","REX","CGS","AEM","EVHA",
               "SGX","SHTG","RTS","NSX","SL1","CGX","HLX","ZGCD","IMR","NAX"]
ROUTE_FALLBACK = "RTS"   # 源航线码为空时回退
WINDOW_DAYS = 30

# 航线修正(用户2026-07-30确认): 文件夹 -> 规范航线码(覆盖源R1C1的改名/错误)
ROUTE_OVERRIDE = {"CUL NANSHA": "CCT"}     # 源R1C1误为HDT, 实为CCT
# 航线合并: 源航线码 -> 规范航线码(同一条航线在源里有不同叫法)
ROUTE_ALIAS = {"AM1": "AEM"}                # AEM 与 AM1 是同一航线

# 大Excel列头(沿用旧文件标签, 与源C1..C16位置一一对应)
COL_HEADERS = ["PORT","man in","wait","Proforma","ltm eta","ltm etd","VOY. NO",
               "date","ETA","ETB","ETD","run","Port Stay(hr)","fsp distance",
               "speed","ETA Delay/Ahead"]

SEGMENT_TITLE = "CUL VESSEL DAILY MOVEMENT  "

def norm(s): return (s or "").strip().upper().replace(" ", "")
def norm_voy(s):
    if s is None: return ""
    return str(s).strip().upper().replace(" ", "")

# ── 表头对齐: 大Excel目标列 -> 源文件表头候选(归一化名) ──
# 源文件列顺序不统一(如ASR多一列TERMINAL把VOY.NO推到C8), 故按"表头名"映射而非固定列位。
def norm_h(s):
    return (s or "").strip().upper().replace(" ", "")

def excel_serial_to_dt(v):
    try:
        return datetime(1899, 12, 30) + timedelta(days=float(v))
    except Exception:
        return None

def norm_date_value(v):
    """真实日期 -> datetime(统一); 文本标记(OMIT/Mon/...) -> 原字符串; 空 -> None。"""
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, (int, float)):
        d = excel_serial_to_dt(v)
        return d if d else None
    s = str(v).strip()
    if s == "": return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%m/%d %H:%M", "%m/%d", "%d/%m %H:%M", "%d/%m"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return s

TARGET_HEADERS = {
    1:  ["PORT"],
    2:  ["MANIN", "MAN IN"],
    3:  ["WAIT"],
    4:  ["PROFORMA"],
    5:  ["LTS ETB", "LTM ETB"],          # 源LTS/LTM ETB -> 大Excel C5(ltm eta位置)
    6:  ["LTS ETD", "LTM ETD"],
    7:  ["VOY.NO", "VOYNO.", "VOYNO", "VOY. NO"],
    8:  ["DATE"],
    9:  ["ETA"],
    10: ["ETB"],
    11: ["ETD"],
    12: ["RUN"],
    13: ["PORTSTAY(HR)", "PORTSTAY"],
    14: ["FSPDISTANCE", "FSP DISTANCE"],
    15: ["SPEED"],
    16: ["ETADELAY/AHEAD", "ETA DELAY/AHEAD"],
}

def nearest_voy_upward(rows, idx):
    """ETB最近行航次号为空时, 向上(行号更小=表中更靠上)找最近的、有航次号的行。"""
    for j in range(idx - 1, -1, -1):
        if rows[j]["voy"]:
            return rows[j]["voy"]
    return ""

def latest_xlsx(folder):
    files = [f for f in glob.glob(os.path.join(folder, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not files: return None
    files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return files[0]

def read_source(path):
    """读源第一个sheet。按表头名(非固定列位)映射到大Excel列, 兼容源列序差异(如ASR多TERMINAL列)。
    返回 dict: route, code, rows[ {display:{col:val}, etb:datetime|None, voy, port, remark} ]。
    - 日期列(C5/C6/C8/C9/C10/C11)统一为 datetime(真实日期) 或 文本标记(OMIT/Mon..)。
    - Voy.No 若空, 向上就近取最近的有航次号的行。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    route = ws.cell(1, 1).value
    code = ws.cell(1, 9).value
    hr = None
    for r in range(1, min(ws.max_row, 200) + 1):
        if norm(ws.cell(r, 1).value) == "PORT":
            hr = r; break
    if hr is None:
        return {"route": route, "code": code, "rows": []}
    # 源表头归一名 -> 源列号
    src_hdr = {}
    for c in range(1, ws.max_column + 1):
        h = norm_h(ws.cell(hr, c).value)
        if h and h not in src_hdr:
            src_hdr[h] = c
    # 目标列 -> 源列号
    col_map = {}
    for tcol, cands in TARGET_HEADERS.items():
        for cand in cands:
            sc = src_hdr.get(norm_h(cand))
            if sc:
                col_map[tcol] = sc; break
    raw = []
    current_route = route   # 初始航线=R1C1, 遇到中间段标题行会切换
    for r in range(hr + 1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None: continue
        s1 = norm(c1)
        if s1 == "PORT": continue          # 多段航次表, 跳过重复列头继续读
        # ── 检测中间段标题行(如 R33: C1="NP2", C4="ZHI YING HE SHUN", C9="ZYHS") ──
        # 特征: C4 有文本值(船名) 且 C1 不是港口名(非纯大写英文港口码或含空格/中文)
        c4_val = ws.cell(r, 4).value
        if c4_val and isinstance(c4_val, str) and c4_val.strip():
            # 这是一个段标题行(换航线了), 更新后续行的航线
            current_route = str(c1).strip()
            continue    # 段标题行本身不作为数据行
        display = {}
        for tcol in range(1, 17):
            sc = col_map.get(tcol)
            display[tcol] = ws.cell(r, sc).value if sc else None
        etb = display.get(10)
        etb = etb if isinstance(etb, datetime) else None
        raw.append({
            "display": display,
            "etb": etb,
            "voy_raw": norm_voy(display.get(7)),
            "port": s1,
            "remark": ws.cell(r, 18).value,
            "row_route": current_route,      # ← 该行所属的实际航线(支持一船多段)
        })
    # Voy.No 向上就近: 每行若空, 向上(表中更靠上)找最近的有航次号的行
    for i, rr in enumerate(raw):
        if rr["voy_raw"]:
            continue
        for j in range(i - 1, -1, -1):
            if raw[j]["voy_raw"]:
                rr["voy_raw"] = raw[j]["voy_raw"]; break
    # 统一日期格式 + 写入display
    for rr in raw:
        for tcol in (5, 6, 8, 9, 10, 11):
            rr["display"][tcol] = norm_date_value(rr["display"].get(tcol))
        rr["display"][7] = rr["voy_raw"]
        rr["voy"] = rr["voy_raw"]
    return {"route": route, "code": code, "rows": raw}

def load_vessel_csv(path):
    """GitHub 仓库内 vessel.csv: 船名(或文件夹名)-> {code, display}。用户手动维护。
    显示名取 ship 列原始值; 代码取 code 列。两机 git pull 同步。"""
    d = {}
    if not os.path.exists(path):
        return d
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            ship = (row.get("ship") or "").strip()
            code = (row.get("code") or "").strip()
            if ship:
                d[norm(ship)] = {"code": code, "display": ship}
    return d

def load_pic(path):
    """PIC 权威表: 优先 xlsx(人工编辑主文件), 回退 csv(镜像)。按文件夹名取PIC。
    取不到返回空dict(回退旧大Excel)。"""
    if str(path).lower().endswith(".xlsx"):
        return _load_pic_xlsx(path)
    return _load_pic_csv(path)

def _load_pic_xlsx(path):
    d = {}
    if not os.path.exists(path):
        return d
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return d
    ws = wb[wb.sheetnames[0]]
    # 表头: 航线|船名(显示)|文件夹名|船代码|PIC|状态 -> C3=文件夹名, C5=PIC
    for r in range(2, ws.max_row + 1):
        fol = ws.cell(r, 3).value
        pic = ws.cell(r, 5).value
        if fol is not None and str(fol).strip():
            d[norm(str(fol).strip())] = str(pic or "").strip()
    return d

def _load_pic_csv(path):
    d = {}
    if not os.path.exists(path):
        return d
    with open(path, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            fol = (row.get("文件夹名") or row.get("folder") or "").strip()
            pic = (row.get("PIC") or row.get("pic") or "").strip()
            if fol:
                d[norm(fol)] = pic
    return d

def canon_route(folder, src_route):
    """算规范航线码: 先应用文件夹级覆盖, 再做同航线合并, 空则回退。"""
    if folder in ROUTE_OVERRIDE:
        r = ROUTE_OVERRIDE[folder]
    else:
        r = src_route
    r = ROUTE_ALIAS.get(norm(r), r)
    if norm(r) in ("", None):
        r = ROUTE_FALLBACK
    return r

def build_route_order(groups):
    """分组顺序: 跟随 ROUTE_ORDER 常量(固化, 可编辑); 常量里没有的新航线追加到末尾(按字母序)。"""
    seen = [r for r in ROUTE_ORDER if norm(r) in groups]
    extra = sorted([r for r in groups if norm(r) not in [norm(x) for x in ROUTE_ORDER]])
    return seen + extra

def main():
    ap = argparse.ArgumentParser(description="按当前船队重建 CUL DAILY MOVEMENT 大Excel")
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--vessel", default=VESSEL_CSV, help="GitHub vessel.csv (船名->代码/显示名)")
    ap.add_argument("--pic", default=DEFAULT_PIC, help="PIC 汇总.xlsx (人工编辑主文件, 文件夹名->PIC)")
    ap.add_argument("--output", default=DEFAULT_OUT)
    ap.add_argument("--today", default=None, help="基准日 YYYY-MM-DD, 默认今天")
    ap.add_argument("--window", type=int, default=WINDOW_DAYS)
    ap.add_argument("--no-backup", action="store_true")
    args = ap.parse_args()

    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()
    lo = today - timedelta(days=args.window)
    hi = today + timedelta(days=args.window)

    print(f"[基准日] {today}  窗口 ±{args.window}天 -> [{lo} ~ {hi}]")
    print("=== 1/4 读权威表(vessel.csv / P盘PIC) ===")
    vessel = load_vessel_csv(args.vessel)
    pic_tbl = load_pic(args.pic)
    print(f"  vessel.csv: {len(vessel)} 条 | P盘PIC表: {len(pic_tbl)} 条")

    print("=== 2/4 扫描当前船队(2026/文件夹) ===")
    folders = sorted([d for d in os.listdir(args.src)
                      if os.path.isdir(os.path.join(args.src, d)) and d != "已下线船舶"])
    ships = []   # {folder, route, code, display, rows}
    for fol in folders:
        p = latest_xlsx(os.path.join(args.src, fol))
        if not p:
            print(f"  [WARN] 无xlsx跳过: {fol}"); continue
        d = read_source(p)
        route = canon_route(fol, d["route"])   # 应用覆盖+合并
        key = norm(fol)
        # 船名代码 & 显示名: vessel.csv(权威) -> 源R1C9 / 文件夹名
        vent = vessel.get(key)
        if vent:
            code = vent.get("code") or d["code"]
            disp = vent.get("display") or fol
        else:
            code = d["code"]
            disp = fol
            print(f"  [WARN] 船未在 vessel.csv 登记: {fol} (code/PIC 回退 源R1C9/文件夹名, 建议补登)")
        # ── 按逐行航线拆分子块(支持一船多段, 如 ZYHS SGX→NP2) ──
        sub_routes = {}
        for rr in d["rows"]:
            sr = canon_route(fol, rr.get("row_route", route))   # 每行实际航线(应用覆盖+合并)
            sub_routes.setdefault(sr, []).append(rr)
        if len(sub_routes) > 1:
            print(f"  [{fol}] 拆为 {len(sub_routes)} 个航线段: {', '.join(sub_routes.keys())}")
        for sub_r, sub_rows in sub_routes.items():
            ships.append({"folder": fol, "route": sub_r, "code": code, "display": disp, "rows": sub_rows})
    print(f"  当前船文件夹数: {len(ships)}")

    print("=== 3/4 分组排序 + 写表 ===")
    # 分组
    groups = {}
    for s in ships:
        groups.setdefault(norm(s["route"]), []).append(s)
    # 顺序: 跟随 ROUTE_ORDER 常量(固化), 新航线追加末尾
    ordered_routes = build_route_order(groups)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CUL DAILY MOVEMENT"
    # 列宽
    widths = [10,8,7,10,11,11,10,10,16,16,16,7,11,11,8,14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1
    seg_fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    hdr_fill = openpyxl.styles.PatternFill("solid", fgColor="D9E1F2")
    bold = openpyxl.styles.Font(bold=True)
    thin = openpyxl.styles.Side(style="thin", color="BFBFBF")
    border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)

    def setc(r, c, v, font=None, fill=None, bd=False, merge_to=None):
        cell = ws.cell(r, c, v)
        if font: cell.font = font
        if fill: cell.fill = fill
        if bd: cell.border = border
        if merge_to:
            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
        return cell

    blocks_written = 0
    for route in ordered_routes:
        grp = sorted(groups[norm(route)], key=lambda s: norm(s["folder"]))
        # 段标题(C1='#VALUE!' 与 C4=标题 为两个独立单元格, 不合并)
        setc(row, 1, "#VALUE!", font=bold, fill=seg_fill)
        setc(row, 4, SEGMENT_TITLE, font=bold, fill=seg_fill)
        row += 1
        for s in grp:
            fol = s["folder"]
            disp = s["display"]
            # PIC: P盘PIC汇总.xlsx(权威); 取不到则留空待补
            pic_raw = pic_tbl.get(norm(fol))
            pic_clean = str(pic_raw).replace("PIC:", "").replace("PIC :", "").strip() if pic_raw else ""
            # 块头
            setc(row, 1, route, font=bold)
            setc(row, 4, disp, font=bold)
            setc(row, 8, "DATE")
            setc(row, 9, s["code"])
            setc(row, 16, "PIC: " + pic_clean)   # 始终带PIC:前缀, 保证网页解析识别块(含无PIC新船)
            row += 1
            # 列头
            for c, h in enumerate(COL_HEADERS, 1):
                setc(row, c, h, font=bold, fill=hdr_fill, bd=True)
            row += 1
            # 港口行(窗口内)
            shown = 0
            for rr in s["rows"]:
                if rr["etb"] is None:
                    continue
                d = rr["etb"].date()
                if not (lo <= d <= hi):
                    continue
                for c in range(1, 17):
                    setc(row, c, rr["display"].get(c), bd=True)
                row += 1; shown += 1
            # Remark: ETB 最接近今日 那行; 航次号为空则向上找最近的有航次号的行
            best_idx = None; best_diff = None
            for i, rr in enumerate(s["rows"]):
                if rr["etb"] is None: continue
                diff = abs((rr["etb"].date() - today).days)
                if best_diff is None or diff < best_diff:
                    best_diff = diff; best_idx = i
            if best_idx is not None:
                best = s["rows"][best_idx]
                voy = best["voy"] or nearest_voy_upward(s["rows"], best_idx)
                rem = best["remark"]
                if rem:   # 仅当有 remark 内容才写; 否则整行不写
                    txt = f"Remark:{voy} {best['port']} {rem}"
                    setc(row, 1, txt)
                    row += 1
            row += 1   # 块间空行
            blocks_written += 1
        row += 1   # 段间空行

    print("=== 4/4 保存 ===")
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    if os.path.exists(args.output) and not args.no_backup:
        bak = args.output + ".bak_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(args.output, bak)
        print(f"  已备份 -> {bak}")
    wb.save(args.output)
    print(f"  航线组数: {len(ordered_routes)}  写出船块数: {blocks_written}")
    print(f"  输出: {args.output}")

if __name__ == "__main__":
    raise SystemExit(main())
