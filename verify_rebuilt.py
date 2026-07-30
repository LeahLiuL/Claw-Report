#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""核查 rebuilt 文件: 段/块结构, 窗口内港口行数, remark 重建格式。"""
import openpyxl
F = r"P:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\CUL DAILY MOVEMENT.rebuilt.xlsx"
wb = openpyxl.load_workbook(F, data_only=True)
ws = wb[wb.sheetnames[0]]
def norm(s): return (s or "").strip().upper()
def is_block_header(ws, r):
    if r >= ws.max_row: return False
    if norm(ws.cell(r+1,1).value) != "PORT": return False
    c4 = ws.cell(r,4).value
    if not c4 or "VESSEL" in norm(c4): return False
    return True

print("===== 段标题(航线组)顺序 & 每组船块 =====")
r = 1
grp_idx = 0
while r <= ws.max_row:
    if ws.cell(r,1).value == "#VALUE!":
        grp_idx += 1
        # 找该段下所有块
        ships = []
        rr = r + 1
        while rr <= ws.max_row and ws.cell(rr,1).value != "#VALUE!":
            if is_block_header(ws, rr):
                ships.append((ws.cell(rr,1).value, ws.cell(rr,4).value, ws.cell(rr,16).value))
            rr += 1
        print(f"  [组{grp_idx}] 船数={len(ships)}: {[(s[0],s[1]) for s in ships]}")
        r = rr
        continue
    r += 1

print("\n===== 抽样检查: 每航线组第1个船块 的 港口行数 / remark =====")
r = 1
checked = 0
while r <= ws.max_row and checked < 6:
    if ws.cell(r,1).value == "#VALUE!":
        rr = r + 1
        while rr <= ws.max_row and ws.cell(rr,1).value != "#VALUE!":
            if is_block_header(ws, rr):
                route = ws.cell(rr,1).value; ship = ws.cell(rr,4).value; pic = ws.cell(rr,16).value
                # 数港口行 + 找remark
                port_n = 0; remark = None
                x = rr + 2
                while x <= ws.max_row:
                    c1 = norm(ws.cell(x,1).value)
                    if c1 == "PORT" or c1 == "#VALUE!" or (ws.cell(x,4).value and "VESSEL" in norm(ws.cell(x,4).value)) or c1.startswith("PIC") or is_block_header(ws, x):
                        break
                    if c1 == "" or c1.startswith("REMARK"):
                        if c1.startswith("REMARK"): remark = ws.cell(x,1).value
                        x += 1; continue
                    port_n += 1; x += 1
                print(f"  {route:6s} | {str(ship):20s} | PIC={pic!r} | 港口行={port_n} | remark={remark!r}")
                checked += 1
                break
            rr += 1
        r = rr
        continue
    r += 1

print("\n===== 全部 remark 行(确认格式 Remark:航次 港口 原文) =====")
r = 1
n_rem = 0
while r <= ws.max_row:
    v = ws.cell(r,1).value
    if isinstance(v,str) and v.startswith("Remark:"):
        n_rem += 1
        if n_rem <= 12:
            print(f"  {v}")
    r += 1
print(f"  ...共 {n_rem} 条 remark")
