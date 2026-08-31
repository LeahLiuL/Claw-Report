# -*- coding: utf-8 -*-
"""
ROB 盘油记录自动刷新（Claw-Report 仓库）

流水线:
  1. 船清单: 解析本仓库 cul_daily_movement.html 的 TODAY_DATA (失败则回退 rob_data/dm_summary.json)
  2. ROB:    Outlook CULINES store 实时抓各船最新 Noon/Berth/Sailing Report
             策略: ① Vessel/<船名> 子文件夹 ② 已知 sender 邮箱 ③ 收件箱主题含船名
             抓不到的船保留上次数据(不丢数据)
  3. 输出:   rob_data/rob_results.json (持久化, 记录每船 sender 便于下次定位)
             rob_oil_report.html (AES 加密网页, 密码见 PASSWORD)
             rob_data/rob_oil_table.xlsx + 本机 C:\\CULINES 同步(仅当该目录存在)

用法:
  python rob_refresh.py               # 完整刷新
  python rob_refresh.py --no-outlook  # 只用已有数据重新生成网页(调试)
  python rob_refresh.py --vessel "CHANG SHENG JI 8"  # 只刷新指定船
"""
import sys, os, re, json, base64, hashlib, argparse, tempfile, csv
sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime, timedelta

BASE = os.path.dirname(os.path.abspath(__file__))
ROB_DIR = os.path.join(BASE, "rob_data")
DM_HTML = os.path.join(BASE, "cul_daily_movement.html")
RESULTS = os.path.join(ROB_DIR, "rob_results.json")
DM_FALLBACK = os.path.join(ROB_DIR, "dm_summary.json")
SENDER_MAP_FILE = os.path.join(ROB_DIR, "vessel_senders.json")
OUT_HTML = os.path.join(BASE, "rob_oil_report.html")
OUT_XLSX = os.path.join(ROB_DIR, "rob_oil_table.xlsx")
CULINES_DIR = r"C:\CULINES\Claw Report"
HISTORY_DIR = os.path.join(ROB_DIR, "history")
# 累加历史 CSV: 单一权威文件, 顶层 rob_data/rob_history.csv (双机协作版规范, 方案A)
# 14 列, 与另一台每日更新 ROB 的机器完全一致, 避免列错位/互相覆盖。
# 注意: sender 不进 CSV(只保留在 rob_results.json), 否则两机 schema 不一致会写坏累计数据。
HISTORY_CSV = os.path.join(ROB_DIR, "rob_history.csv")
# 累加历史 CSV 的列(逐船逐次快照, 一行一船)
SNAP_FIELDS = ["date", "vessel", "code", "lane", "pic",
               "lsfo", "hsfo", "mgo", "ulsfo", "bw", "fw", "refeer",
               "found", "report_time"]

PASSWORD = "jimmy"          # 网页密码(AES, 源码看不到明文; 注意本仓库公开, 密码也在脚本里)
REPORT_KEYS = ("NOON", "BERTH", "SAILING", "ANCHOR", "DRIFT")
OIL_KEYS = ("LSFO", "HSFO", "MGO", "ULSFO")


def norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


# ---------------------------------------------------------------- 1. 船清单
# 已下线船舶: 船名带『已下线』后缀(船期表 cul_daily_movement.html 里就是这么标的)。
# 这些船已退租/停航, 船长不再发 ROB 报告, 留着只会占主表、拉低 Data Integrity 统计。
OFFLINE_MARK = "已下线"


def is_offline(name):
    return OFFLINE_MARK in (name or "")


def drop_offline(fleet):
    keep, dropped = [], []
    for v in fleet:
        if is_offline(v.get("vessel")):
            dropped.append(v["vessel"])
        else:
            keep.append(v)
    if dropped:
        print("Fleet: drop %d offline vessels -> %s"
              % (len(dropped), ", ".join(dropped)))
    return keep


def load_fleet():
    if os.path.exists(DM_HTML):
        try:
            html = open(DM_HTML, encoding="utf-8").read()
            m = re.search(r"const\s+TODAY_DATA\s*=\s*", html)
            if m:
                data, _ = json.JSONDecoder().raw_decode(html[m.end():])
                vs = data.get("vessels", [])
                fleet = [{"vessel": (v.get("vessel") or "").strip(),
                          "code": (v.get("code") or "").strip(),
                          "lane": (v.get("route") or "").strip(),
                          "pic": (v.get("pic") or "").strip()} for v in vs]
                if fleet:
                    print("Fleet: %d vessels from cul_daily_movement.html (date %s)"
                          % (len(fleet), data.get("date")))
                    return drop_offline(fleet)
        except Exception as e:
            print("[WARN] parse cul_daily_movement.html failed:", e)
    if os.path.exists(DM_FALLBACK):
        fleet = json.load(open(DM_FALLBACK, encoding="utf-8"))
        print("Fleet: %d vessels from fallback rob_data/dm_summary.json" % len(fleet))
        return drop_offline(fleet)
    raise SystemExit("No fleet source: neither cul_daily_movement.html nor rob_data/dm_summary.json")


# ---------------------------------------------------------------- 2. Outlook ROB
def connect_outlook():
    """返回主邮箱 store。注意: NS.Stores 里『联机存档 - leahliu@culines.com』也含
    CULINES.COM 且可能排在前面, 其收件箱是空的 —— 必须排除存档, 否则什么都抓不到。"""
    import win32com.client
    OL = win32com.client.Dispatch("Outlook.Application")
    NS = OL.GetNamespace("MAPI")
    stores = [s for s in NS.Stores if "CULINES.COM" in (s.DisplayName or "").upper()]
    main = [s for s in stores
            if "存档" not in (s.DisplayName or "")
            and "ARCHIVE" not in (s.DisplayName or "").upper()]
    if main:
        return main[0]
    return stores[0] if stores else None


def get_sender(it):
    """尽量拿到真实 SMTP 发件地址。Exchange 账号下 SenderEmailAddress 常返回
    X.500/空, 必须再走 Sender.GetExchangeUser().PrimarySmtpAddress 兜一层。"""
    for attr in ("SenderEmailAddress",):
        try:
            v = getattr(it, attr, "")
            if v and "@" in v:
                return v
        except Exception:
            pass
    try:
        ex = it.Sender
        if ex:
            smtp = ex.GetExchangeUser().PrimarySmtpAddress
            if smtp and "@" in smtp:
                return smtp
    except Exception:
        pass
    return ""


def load_sender_map():
    """船名(norm) -> 船长邮箱, 固化随仓库走。换电脑不依赖运行时历史。"""
    m = {}
    if os.path.exists(SENDER_MAP_FILE):
        try:
            raw = json.load(open(SENDER_MAP_FILE, encoding="utf-8"))
            for k, v in raw.items():
                if v:
                    m[norm(k)] = v
        except Exception as e:
            print("[WARN] load vessel_senders.json failed:", e)
    return m


def save_sender_map(m):
    try:
        with open(SENDER_MAP_FILE, "w", encoding="utf-8") as f:
            json.dump(m, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("[WARN] save vessel_senders.json failed:", e)


def build_folder_cache(inbox):
    """Vessel 子文件夹: norm(文件夹名) -> folder(递归多层)。
    注: dict 按名字做 key, 多艘船同名子文件夹(如 'Master')会互相覆盖,
    因此全量扫描一律改用 build_folder_list()。"""
    cache = {}
    try:
        vf = inbox.Folders["Vessel"]

        def walk(f):
            for c in f.Folders:
                cache.setdefault(norm(c.Name), c)
                try:
                    walk(c)
                except Exception:
                    pass
        walk(vf)
    except Exception:
        pass
    return cache


def build_folder_list(inbox):
    """收件箱下全部子文件夹对象列表(递归多层, 同名不去重)。
    必须覆盖收件箱顶层而不仅是 Vessel/: 相当多船的报告文件夹(SHTG、ZLST、cusk 等)
    直接位于收件箱顶层, 与 Vessel/ 平级; 且多艘船都有 'Master' 同名子文件夹,
    用 dict 存会互相覆盖导致漏扫。这里保留每一个文件夹对象。"""
    out = []

    def walk(f, depth=0):
        if depth > 4:
            return
        try:
            for c in f.Folders:
                out.append(c)
                walk(c, depth + 1)
        except Exception:
            pass
    walk(inbox)
    return out


def build_sender_index(folders, sender_map, per_folder=300, cap_per_sender=15):
    """按用户提供的船长邮箱, 在全部文件夹里预建索引: sender -> 最近若干封邮件(新在前)。
    用 get_sender() 解析真实 SMTP(CULINES 内部 Exchange 存的是 X.500, 直接按 SMTP
    字符串 Restrict 永远匹配不上)。报告无论落在哪个文件夹都能被 sender 精确定位。"""
    targets = set(v.lower() for v in sender_map.values() if v)
    if not targets:
        return {}
    idx = {}
    for fobj in folders:
        try:
            items = fobj.Items
            items.Sort("[ReceivedTime]", True)
        except Exception:
            continue
        cnt = 0
        for it in items:
            if cnt >= per_folder:
                break
            cnt += 1
            sa = get_sender(it)
            if not sa:
                continue
            key = sa.lower()
            if key not in targets:
                continue
            lst = idx.get(key)
            if lst is None:
                lst = []
                idx[key] = lst
            if len(lst) < cap_per_sender:
                lst.append(it)
    return idx


def match_folder(cache, vessel, code=None):
    """返回 (folder, exact)。exact=False 表示前缀匹配的共用/变体文件夹(如 'MEDKON'
    同时放 MEDKON DON / MEDKON LIA 的邮件), 需按主题过滤防误抓。
    code: 船代码。很多船的报告文件夹按代码命名(SHTG / ZLST / cusk), 只按船名匹配不到。"""
    nv = norm(vessel)
    nc = norm(code or "")
    if not nv and not nc:
        return None, False
    # 精确: 船名或代码直接命中文件夹名
    if nv and nv in cache:
        return cache[nv], True
    if nc and nc in cache:
        return cache[nc], True
    # 前缀: 共用/变体文件夹, 需主题过滤
    for k, f in cache.items():
        if len(k) < 6:
            continue
        if nv and (k.startswith(nv) or nv.startswith(k)):
            return f, False
        if nc and (k.startswith(nc) or nc.startswith(k)):
            return f, False
    return None, False


def pick_report_attachment(it, strict=False):
    xlsx = None
    for a in it.Attachments:
        fn = a.FileName
        if fn.lower().endswith(".xlsx") and any(k in fn.upper() for k in REPORT_KEYS):
            return a
    if strict:
        return None
    for a in it.Attachments:
        if a.FileName.lower().endswith(".xlsx"):
            return a
    return None


def extract_rob(att):
    import openpyxl
    fd, p = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        att.SaveAsFile(p)
        wb = openpyxl.load_workbook(p, data_only=True)
    except Exception:
        try:
            os.remove(p)
        except Exception:
            pass
        return {}
    # CUL NOON/BERTH/SAILING 报告的油表行有多种写法, 只认行首 "ROB " 会漏抓:
    #   NOON  : "ROB LSFO | 598.17 | MT | ROB MGO | 372.84"   (一行多个键值对)
    #   BERTH : "POB/FWE/ANCHOR AWEIGH ROB LSFO | 602.35"     (阶段前缀)
    # 规则: 单元格含 "ROB" 且 ROB 之后的 token 是已知油种代码时, 取该单元格右侧第一个
    # 数值。同一油种重复出现时后值覆盖前值(BERTH 按 POB/FWE/ANCHOR AWEIGH 排列,
    # 最晚阶段即最新存量)。
    oil_codes = ("LSFO", "HSFO", "ULSFO", "MGO", "BW", "FW", "REFEER", "REEFER")

    import re as _re

    def _as_num(v):
        """单元格 -> float, 非数字返回 None。支持 '1,234.5' / ' 598.17 ' / 数字类型。"""
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        if not s:
            return None
        if not _re.match(r"^-?\d+(\.\d+)?$", s):
            return None
        return float(s)

    def _take(cu, i, row):
        if "ROB" not in cu:
            return None
        pos = cu.rfind("ROB")
        tail = cu[pos + 3:].strip()
        if not tail:
            return None
        tok = tail.split()[0].strip(" :|-")
        if tok not in oil_codes:
            return None
        # 向右找数值, 但遇到"有内容的文本"(单位 MT / 下一个键名 'POB ROB HSFO')
        # 必须停止 —— BERTH 报告一行放多组 "键 | 值 | MT", 若跨过 MT 继续扫,
        # 空值会误抓到下一组的 0(例: 'POB ROB LSFO | (空) | MT | POB ROB HSFO | 0')
        # 导致整船 ROB 被清零(2026-08-30 CUL HAIPHONG / KAI DA HONG ZHOU 事故)。
        # 空单元格跳过(合并单元格/排版留白), 非空非数字立即终止。
        for j in range(i + 1, len(row)):
            v = row[j]
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            n = _as_num(v)
            if n is None:
                break
            return (tok, round(n, 3))   # 收敛浮点尾巴(3651.5230000000006 -> 3651.523)
        return None

    rob = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for i, c in enumerate(row):
                if c and isinstance(c, str):
                    got = _take(c.upper().strip(), i, row)
                    if got:
                        k, v = got
                        rob["REFEER" if k == "REEFER" else k] = v
    # 低硫船只报 ULSFO 时归一到 LSFO, 统一主表/趋势口径
    if "ULSFO" in rob and "LSFO" not in rob:
        rob["LSFO"] = rob["ULSFO"]
    try:
        os.remove(p)
    except Exception:
        pass
    return rob


def scan_for_rob(items, max_attach=40, max_walk=1200, subject_token=None):
    """倒序遍历邮件, 返回最新一封报告附件含 ROB 的 (rob, receivedTime, subject, sender)。
    subject_token: 若设置(norm 后的船名), 邮件主题必须含该 token 才候选 —— 用于共用文件夹
    防止误抓其他船的报告(如 MEDKON 文件夹里 MEDKON DON 的邮件)。"""
    tried = walked = 0
    for it in items:
        walked += 1
        if walked > max_walk:
            break
        try:
            if subject_token:
                try:
                    subj0 = it.Subject or ""
                except Exception:
                    subj0 = ""
                if subject_token not in norm(subj0):
                    continue
            att = pick_report_attachment(it)
            if att is None:
                continue
            if tried >= max_attach:
                break
            tried += 1
            rob = extract_rob(att)
            if any(k in rob for k in OIL_KEYS):
                se = get_sender(it)
                try:
                    subj = it.Subject or ""
                except Exception:
                    subj = ""
                return rob, it.ReceivedTime, subj, se
        except Exception:
            continue
    return None


def apply_hit(rec, hit):
    rob, recv, subj, se = hit
    rec.update({
        "rob_lsfo": rob.get("LSFO"), "rob_hsfo": rob.get("HSFO"), "rob_mgo": rob.get("MGO"),
        "rob_ulsfo": rob.get("ULSFO"), "rob_bw": rob.get("BW"), "rob_fw": rob.get("FW"),
        "rob_refeer": rob.get("REFEER"),
        "report_time": recv.strftime("%Y-%m-%d %H:%M:%S"),
        "source": subj, "sender": se, "found": True,
    })
    return True


def refresh_vessel(inbox, cache, rec, sender_map=None, sender_index=None, folder_list=None,
                   fleet_norms=None):
    """多来源合并取全局最新: ①船名/代码文件夹树 ②sender 索引 ③收件箱 sender Restrict
    ④主题 Restrict(全文件夹)。所有候选按 EntryID 去重、ReceivedTime 全局倒序后,
    从最新一封往下找第一份能解析出 ROB 的报告 —— 无论报告落在哪个文件夹、由哪个发件人
    发出, 永远取时间上最新的那份。
    (旧实现是"哪个分支先命中就用谁", 文件夹里的旧报告会压过 sender 找到的新报告。)"""
    vname = rec["vessel"]
    nv = norm(vname)
    nc = norm(rec.get("code") or "")
    sender_map = sender_map or {}
    # 有效发件人: 运行时历史 优先, 否则用固化映射(换电脑也能用)
    eff_sender = rec.get("sender") or sender_map.get(nv) or sender_map.get(vname)
    # 该发件人是否对应多艘船(共用邮箱, 如 MEDKON DON/LIA)——需主题过滤防误抓
    shared = sum(1 for v in sender_map.values() if v == eff_sender) if eff_sender else 0

    cands = {}  # entryID -> (receivedTime, item)

    def add_items(items, token=None, limit=400):
        n = 0
        try:
            for it in items:
                n += 1
                if n > limit:
                    break
                try:
                    eid = it.EntryID
                    rt = it.ReceivedTime
                except Exception:
                    continue
                if token is not None:
                    try:
                        subj0 = it.Subject or ""
                    except Exception:
                        subj0 = ""
                    ns = norm(subj0)
                    toks = token if isinstance(token, (list, tuple, set)) else (token,)
                    if not any(t and t in ns for t in toks):
                        continue
                old = cands.get(eid)
                if old is None or rt > old[0]:
                    cands[eid] = (rt, it)
        except Exception:
            pass

    # ① 船名/代码命名的文件夹树(Vessel/ 下 + 收件箱顶层如 SHTG/ZLST)
    folder, exact = match_folder(cache, vname, rec.get("code"))
    if folder is None and folder_list:
        nc2 = norm(rec.get("code") or "")
        for fo in folder_list:
            try:
                fn2 = norm(fo.Name)
            except Exception:
                continue
            if fn2 == nv or (nc2 and fn2 == nc2):
                folder, exact = fo, True
                break
    if folder is not None:
        scan_folders = [folder]
        try:
            stack = list(folder.Folders)
            while stack:
                sub = stack.pop(0)
                scan_folders.append(sub)
                try:
                    stack.extend(sub.Folders)
                except Exception:
                    pass
        except Exception:
            pass
        # 主题过滤判定 —— 三种情况都强制按主题认船, 否则会串船:
        #   a) 前缀匹配到的共用文件夹(不精确命中)
        #   b) 共用发件人(一个邮箱发多艘船)
        #   c) **共用文件夹**: 文件夹名还是别的船名/代码的前缀。
        #      例: MEDKON 文件夹同时放 MEDKON DON 与 MEDKON LIA 的报告, 两者都
        #      "精确"命中该文件夹, 谁也不过滤主题 -> 两船都取到文件夹里最新那份
        #      (2026-08-30 实测两船显示完全相同的 276.4/103.3, 实为 DON 的数据)。
        try:
            fnm = norm(folder.Name)
        except Exception:
            fnm = ""
        shared_folder = bool(fnm) and len(fnm) >= 5 and any(
            o and o != fnm and o != nv and o.startswith(fnm)
            for o in (fleet_norms or ()))
        need_subj = shared_folder or (not exact) or shared > 1
        tokens = [t for t in (nv, nc if len(nc) >= 4 else "") if t]
        token1 = tokens if need_subj else None
        for fo in scan_folders:
            try:
                fo.Items.Sort("[ReceivedTime]", True)
            except Exception:
                pass
            add_items(fo.Items, token=token1)

    # ②' sender 预建索引(收件箱全部子文件夹, 已解析 X.500); 共用邮箱跳过
    if eff_sender and sender_index and shared <= 1:
        add_items(sender_index.get(eff_sender.lower()) or [])

    # ② 收件箱 sender Restrict 兜底(外部 SMTP 发件人落在收件箱时有效)
    if eff_sender:
        try:
            items = inbox.Items.Restrict("[SenderEmailAddress]='%s'" % eff_sender)
            items.Sort("[ReceivedTime]", True)
            add_items(items, token=(nv if shared > 1 else None))
        except Exception:
            pass

    # ③ 主题含船名(收件箱 + 全部子文件夹, 含顶层船文件夹/嵌套/同名)
    try:
        token_sql = "@SQL=\"urn:schemas:httpmail:subject\" like '%s'" % vname
        for fo in [inbox] + (list(folder_list) if folder_list else list(cache.values())):
            try:
                items = fo.Items.Restrict(token_sql)
                items.Sort("[ReceivedTime]", True)
                add_items(items, limit=60)
            except Exception:
                continue
    except Exception:
        pass

    if not cands:
        return False
    ordered = sorted(cands.values(), key=lambda x: x[0], reverse=True)
    hit = scan_for_rob([it for _, it in ordered], max_walk=400)
    if hit:
        if eff_sender:
            rec["sender"] = eff_sender
        return apply_hit(rec, hit)
    return False


# ---- 偏旧兜底 ----
# 常规多来源检索每条分支都有扫描上限(每文件夹前 N 封 / sender 最近 15 封), 个别船的
# 报告如果落在很深的层级或被大量别的邮件挡住, 就会一直停在旧日期(例: M. ODYSSEY
# 卡在 08-26)。下面按 ReceivedTime 窗口对全部文件夹 Restrict 一次(不受"前 N 封"限制),
# 再按 sender / 主题精确匹配, 保证只要邮箱里有更新的报告就一定能捞出来。
STALE_HOURS = 26          # 报告时间超过这个小时数(或压根没抓到)才进深度扫描
DEEP_LOOKBACK_DAYS = 6    # 深度扫描的回看天数


def is_stale(rec, now=None):
    rt = rec.get("report_time")
    if not rt:
        return True
    try:
        t = datetime.strptime(rt[:19], "%Y-%m-%d %H:%M:%S")
    except Exception:
        return True
    return ((now or datetime.now()) - t).total_seconds() > STALE_HOURS * 3600


def deep_refresh_stale(inbox, folder_list, recs, sender_map, lookback_days=None):
    """对"报告偏旧/抓不到"的船做一次全文件夹深度扫描, 返回补抓成功的条数。"""
    if not recs:
        return 0
    lookback_days = lookback_days or DEEP_LOOKBACK_DAYS
    cutoff = (datetime.now() - timedelta(days=lookback_days)).strftime("%m/%d/%Y %H:%M %p")
    folders = [inbox] + list(folder_list or [])
    pool = []
    for fo in folders:
        try:
            items = fo.Items.Restrict("[ReceivedTime] >= '%s'" % cutoff)
        except Exception:
            continue
        try:
            items.Sort("[ReceivedTime]", True)
        except Exception:
            pass
        try:
            for it in items:
                try:
                    pool.append((it.ReceivedTime, it))
                except Exception:
                    continue
        except Exception:
            continue
    pool.sort(key=lambda x: x[0], reverse=True)
    print("deep scan: %d items (last %dd, %d folders), %d stale vessels"
          % (len(pool), lookback_days, len(folders), len(recs)))
    fixed = 0
    for rec in recs:
        vname = rec.get("vessel", "")
        nv, nc = norm(vname), norm(rec.get("code") or "")
        eff = rec.get("sender") or sender_map.get(nv) or sender_map.get(vname)
        shared = sum(1 for v in (sender_map or {}).values() if v == eff) if eff else 0
        hits = []
        for _, it in pool:
            try:
                subj = norm(it.Subject or "")
            except Exception:
                subj = ""
            ok = bool(nv and nv in subj) or bool(nc and len(nc) >= 4 and nc in subj)
            if not ok and eff:
                try:
                    se = (get_sender(it) or "").lower()
                except Exception:
                    se = ""
                ok = (se == eff.lower())
            if ok:
                hits.append(it)
        if not hits:
            continue
        hit = scan_for_rob(hits, max_walk=200,
                           subject_token=(nv if (shared > 1 or not eff) else None))
        if hit:
            old_t = rec.get("report_time")
            apply_hit(rec, hit)
            if eff:
                rec["sender"] = eff
            fixed += 1
            print("   [DEEP] %-24s %s -> %s  (LSFO=%s MGO=%s)"
                  % (vname[:24], (old_t or "MISS")[:16], rec["report_time"][:16],
                     rec.get("rob_lsfo"), rec.get("rob_mgo")))
    print("deep scan fixed: %d/%d" % (fixed, len(recs)))
    return fixed


# ---------------------------------------------------------------- 3. AES (CryptoJS 兼容)
def evp_bytes_to_key(password, salt, key_len=32, iv_len=16):
    d = b""
    prev = b""
    while len(d) < key_len + iv_len:
        prev = hashlib.md5(prev + password + salt).digest()
        d += prev
    return d[:key_len], d[key_len:key_len + iv_len]


def cryptojs_encrypt(plaintext, passphrase):
    from Crypto.Cipher import AES
    from Crypto.Random import get_random_bytes
    salt = get_random_bytes(8)
    key, iv = evp_bytes_to_key(passphrase.encode(), salt)
    cipher = AES.new(key, AES.MODE_CBC, iv)
    data = plaintext.encode("utf-8")
    pad = 16 - len(data) % 16
    data += bytes([pad]) * pad
    return base64.b64encode(b"Salted__" + salt + cipher.encrypt(data)).decode()


def cryptojs_decrypt(enc, passphrase):
    """Python 端验证解密(模拟 CryptoJS.AES.decrypt(passphrase 模式))"""
    from Crypto.Cipher import AES
    raw = base64.b64decode(enc)
    assert raw[:8] == b"Salted__", "not CryptoJS format"
    salt = raw[8:16]
    ct = raw[16:]
    key, iv = evp_bytes_to_key(passphrase.encode(), salt)
    pt = AES.new(key, AES.MODE_CBC, iv).decrypt(ct)
    pad = pt[-1]
    return pt[:-pad].decode("utf-8")


# ---------------------------------------------------------------- 4. HTML
# 前端模板独立文件: templates/rob.html
# 抓取逻辑(本文件)与页面(HTML/CSS/JS)彻底解耦 —— 另一台电脑改页面只需动
# templates/rob.html, 不会和抓邮件的 Python 逻辑互相覆盖; 本脚本只负责读模板并
# 把加密数据注入 __ENC__ 占位符。
TEMPLATE_FILE = os.path.join(BASE, "templates", "rob.html")


def load_template():
    if not os.path.exists(TEMPLATE_FILE):
        raise SystemExit("Missing template: %s (前端模板缺失, 请从仓库恢复)" % TEMPLATE_FILE)
    return open(TEMPLATE_FILE, encoding="utf-8").read()


def _num(x):
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def _hours_of(s):
    """'YYYY-MM-DD HH:MM' -> 当天小时数(浮点), 用于挑最接近锚点的报告。"""
    import re as _re
    m = _re.search(r"(\d{1,2}):(\d{2})", s or "")
    if not m:
        return 12.0
    return int(m.group(1)) + int(m.group(2)) / 60.0


def build_html(results):
    vessels = []
    ordered = sorted(results, key=lambda r: (r.get("lane", ""), r.get("code", "")))
    for i, r in enumerate(ordered, 1):
        vessels.append({
            "seq": i, "vessel": r.get("vessel", ""), "code": r.get("code", ""),
            "lane": r.get("lane", ""), "pic": r.get("pic", ""),
            "rob_lsfo": r.get("rob_lsfo"), "rob_hsfo": r.get("rob_hsfo"),
            "rob_ulsfo": r.get("rob_ulsfo"), "rob_mgo": r.get("rob_mgo"),
            "found": bool(r.get("found")),
            "remark": "No ROB report from Master found in mailbox" if not r.get("found") else "",
            "report_time": (r.get("report_time") or "")[:19],
        })
    # ---- 历史存档(供网页趋势/消耗分析) ----
    history = []
    if os.path.exists(HISTORY_CSV):
        try:
            with open(HISTORY_CSV, encoding="utf-8") as f:
                for row in csv.DictReader(f):
                    # 14 列规范: date,vessel,code,lane,pic,lsfo,hsfo,mgo,ulsfo,bw,fw,refeer,found,report_time
                    if is_offline(row.get("vessel", "")):
                        continue          # 已下线船的历史不再展示(CSV 归档行保留, 不删)
                    oils = (_num(row.get("lsfo")), _num(row.get("hsfo")),
                            _num(row.get("mgo")), _num(row.get("ulsfo")))
                    if int(row.get("found") or 0) and all(o is None for o in oils):
                        continue          # 脏行: 标记为抓到但四个油种全空(写入时被清零)
                    history.append({
                        "t": (row.get("report_time") or row.get("date") or "")[:16],
                        "v": row.get("vessel", ""),
                        "c": row.get("code", ""),
                        "l": row.get("lane", ""),
                        "ls": oils[0], "hs": oils[1], "mg": oils[2], "us": oils[3],
                        "bw": _num(row.get("bw")),
                        "fw": _num(row.get("fw")),
                        "f": int(row.get("found") or 0),
                        "rt": (row.get("report_time") or "")[:19],    # 船长报告接收时间
                    })
        except Exception as e:
            print("[WARN] read history csv failed:", e)
    # ---- 加油事件(可选, 由用户后续提供; 当前为空则按 ROB 增幅自动识别) ----
    # 文件格式: {"船名": {"YYYY-MM-DD": 加油量MT, ...}, ...}
    bunkering = {}
    BUNKER_JSON = os.path.join(ROB_DIR, "bunkering.json")
    if os.path.exists(BUNKER_JSON):
        try:
            with open(BUNKER_JSON, encoding="utf-8") as f:
                bunkering = json.load(f)
        except Exception as e:
            print("[WARN] read bunkering.json failed:", e)
    payload = {"updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
               "vessels": vessels,
               "history": history,
               "bunkering": bunkering}
    enc = cryptojs_encrypt(json.dumps(payload, ensure_ascii=False), PASSWORD)
    # Python 端自校验(确保 JS 端能解开)
    back = cryptojs_decrypt(enc, PASSWORD)
    assert json.loads(back)["updated"] == payload["updated"]
    assert "history" in json.loads(back)
    html = load_template().replace("__ENC__", enc)
    with open(OUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print("HTML -> %s (%d vessels, history=%d rows, encrypted OK, %d bytes)"
          % (OUT_HTML, len(vessels), len(history), len(html)))
    print("HTML -> %s (%d vessels, encrypted OK, %d bytes)" % (OUT_HTML, len(vessels), len(html)))


# ---------------------------------------------------------------- 5. xlsx (可选, 本机)
def build_xlsx(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception as e:
        print("[WARN] openpyxl missing, skip xlsx:", e)
        return
    snap = datetime.now().strftime("%Y.%-m.%-d") if os.name != "nt" else datetime.now().strftime("%Y.%#m.%#d")
    ordered = sorted(results, key=lambda r: (r.get("lane", ""), r.get("code", "")))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = snap
    ws.cell(1, 1, snap).font = Font(bold=True, size=14)
    ws.cell(1, 2, "盘油记录（ROB 取自各船船长 Noon/Berth/Sailing Report）").font = Font(bold=True, size=11)
    headers = ["序号", "Vessel Name全称", "Vessel Code", "Lane Code", "燃油负责人", "PIC",
               "ROB LSFO", "ROB HSFO", "ROB ULSFO", "ROB MGO", "订油状态", "订油情况",
               "REMARK", "特殊", "拟采购日期", "ROB报告时间"]
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    def fmt_num(x):
        if x is None:
            return None
        if isinstance(x, float):
            return int(x) if x == int(x) else round(x, 2)
        return x

    r = 3
    for i, rec in enumerate(ordered, 1):
        rowvals = [i, rec.get("vessel"), rec.get("code"), rec.get("lane"), "", rec.get("pic"),
                   fmt_num(rec.get("rob_lsfo")), fmt_num(rec.get("rob_hsfo")),
                   fmt_num(rec.get("rob_ulsfo")), fmt_num(rec.get("rob_mgo")),
                   "", "", "", "", "",
                   (rec.get("report_time") or "")[:19]]
        if not rec.get("found"):
            rowvals[12] = "邮箱未找到船长存油报告"
        for c, val in enumerate(rowvals, 1):
            cell = ws.cell(r, c, val)
            cell.border = border
            if c in (1, 2, 3, 4, 7, 8, 9, 10, 16):
                cell.alignment = Alignment(horizontal="center")
        r += 1
    for col, w in zip("ABCDEFGHIJKLMNO", [6, 22, 16, 12, 12, 16, 11, 11, 11, 10, 18, 28, 8, 14, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    try:
        wb.save(OUT_XLSX)
        print("XLSX -> %s" % OUT_XLSX)
    except PermissionError:
        alt = OUT_XLSX.replace(".xlsx", "_new.xlsx")
        wb.save(alt)
        print("[WARN] xlsx locked, saved %s" % alt)
    # 本机 CULINES 同步(仅目录存在时)
    if os.path.isdir(CULINES_DIR):
        import shutil
        try:
            dst = os.path.join(CULINES_DIR, "盘油记录.auto.xlsx")
            shutil.copy2(OUT_XLSX, dst)
            print("Synced -> %s" % dst)
        except Exception as e:
            print("[WARN] CULINES sync failed:", e)


# ---------------------------------------------------------------- 6. 每日历史存档(累加)
def write_daily_history(results):
    """累加式每日历史, 两份产物:
       - rob_data/history/rob_YYYY-MM-DD.json : 当天完整快照(按运行覆盖当天)
       - rob_data/history/rob_history.csv      : 逐船逐次快照, 追加(不覆盖历史,
         可随时按 snapshot_time / vessel 筛选任意历史日)
    """
    os.makedirs(HISTORY_DIR, exist_ok=True)
    now = datetime.now()
    snap_time = now.strftime("%Y-%m-%d %H:%M")
    date_tag = now.strftime("%Y-%m-%d")
    # 1) 当日 json 快照(覆盖当天, 不同日期是不同文件 => 自然累积)
    day_file = os.path.join(HISTORY_DIR, "rob_%s.json" % date_tag)
    try:
        with open(day_file, "w", encoding="utf-8") as f:
            json.dump({"snapshot_time": snap_time, "vessels": results},
                      f, ensure_ascii=False, indent=2)
    except Exception as e:
        print("[WARN] write daily json failed:", e)
    # 2) 累加 csv(幂等, 绝不覆盖历史行; 14 列规范与另一台机器一致):
    #    - 有报告时间的行按 (vessel, report_time) 去重 —— 同一份 Noon 报告无论由
    #      每日运行还是回补(--backfill)、哪台机器写入, 都只留一行;
    #    - MISS 行(report_time 为空)按 (date, vessel) 记录, 每次运行留痕,
    #      供网页 Data Integrity 面板展示缺报历史。
    #    date 列统一取"报告日期"(report_time[:10]), 而非运行日, 避免批量回补/双机
    #    运行日差异导致整月历史塌缩到同一天。
    existing_rt, existing_miss = set(), set()
    if os.path.exists(HISTORY_CSV):
        try:
            with open(HISTORY_CSV, encoding="utf-8", newline="") as f:
                for row in csv.reader(f):
                    if len(row) < 14:
                        continue
                    rt = (row[13] or "")[:19]
                    if rt:
                        existing_rt.add((row[1], rt))
                    else:
                        existing_miss.add((row[0], row[1]))
        except Exception:
            pass
    write_header = not os.path.exists(HISTORY_CSV)
    new_rows = []
    for r in results:
        rt19 = (r.get("report_time") or "")[:19]
        vessel = r.get("vessel", "")
        if rt19:
            if (vessel, rt19) in existing_rt:
                continue
            existing_rt.add((vessel, rt19))
            d = rt19[:10]
        else:
            d = snap_time[:10]
            if (d, vessel) in existing_miss:
                continue
            existing_miss.add((d, vessel))
        new_rows.append([
            d, vessel, r.get("code", ""), r.get("lane", ""),
            r.get("pic", ""),
            r.get("rob_lsfo"), r.get("rob_hsfo"), r.get("rob_mgo"),
            r.get("rob_ulsfo"), r.get("rob_bw"), r.get("rob_fw"),
            r.get("rob_refeer", ""),
            int(bool(r.get("found"))), rt19,
        ])
    try:
        with open(HISTORY_CSV, "a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(SNAP_FIELDS)
            for row in new_rows:
                w.writerow(row)
        if new_rows:
            print("history -> %s (+%d rows, day=%s)" % (HISTORY_CSV, len(new_rows), day_file))
        else:
            print("history -> %s (no new rows this run, cumulative preserved)" % HISTORY_CSV)
    except Exception as e:
        print("[WARN] write history csv failed:", e)


def append_history_rows(recs, fleet_lookup=None):
    """把回补解析到的报告追加进累加 CSV(14 列规范, 与另一台机器一致)。
    去重键 (vessel, report_time): 同一封报告无论由每日运行/回补/哪台机器写入都只留一行。
    一天多份报告(Noon/Berth/Sailing)折叠为"每天每船一条": 取当天最接近 Noon(12:00)
    的那条, 与另一台每日一行模型一致, 也正好是趋势锚点想要的每天同一时刻 ROB。
    date 列 = 报告日期(report_time[:10])。"""
    fleet_lookup = fleet_lookup or {}
    existing = set()
    if os.path.exists(HISTORY_CSV):
        try:
            for row in csv.reader(open(HISTORY_CSV, encoding="utf-8")):
                if len(row) >= 14 and row[13]:
                    existing.add((row[1], row[13][:19]))
        except Exception:
            pass
    best = {}   # (date, vessel) -> (score, rec, rt19, date)
    for r in recs:
        rt19 = (r.get("report_time") or "")[:19]
        if not rt19:
            continue
        if (r["vessel"], rt19) in existing:
            continue
        d = rt19[:10]
        key = (d, r["vessel"])
        score = abs(_hours_of(rt19) - 12.0)
        cur = best.get(key)
        if cur is None or score < cur[0]:
            best[key] = (score, r, rt19, d)
    rows = []
    for (d, vessel), (score, r, rt19, dd) in best.items():
        fl = fleet_lookup.get(norm(r["vessel"]), {})
        rows.append([
            dd, vessel, fl.get("code", r.get("code", "")), fl.get("lane", r.get("lane", "")),
            "",  # pic: 回补不抓船长名, 留空(另一台机器的行会带)
            r.get("rob_lsfo"), r.get("rob_hsfo"), r.get("rob_mgo"),
            r.get("rob_ulsfo"), r.get("rob_bw"), r.get("rob_fw"),
            "",  # refeer: 回补不抓, 留空
            1, rt19,
        ])
    try:
        write_header = not os.path.exists(HISTORY_CSV)
        with open(HISTORY_CSV, "a", encoding="utf-8", newline="") as f:
            w = csv.writer(f)
            if write_header:
                w.writerow(SNAP_FIELDS)
            for row in rows:
                w.writerow(row)
    except Exception as e:
        print("[WARN] append history failed:", e)
    return len(rows)


def backfill_history(days, sender_map, fleet_lookup=None):
    """回补过去 days 天的每一份 ROB 报告(Noon/Berth/Sailing), 按船长邮箱逆向映射识别船。
    返回 (recs 已解析, unknown 未知发件人列表)。未知发件人需交用户确认后固化进 vessel_senders.json。"""
    import win32com.client
    store = connect_outlook()
    if store is None:
        print("[WARN] no CULINES store, backfill aborted"); return [], []
    inbox = store.GetDefaultFolder(6)
    cache = build_folder_cache(inbox)
    rev = {}
    for v, e in sender_map.items():
        if e:
            rev.setdefault(e.lower(), []).append(v)
    cutoff = datetime.now().astimezone() - timedelta(days=days)
    cutoff_naive = cutoff.replace(tzinfo=None)
    filt = cutoff_naive.strftime("%m/%d/%Y %H:%M %p")
    folders = [inbox]
    seen_folders = {id(inbox)}
    def _walk(f):
        try:
            for c in f.Folders:
                if id(c) in seen_folders:
                    continue
                seen_folders.add(id(c))
                folders.append(c)
                _walk(c)
        except Exception:
            pass
    try:
        _walk(inbox)
    except Exception:
        pass
    unknown, recs, seen = [], [], set()
    for f in folders:
        try:
            items = f.Items.Restrict("[ReceivedTime] >= '%s'" % filt)
        except Exception:
            try:
                items = f.Items
            except Exception:
                continue
        try:
            items.Sort("[ReceivedTime]", True)
        except Exception:
            pass
        in_vessel_tree = (f is not inbox)
        for it in items:
            try:
                rt = it.ReceivedTime
            except Exception:
                continue
            if rt < cutoff:
                continue
            att = pick_report_attachment(it, strict=True)
            if att is None:
                continue
            rob = extract_rob(att)
            if not any(k in rob for k in OIL_KEYS):
                continue
            # 这是一份 ROB 报告, 识别船(按发件人邮箱优先, 其次 Vessel 子文件夹)
            se = get_sender(it)
            vessel = None
            cands = rev.get(se.lower(), []) if se else []
            folder_v = norm(f.Name) if in_vessel_tree else None
            if cands:
                if len(cands) == 1:
                    vessel = cands[0]
                else:
                    try:
                        subj = norm(it.Subject or "")
                    except Exception:
                        subj = ""
                    hit = [v for v in cands if v in subj]
                    vessel = hit[0] if hit else cands[0]
            elif folder_v and folder_v in cache:
                vessel = f.Name
            if not vessel:
                unknown.append({"sender": se, "subject": (it.Subject or "")[:90],
                                "received": rt.strftime("%Y-%m-%d %H:%M")})
                continue
            rep_t = rt.strftime("%Y-%m-%d %H:%M:%S")
            key = (rep_t, vessel)
            if key in seen:
                continue
            seen.add(key)
            recs.append({
                "vessel": vessel, "report_time": rep_t, "sender": se,
                "rob_lsfo": rob.get("LSFO"), "rob_hsfo": rob.get("HSFO"),
                "rob_mgo": rob.get("MGO"), "rob_ulsfo": rob.get("ULSFO"),
                "rob_bw": rob.get("BW"), "rob_fw": rob.get("FW"),
            })
    return recs, unknown


# ---------------------------------------------------------------- main
def backfill_mode(days):
    sender_map = load_sender_map()
    fleet = load_fleet()
    fleet_lookup = {norm(v["vessel"]): v for v in fleet}
    # norm -> 显示名(fleet + vessel.csv), 回补按发件人映射归船拿到的是 norm 键,
    # 写入历史前统一转显示名, 避免与每日运行的大写船名在趋势页分裂成两条船。
    disp = {norm(v["vessel"]): v["vessel"] for v in fleet}
    try:
        with open(os.path.join(BASE, "vessel.csv"), encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if len(row) >= 2 and row[0].strip():
                    nv = norm(row[0])
                    disp.setdefault(nv, row[0].strip())
                    if nv not in fleet_lookup:
                        fleet_lookup[nv] = {"vessel": row[0].strip(),
                                            "code": row[1].strip(), "lane": ""}
    except Exception:
        pass
    recs, unknown = backfill_history(days, sender_map, fleet_lookup)
    for r in recs:
        r["vessel"] = disp.get(r["vessel"], r["vessel"].upper())
    added = append_history_rows(recs, fleet_lookup)
    print("backfill: parsed %d reports, appended %d new rows (past %d days)"
          % (len(recs), added, days))
    results = []
    if os.path.exists(RESULTS):
        results = json.load(open(RESULTS, encoding="utf-8"))
    build_html(results)
    build_xlsx(results)
    if unknown:
        print("\n=== UNKNOWN SENDERS (%d) — ROB reports from unmapped captains ==="
              % len(unknown))
        for u in unknown[:300]:
            print("  SENDER=%s | %s | %s" % (u["sender"], u["received"], u["subject"]))
    print("DONE")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-outlook", action="store_true", help="不抓 Outlook, 只重建网页")
    ap.add_argument("--vessel", default=None, help="只刷新指定船(名称子串)")
    ap.add_argument("--backfill", type=int, default=0,
                    help="回补过去 N 天的每份 ROB 报告(按船长邮箱识别船), 追加进历史 CSV")
    args = ap.parse_args()
    if args.backfill:
        backfill_mode(args.backfill)
        return

    os.makedirs(ROB_DIR, exist_ok=True)
    fleet = load_fleet()

    results = []
    if os.path.exists(RESULTS):
        results = json.load(open(RESULTS, encoding="utf-8"))
    old_by_vessel = {}
    for r in results:
        old_by_vessel[norm(r.get("vessel", ""))] = r

    merged = []
    for v in fleet:
        old = old_by_vessel.get(norm(v["vessel"]), {})
        merged.append({
            "vessel": v["vessel"],
            "code": v["code"] or old.get("code", ""),
            "lane": v["lane"] or old.get("lane", ""),
            "pic": v["pic"] or old.get("pic", ""),
            "rob_lsfo": old.get("rob_lsfo"), "rob_hsfo": old.get("rob_hsfo"),
            "rob_mgo": old.get("rob_mgo"), "rob_ulsfo": old.get("rob_ulsfo"),
            "rob_bw": old.get("rob_bw"), "rob_fw": old.get("rob_fw"),
            "rob_refeer": old.get("rob_refeer"),
            "report_time": old.get("report_time"), "source": old.get("source"),
            "sender": old.get("sender"), "found": old.get("found", False),
        })

    if not args.no_outlook:
        sender_map = load_sender_map()
        store = connect_outlook()
        if store is None:
            print("[WARN] CULINES Outlook store not found, keep old data")
        else:
            inbox = store.GetDefaultFolder(6)
            cache = build_folder_cache(inbox)
            # 全量文件夹列表(含收件箱顶层船文件夹 + 嵌套 + 同名不去重), 供 sender 索引和主题兜底
            folder_list = build_folder_list(inbox)
            sender_index = build_sender_index([inbox] + folder_list, sender_map)
            # 全船队 norm(船名 + 船代码): 用于识别"共用文件夹"(MEDKON 里同时有
            # MEDKON DON / MEDKON LIA), 命中这类文件夹时必须按主题认船防串数据
            fleet_norms = set()
            for v in merged:
                fleet_norms.add(norm(v["vessel"]))
                if v.get("code"):
                    fleet_norms.add(norm(v["code"]))
            print("Outlook store OK, Vessel folders: %d, all folders: %d, "
                  "sender map: %d, sender index: %d"
                  % (len(cache), len(folder_list), len(sender_map), len(sender_index)))
            n_new = 0
            targets = merged
            if args.vessel:
                targets = [r for r in merged if args.vessel.upper() in r["vessel"].upper()]
            for i, rec in enumerate(targets, 1):
                got = refresh_vessel(inbox, cache, rec, sender_map,
                                     sender_index, folder_list, fleet_norms)
                n_new += 1 if got else 0
                mark = "NEW" if got else ("keep" if rec.get("found") else "MISS")
                print("[%2d/%2d] %-24s %-8s %-5s LSFO=%-8s MGO=%-8s t=%s"
                      % (i, len(targets), rec["vessel"], rec["code"], mark,
                         rec.get("rob_lsfo"), rec.get("rob_mgo"),
                         (rec.get("report_time") or "")[:16]))
                # 自动学习: 成功抓到且有真实发件人, 回写固化映射
                if got:
                    s = rec.get("sender")
                    if s and norm(rec["vessel"]) not in sender_map:
                        sender_map[norm(rec["vessel"])] = s
                        print("   + learned sender for %s: %s" % (rec["vessel"], s))
            # 兜底: 常规检索后仍偏旧/没抓到的船, 再做一次全文件夹深度扫描
            stale = [r for r in targets if is_stale(r)]
            if stale:
                print("stale vessels (older than %dh or MISS): %d -> %s"
                      % (STALE_HOURS, len(stale),
                         ", ".join(r["vessel"] for r in stale)))
                deep_refresh_stale(inbox, folder_list, stale, sender_map)
            save_sender_map(sender_map)
            print("refreshed this run: %d/%d" % (n_new, len(targets)))

    json.dump(merged, open(RESULTS, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    found = sum(1 for r in merged if r.get("found"))
    print("found ROB: %d/%d -> %s" % (found, len(merged), RESULTS))

    build_html(merged)
    build_xlsx(merged)
    write_daily_history(merged)
    print("DONE")


if __name__ == "__main__":
    main()
