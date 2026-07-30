#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gen_pic.py —— 汇总/维护当前船队的 PIC 对照表(P盘大Excel同级)
- 船代码: 从 GitHub vessel.csv 取(权威)。
- PIC:    若 P盘 PIC汇总.xlsx(人工编辑主文件)已存在, 保留其中手工维护的PIC(不覆盖);
          仅缺失的船回退 旧大Excel 块头C16。.csv 为同步镜像。
输出 PIC汇总.xlsx 与 PIC汇总.csv。
"""
import os, glob, csv, argparse
import openpyxl
from build_fleet_movement import (norm, FOLDER_ALIAS, ROUTE_OVERRIDE, ROUTE_ALIAS,
                                  canon_route, load_old_ref, load_vessel_csv, load_pic,
                                  DEFAULT_SRC, DEFAULT_OLD, VESSEL_CSV, DEFAULT_PIC, UPD_DIR)

# 输出与源数据同目录(随 UPD_DIR 自动切换 P:/Z:), 两机通用。
OUT_XLSX = os.path.join(UPD_DIR, "PIC汇总.xlsx")
OUT_CSV  = os.path.join(UPD_DIR, "PIC汇总.csv")

def latest_xlsx(folder):
    fs = [f for f in glob.glob(os.path.join(folder, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not fs: return None
    fs.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return fs[0]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--old", default=DEFAULT_OLD)
    ap.add_argument("--vessel", default=VESSEL_CSV)
    ap.add_argument("--pic", default=DEFAULT_PIC, help="已有PIC表(保留手工PIC)")
    args = ap.parse_args()
    old_ref, _ = load_old_ref(args.old)
    vessel = load_vessel_csv(args.vessel)
    existing_pic = load_pic(args.pic)   # 从 xlsx 主文件保留手工维护的PIC
    src = args.src
    folders = sorted([d for d in os.listdir(src)
                      if os.path.isdir(os.path.join(src, d)) and d != "已下线船舶"])
    rows = []
    for fol in folders:
        p = latest_xlsx(os.path.join(src, fol))
        if not p: continue
        wb = openpyxl.load_workbook(p, data_only=True)
        ws = wb[wb.sheetnames[0]]
        src_route = ws.cell(1, 1).value
        route = canon_route(fol, src_route)
        key = norm(FOLDER_ALIAS.get(fol, fol))
        ref = old_ref.get(key, {})
        # 船代码: vessel.csv(权威) -> 旧大ExcelC9 -> 源R1C9
        code = vessel.get(norm(fol)) or vessel.get(key) or ref.get("code") or ws.cell(1, 9).value
        # PIC: 已有PIC表(手工)优先; 缺失才回退旧大Excel
        pic = existing_pic.get(norm(fol))
        if pic is None:
            pic = str(ref.get("pic") or "").replace("PIC:", "").replace("PIC :", "").strip()
        disp = ref.get("display") or fol
        status = "已有" if pic else "⚠ 缺失待补"
        rows.append((route, disp, fol, code or "", pic, status))

    # 按航线、船名排序
    rows.sort(key=lambda x: (norm(x[0]), norm(x[1])))

    # 写 xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PIC汇总"
    hdr = ["航线", "船名(显示)", "文件夹名", "船代码", "PIC", "状态"]
    fill_h = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
    bold = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(1, c, h); cell.font = bold; cell.fill = fill_h
        cell.border = openpyxl.styles.Border(*[openpyxl.styles.Side(style="thin", color="BFBFBF")]*4)
    miss_fill = openpyxl.styles.PatternFill("solid", fgColor="FFF2CC")
    for i, (route, disp, fol, code, pic, status) in enumerate(rows, 2):
        vals = [route, disp, fol, code, pic, status]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.border = openpyxl.styles.Border(*[openpyxl.styles.Side(style="thin", color="BFBFBF")]*4)
            if status.startswith("⚠"):
                cell.fill = miss_fill
    widths = [10, 24, 22, 10, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    try:
        wb.save(OUT_XLSX)
    except PermissionError:
        # 文件可能被 Excel 打开而锁定 -> 落到 .new.xlsx 避免覆盖失败
        alt = OUT_XLSX[:-5] + ".new.xlsx"
        wb.save(alt)
        print(f"  [WARN] {OUT_XLSX} 被占用(可能Excel打开), 已写入 {alt}")

    # 写 csv
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(hdr)
        for r in rows:
            w.writerow(r)

    total = len(rows)
    miss = sum(1 for r in rows if r[5].startswith("⚠"))
    print(f"当前船队: {total} 艘 | 有PIC: {total-miss} | 缺失待补: {miss}")
    print(f"  -> {OUT_XLSX}")
    print(f"  -> {OUT_CSV}")
    if miss:
        print("缺失PIC的船:")
        for r in rows:
            if r[5].startswith("⚠"):
                print(f"   {r[0]:<6} {r[1]}")

if __name__ == "__main__":
    main()
