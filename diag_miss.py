#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""诊断：对比大Excel各块(PORT,VOY)键 vs 源lookup键，定位0命中/低命中的原因。"""
import os, sys, glob, argparse
import openpyxl

DEFAULT_SRC = r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\2026"
DEFAULT_TEMPLATE = r"Z:\04 上海操作中心\01 船期管理科\船期管理\VSL Daily Movement\更新\CUL DAILY MOVEMENT.xlsx"
NAME_ALIASES = {"HONG DA XIN 728": "HDX 728", "DONG FANG MIN HAI": "DONG FANG MING HAI"}

def norm(s): return (s or "").strip().upper()
def norm_voy(s):
    if s is None: return ""
    return str(s).strip().upper().replace(" ", "")

def latest_xlsx(folder):
    files = [f for f in glob.glob(os.path.join(folder, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not files: return None
    files.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return files[0]

def read_source_keys(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    code = ws.cell(1, 9).value
    header_row = None
    for r in range(1, min(ws.max_row, 80) + 1):
        if norm(ws.cell(r, 1).value) == "PORT":
            header_row = r; break
    keys = set()
    if header_row is None:
        return code, keys
    for r in range(header_row + 1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None: continue
        s1 = norm(c1)
        if s1 == "PORT": continue
        if s1 == "" or s1.startswith("REMARK"): continue
        keys.add((s1, norm_voy(ws.cell(r, 7).value)))
    return code, keys

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--only", default="", help="只诊断指定船名(逗号分隔)，如 ASR,NEW THINKER")
    args = ap.parse_args()

    src_dir = args.src
    template = args.template
    only = [norm(x) for x in args.only.split(",") if x.strip()] + \
           [norm(NAME_ALIASES[x]) for x in args.only.split(",") if x.strip() in NAME_ALIASES]

    # 源索引
    index = {}
    for fol in sorted([d for d in os.listdir(src_dir) if os.path.isdir(os.path.join(src_dir, d)) and d != "已下线船舶"]):
        folder = os.path.join(src_dir, fol)
        p = latest_xlsx(folder)
        if not p: continue
        code, keys = read_source_keys(p)
        index[fol] = (code, keys, p)

    wb_ro = openpyxl.load_workbook(template, data_only=True)
    ws = wb_ro[wb_ro.sheetnames[0]]
    total = ws.max_row

    def is_block_header(ws, r):
        if r >= ws.max_row: return False
        if norm(ws.cell(r + 1, 1).value) != "PORT": return False
        c4 = ws.cell(r, 4).value
        if not c4 or "VESSEL" in norm(c4): return False
        return True

    def match_folder(vessel, code):
        bv = norm(vessel); bc = norm(code)
        cands = {bv}
        if bv in NAME_ALIASES: cands.add(norm(NAME_ALIASES[bv]))
        for fol, (fc, fk, fp) in index.items():
            if norm(fol) in cands: return fol, fk, fp
        for fol, (fc, fk, fp) in index.items():
            if bc and norm(fc) == bc: return fol, fk, fp
        return None, None, None

    r = 1
    while r <= total:
        if is_block_header(ws, r):
            vessel = str(ws.cell(r, 4).value).strip()
            code = ws.cell(r, 9).value
            fol, src_keys, src_path = match_folder(vessel, code)
            # 收集块内键
            blk_keys = set()
            rr = r + 2
            while rr <= total:
                sc1 = norm(ws.cell(rr, 1).value)
                if sc1 == "PORT" or sc1 == "#VALUE!" or (ws.cell(rr, 4).value and "VESSEL" in norm(ws.cell(rr, 4).value)) or sc1.startswith("PIC") or is_block_header(ws, rr):
                    break
                if sc1 == "" or sc1.startswith("REMARK"):
                    rr += 1; continue
                blk_keys.add((sc1, norm_voy(ws.cell(rr, 7).value)))
                rr += 1
            r = rr
            if only and norm(vessel) not in only:
                continue
            inter = blk_keys & src_keys
            print(f"\n=== {vessel} (code={code}) -> folder={fol} ===")
            print(f"  块内键数={len(blk_keys)} 源键数={len(src_keys)} 交集={len(inter)}")
            if only:
                print(f"  源文件: {os.path.basename(src_path)}")
                print(f"  块内独有(源无): {sorted(blk_keys - src_keys)}")
                print(f"  源独有(块内无): {sorted(src_keys - blk_keys)[:20]}")
            continue
        r += 1

if __name__ == "__main__":
    main()
